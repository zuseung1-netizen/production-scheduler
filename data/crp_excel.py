"""
CRP Excel integration.
CRP data (total headcount per shift per date) lives in Excel.
The scheduler distributes this total HC to rooms/processes automatically.

Sheet: "CRP"
Row 1: Headers — ShiftNo | <date1> | <date2> | ...
Row 2+: shift_no | total_hc | total_hc | ...
"""
import os
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from data.repositories import ConfigRepo

CRP_SHEET = "CRP"
HOLD_SHEET = "HOLD"


class CRPManager:
    """
    Loads CRP data from Excel and caches it in memory.
    Stores total headcount per (date, shift_no).
    The scheduler is responsible for distributing HC to rooms/processes.
    """

    def __init__(self):
        self._total_hc: Dict[Tuple[str, int], int] = {}
        # key: (date_str YYYY-MM-DD, shift_no) -> total headcount for that shift
        self._holds: Dict[Tuple[str, str, int], bool] = {}
        # key: (date_str, room_code, shift_no) -> True if on hold
        self._loaded = False
        self._path: Optional[str] = None

    def _excel_path(self) -> Optional[str]:
        return ConfigRepo.get("crp_excel_path") or None

    def refresh(self) -> Tuple[bool, str]:
        """Reload from Excel. Returns (success, message)."""
        if not HAS_OPENPYXL:
            return False, "openpyxl not installed"
        path = self._excel_path()
        if not path or not os.path.exists(path):
            return False, f"CRP Excel not found: {path}"
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            self._total_hc.clear()
            self._holds.clear()

            if CRP_SHEET in wb.sheetnames:
                ws = wb[CRP_SHEET]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return False, "CRP sheet is empty"
                header = rows[0]

                # Auto-detect format:
                # New format: col0="Date", col1+="Shift N"  → rows=dates, cols=shifts
                # Old format: col0="ShiftNo", col1+=dates   → rows=shifts, cols=dates
                first_cell = str(header[0]).strip().lower() if header[0] is not None else ""
                new_format = first_cell == "date"

                if new_format:
                    # col 1+ are shift numbers
                    shift_cols = []
                    for i, h in enumerate(header):
                        if i >= 1 and h is not None:
                            try:
                                shift_cols.append((i, int(str(h).replace("Shift", "").strip())))
                            except (TypeError, ValueError):
                                pass
                    for row in rows[1:]:
                        if not row[0]:
                            continue
                        d = row[0]
                        if isinstance(d, (datetime, date)):
                            date_str = d.strftime("%Y-%m-%d")
                        else:
                            date_str = str(d)[:10]
                        for col_idx, shift in shift_cols:
                            val = row[col_idx]
                            try:
                                hc = int(val) if val is not None else 0
                            except (TypeError, ValueError):
                                hc = 0
                            self._total_hc[(date_str, shift)] = hc
                else:
                    # Old format: col0=ShiftNo, col1+=dates
                    date_cols = []
                    for i, h in enumerate(header):
                        if i >= 1 and h is not None:
                            if isinstance(h, (datetime, date)):
                                date_cols.append((i, h.strftime("%Y-%m-%d")))
                            else:
                                try:
                                    date_cols.append((i, str(h)[:10]))
                                except Exception:
                                    pass
                    for row in rows[1:]:
                        if not row[0]:
                            continue
                        try:
                            shift = int(row[0])
                        except (TypeError, ValueError):
                            continue
                        for col_idx, date_str in date_cols:
                            val = row[col_idx]
                            try:
                                hc = int(val) if val is not None else 0
                            except (TypeError, ValueError):
                                hc = 0
                            self._total_hc[(date_str, shift)] = hc

            if HOLD_SHEET in wb.sheetnames:
                ws_h = wb[HOLD_SHEET]
                for row in ws_h.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    d = row[0]
                    if isinstance(d, (datetime, date)):
                        d = d.strftime("%Y-%m-%d")
                    else:
                        d = str(d)[:10]
                    room = str(row[1]).strip() if row[1] else ""
                    try:
                        shift = int(row[2])
                    except (TypeError, ValueError):
                        continue
                    self._holds[(d, room, shift)] = True

            self._loaded = True
            self._path = path
            return True, f"Loaded {len(self._total_hc)} CRP entries from {os.path.basename(path)}"

        except Exception as e:
            return False, f"Failed to load CRP: {e}"

    def get_total_hc(self, date_str: str, shift_no: int) -> int:
        """Return total headcount for this date/shift (sum across all rooms)."""
        if not self._loaded:
            self.refresh()
        return self._total_hc.get((date_str, shift_no), 0)

    def is_held(self, date_str: str, room_code: str, shift_no: int) -> bool:
        if not self._loaded:
            self.refresh()
        return self._holds.get((date_str, room_code, shift_no), False)

    def get_all(self) -> Dict:
        """Returns {(date_str, shift_no): total_hc}."""
        if not self._loaded:
            self.refresh()
        return dict(self._total_hc)

    def get_date_range(self) -> Tuple[Optional[str], Optional[str]]:
        if not self._total_hc:
            return None, None
        dates = [k[0] for k in self._total_hc.keys()]
        return min(dates), max(dates)

    def create_template(self, save_path: str, shifts: List[int],
                        date_from: str, date_to: str) -> Tuple[bool, str]:
        """Generate a blank CRP Excel template (Date × Shift)."""
        if not HAS_OPENPYXL:
            return False, "openpyxl not installed"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = CRP_SHEET

            d0 = datetime.strptime(date_from, "%Y-%m-%d").date()
            d1 = datetime.strptime(date_to, "%Y-%m-%d").date()
            dates = []
            cur = d0
            while cur <= d1:
                dates.append(cur)
                cur += timedelta(days=1)

            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # New format: row0=headers (Date | Shift1 | Shift2 | ...), rows=dates
            headers = ["Date"] + [f"Shift {s}" for s in shifts]
            ws.column_dimensions["A"].width = 14
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if ci >= 2:
                    ws.column_dimensions[cell.column_letter].width = 12

            for ri, d in enumerate(dates, 2):
                ws.cell(row=ri, column=1, value=d.strftime("%Y-%m-%d"))
                for ci in range(2, len(headers) + 1):
                    ws.cell(row=ri, column=ci, value=0).border = border

            # HOLD sheet
            ws_h = wb.create_sheet(HOLD_SHEET)
            for ci, h in enumerate(["Date", "RoomCode", "ShiftNo"], 1):
                cell = ws_h.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font

            wb.save(save_path)
            return True, f"Template saved to {save_path}"
        except Exception as e:
            return False, str(e)

    def write_total_hc(self, updates: Dict[Tuple[str, int], int]) -> Tuple[bool, str]:
        """Write total HC values per (date_str, shift_no) to CRP Excel.
        updates: {(date_str, shift_no): total_hc_int}
        Supports both old (row=shift, col=date) and new (row=date, col=shift) formats.
        """
        if not HAS_OPENPYXL:
            return False, "openpyxl not installed"
        path = self._excel_path()
        if not path or not os.path.exists(path):
            return False, f"CRP Excel not found: {path}"
        try:
            wb = openpyxl.load_workbook(path)
            if CRP_SHEET not in wb.sheetnames:
                return False, f"Sheet '{CRP_SHEET}' not found in {path}"
            ws = wb[CRP_SHEET]
            rows_data = list(ws.iter_rows(values_only=False))
            if not rows_data:
                return False, "CRP sheet is empty"

            header_row = rows_data[0]
            first_cell = str(header_row[0].value).strip().lower() if header_row[0].value is not None else ""
            new_format = first_cell == "date"

            written = 0
            if new_format:
                # New format: row0=header (Date|Shift1|...), rows=dates
                shift_col: dict = {}
                for ci, cell in enumerate(header_row):
                    v = cell.value
                    if ci >= 1 and v is not None:
                        try:
                            shift_col[int(str(v).replace("Shift", "").strip())] = ci
                        except (TypeError, ValueError):
                            pass

                date_row: dict = {}
                for ri, row in enumerate(rows_data[1:], start=1):
                    v = row[0].value
                    if v is None:
                        continue
                    if isinstance(v, (datetime, date)):
                        date_row[v.strftime("%Y-%m-%d")] = ri
                    else:
                        date_row[str(v)[:10]] = ri

                for (d_str, shift), hc in updates.items():
                    row_i = date_row.get(d_str)
                    col_i = shift_col.get(shift)
                    if row_i is not None and col_i is not None:
                        ws.cell(row=row_i + 1, column=col_i + 1, value=hc)
                        self._total_hc[(d_str, shift)] = hc
                        written += 1
            else:
                # Old format: col0=ShiftNo, col1+=dates
                date_col: dict = {}
                for ci, cell in enumerate(header_row):
                    v = cell.value
                    if ci >= 1 and v is not None:
                        if isinstance(v, (datetime, date)):
                            date_col[v.strftime("%Y-%m-%d")] = ci
                        else:
                            try:
                                date_col[str(v)[:10]] = ci
                            except Exception:
                                pass

                shift_row: dict = {}
                for ri, row in enumerate(rows_data[1:], start=1):
                    try:
                        shift_row[int(row[0].value)] = ri
                    except (TypeError, ValueError):
                        continue

                for (d_str, shift), hc in updates.items():
                    col_i = date_col.get(d_str)
                    row_i = shift_row.get(shift)
                    if col_i is not None and row_i is not None:
                        ws.cell(row=row_i + 1, column=col_i + 1, value=hc)
                        self._total_hc[(d_str, shift)] = hc
                        written += 1

            wb.save(path)
            return True, f"Updated {written} HC values in {os.path.basename(path)}"
        except Exception as e:
            return False, f"Failed to write CRP: {e}"


# Singleton
crp_manager = CRPManager()
