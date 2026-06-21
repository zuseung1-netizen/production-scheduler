"""
Excel import / export utilities.
Provides:
  - Template download for SO / SKU / Room master
  - SO upload with change detection
  - Full data export
"""
import json
import os
from datetime import datetime
from typing import Dict, List, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from data.repositories import SKURepo, RoomRepo, SORepo, PlanRepo, ActualRepo


# ─── helpers ─────────────────────────────────────────────────────────────────

def _header_style():
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(color="FFFFFF", bold=True)
    return fill, font

def _write_header(ws, headers: List[str]):
    fill, font = _header_style()
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")

def _col_widths(ws, widths: List[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


# ─── SO Template & Upload ─────────────────────────────────────────────────────

SO_HEADERS = ["SONumber", "SKUCode", "LineItem", "Qty", "DueDate",
              "Priority", "Status", "StartNoEarlier", "Note", "CustomerName",
              "CommittedDueDate"]
SO_COL_WIDTHS = [16, 14, 10, 8, 12, 10, 10, 16, 30, 20, 16]

def download_so_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook()
    ws = wb.active
    ws.title = "SalesOrders"
    _write_header(ws, SO_HEADERS)
    _col_widths(ws, SO_COL_WIDTHS)
    # Example row
    ws.append(["SO-001", "SKU-A", "L01", 100, "2025-09-30", 1, "OPEN", "", "Sample order", "Customer A", ""])
    wb.save(path)
    return True, path

def preview_so_upload(path: str) -> Tuple[bool, str, Dict]:
    """
    Parse SO Excel and diff against current DB — no writes.
    Returns (success, message, preview) where preview = {
        "rows": [{change_type, new, old, changed_fields}, ...],
        "summary": {new, modified, unchanged, closed},
        "errors": [...],
    }
    """
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed", {}
    if not os.path.exists(path):
        return False, f"File not found: {path}", {}

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["SalesOrders"] if "SalesOrders" in wb.sheetnames else wb.active
        excel_rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return False, str(e), {}

    from data.repositories import SORepo
    existing_map = {
        (s["so_number"], s["sku_code"], s["line_item"]): s
        for s in SORepo.all()
    }

    COMPARE = ("qty", "due_date", "committed_due_date", "priority", "status",
               "start_no_earlier", "note", "customer_name")

    preview_rows, uploaded_keys, errors = [], set(), []
    summary = {"new": 0, "modified": 0, "unchanged": 0, "closed": 0}

    for row in excel_rows:
        if not row[0]:
            continue
        try:
            nd = {
                "so_number":        str(row[0]).strip(),
                "sku_code":         str(row[1]).strip(),
                "line_item":        str(row[2]).strip(),
                "qty":              int(row[3]),
                "due_date":         _parse_date(row[4]),
                "priority":         int(row[5]) if row[5] not in (None, "") else None,
                "status":           str(row[6]).strip().upper() if row[6] else "OPEN",
                "start_no_earlier":   _parse_date(row[7]) if row[7] else None,
                "note":               str(row[8]).strip() if row[8] else None,
                "customer_name":      str(row[9]).strip() if len(row) > 9 and row[9] else None,
                "committed_due_date": _parse_date(row[10]) if len(row) > 10 and row[10] else None,
            }
            key = (nd["so_number"], nd["sku_code"], nd["line_item"])
            uploaded_keys.add(key)
            old = existing_map.get(key)

            if old is None:
                preview_rows.append({"change_type": "NEW",      "new": nd, "old": None,  "changed_fields": []})
                summary["new"] += 1
            else:
                changed = [f for f in COMPARE
                           if str(old.get(f) or "") != str(nd.get(f) or "")]
                if changed:
                    preview_rows.append({"change_type": "MODIFIED",  "new": nd, "old": old, "changed_fields": changed})
                    summary["modified"] += 1
                else:
                    preview_rows.append({"change_type": "UNCHANGED", "new": nd, "old": old, "changed_fields": []})
                    summary["unchanged"] += 1
        except Exception as e:
            errors.append(f"Row parse error: {e}")

    for key, so in existing_map.items():
        if key not in uploaded_keys and so["status"] != "CLOSED":
            preview_rows.append({"change_type": "CLOSED", "new": {**so, "status": "CLOSED"}, "old": so, "changed_fields": ["status"]})
            summary["closed"] += 1

    return True, "", {"rows": preview_rows, "summary": summary, "errors": errors}


def upload_so(path: str) -> Tuple[bool, str, Dict]:
    """
    Upload SO Excel. Detects new / modified / closed items.
    Returns (success, message, summary_dict).
    """
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed", {}
    if not os.path.exists(path):
        return False, f"File not found: {path}", {}

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["SalesOrders"] if "SalesOrders" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return False, str(e), {}

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Save snapshot before changes
    SORepo.save_snapshot(batch_id)

    existing_keys = {(so["so_number"], so["sku_code"], so["line_item"])
                     for so in SORepo.all()}
    uploaded_keys = set()
    summary = {"new": 0, "modified": 0, "unchanged": 0, "closed": 0, "errors": []}

    for row in rows:
        if not row[0]:
            continue
        try:
            data = {
                "so_number":       str(row[0]).strip(),
                "sku_code":        str(row[1]).strip(),
                "line_item":       str(row[2]).strip(),
                "qty":             int(row[3]),
                "due_date":        _parse_date(row[4]),
                "priority":        int(row[5]) if row[5] not in (None, "") else None,
                "status":          str(row[6]).strip().upper() if row[6] else "OPEN",
                "start_no_earlier": _parse_date(row[7]) if row[7] else None,
                "note":               str(row[8]).strip() if row[8] else None,
                "customer_name":      str(row[9]).strip() if len(row) > 9 and row[9] else None,
                "committed_due_date": _parse_date(row[10]) if len(row) > 10 and row[10] else None,
            }
            change = SORepo.upsert(data, batch_id=batch_id)
            summary[change.lower()] = summary.get(change.lower(), 0) + 1
            uploaded_keys.add((data["so_number"], data["sku_code"], data["line_item"]))
        except Exception as e:
            summary["errors"].append(str(e))

    # Mark keys present in DB but absent in upload as CLOSED
    for key in existing_keys - uploaded_keys:
        so = SORepo.get(*key)
        if so and so["status"] not in ("CLOSED",):
            SORepo.close(*key, batch_id=batch_id)
            summary["closed"] += 1

    msg = (f"Upload complete [batch {batch_id}]: "
           f"{summary['new']} new, {summary['modified']} modified, "
           f"{summary['unchanged']} unchanged, {summary['closed']} closed")
    if summary["errors"]:
        msg += f", {len(summary['errors'])} errors"
    return True, msg, summary


def _parse_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s


# ─── SKU Master Template & Upload ─────────────────────────────────────────────

SKU_HEADERS = ["SKUCode", "SKUName", "UoM", "PostLeadDays", "Note"]
SKU_COL_WIDTHS = [14, 30, 8, 14, 30]

def download_sku_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook()
    ws = wb.active
    ws.title = "SKUMaster"
    _write_header(ws, SKU_HEADERS)
    _col_widths(ws, SKU_COL_WIDTHS)
    ws.append(["SKU-A", "Sample Product A", 30, 0, ""])
    wb.save(path)
    return True, path

def upload_sku(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["SKUMaster"] if "SKUMaster" in wb.sheetnames else wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            SKURepo.upsert({
                "sku_code": str(row[0]).strip(),
                "sku_name": str(row[1]).strip() if row[1] else "",
                "uom": int(row[2]) if row[2] else 1,
                "post_lead_days": int(row[3]) if row[3] else 0,
                "note": str(row[4]) if row[4] else None,
            })
            count += 1
        return True, f"Imported {count} SKUs"
    except Exception as e:
        return False, str(e)


# ─── Room Master Template & Upload ────────────────────────────────────────────

ROOM_HEADERS = ["RoomCode", "ProcessName", "ProcessType", "RoomType", "UPPH", "UPH_Fixed",
                "HC_Min", "HC_Max", "HC_Fixed", "Changeover_Shifts", "Note"]
ROOM_COL_WIDTHS = [12, 20, 12, 12, 10, 10, 8, 8, 10, 14, 30]

def download_room_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook()
    ws = wb.active
    ws.title = "RoomMaster"
    _write_header(ws, ROOM_HEADERS)
    _col_widths(ws, ROOM_COL_WIDTHS)
    ws.append(["ROOM-A", "PROCESS-1", "MANUAL", "TYPE-A", 50, "", 2, 6, "", 1, ""])
    ws.append(["ROOM-A", "PROCESS-2", "AUTO",   "TYPE-A", "", 200, "", "", 1, 0, ""])
    wb.save(path)
    return True, path

def upload_room(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["RoomMaster"] if "RoomMaster" in wb.sheetnames else wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        has_co = "Changeover_Shifts" in headers
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if has_co:
                # 11-column format (new)
                co = int(row[9]) if row[9] is not None else 0
                note = str(row[10]) if row[10] else None
            else:
                # 10-column legacy format — default changeover to 0
                co = 0
                note = str(row[9]) if row[9] else None
            RoomRepo.upsert({
                "room_code":         str(row[0]).strip(),
                "process_name":      str(row[1]).strip(),
                "process_type":      str(row[2]).strip().upper(),
                "room_type":         str(row[3]).strip() if row[3] else "TYPE-A",
                "upph":              float(row[4]) if row[4] else None,
                "uph_fixed":         float(row[5]) if row[5] else None,
                "hc_min":            int(row[6]) if row[6] else None,
                "hc_max":            int(row[7]) if row[7] else None,
                "hc_fixed":          int(row[8]) if row[8] else None,
                "changeover_shifts": co,
                "note":              note,
            })
            count += 1
        return True, f"Imported {count} room/process rows"
    except Exception as e:
        return False, str(e)


# ─── Full Export ──────────────────────────────────────────────────────────────

def export_all(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        wb = Workbook()

        # SO sheet
        ws_so = wb.active
        ws_so.title = "SalesOrders"
        _write_header(ws_so, SO_HEADERS + ["ReceivedAt"])
        for so in SORepo.all():
            ws_so.append([so["so_number"], so["sku_code"], so["line_item"],
                           so["qty"], so["due_date"], so["priority"],
                           so["status"], so["start_no_earlier"], so["note"],
                           so["received_at"]])

        # SKU sheet
        ws_sku = wb.create_sheet("SKUMaster")
        _write_header(ws_sku, SKU_HEADERS)
        for sku in SKURepo.all():
            ws_sku.append([sku["sku_code"], sku["sku_name"], sku["uom"],
                           sku["post_lead_days"], sku["note"]])

        # Room sheet
        ws_room = wb.create_sheet("RoomMaster")
        _write_header(ws_room, ROOM_HEADERS)
        for r in RoomRepo.all():
            ws_room.append([r["room_code"], r["process_name"], r["process_type"],
                            r["upph"], r["uph_fixed"], r["hc_min"],
                            r["hc_max"], r["hc_fixed"], r["note"]])

        # Production Plan sheet
        ws_plan = wb.create_sheet("ProductionPlan")
        plan_headers = ["PlanID", "SONumber", "SKUCode", "LineItem", "ProcessName",
                        "RoomCode", "PlanDate", "ShiftNo", "QtyPlanned",
                        "QtyProduced", "IsLocked", "Memo"]
        _write_header(ws_plan, plan_headers)
        for p in PlanRepo.all():
            ws_plan.append([p["plan_id"], p["so_number"], p["sku_code"],
                            p["line_item"], p["process_name"], p["room_code"],
                            p["plan_date"], p["shift_no"], p["qty_planned"],
                            p["qty_produced"], p["is_locked"], p["memo"]])

        wb.save(path)
        return True, f"Exported to {path}"
    except Exception as e:
        return False, str(e)


# ─── Gantt Plan Normalized Export ────────────────────────────────────────────

def export_gantt_plan(plans: List[Dict], sos_dict: Dict,
                      mat_groups: Dict, path: str) -> Tuple[bool, str]:
    """
    Export current Gantt view plans as normalized flat tables.

    sos_dict  : {(so_no, sku, line_item): so_row}  (from GanttCanvas._sos)
    mat_groups: {group_id: [members]}               (from GanttCanvas._mat_groups)
    """
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        wb = Workbook()

        # ── Sheet 1: Plan Detail ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "PlanDetail"
        hdr_fill, hdr_font = _header_style()

        COLS = [
            ("PlanID",          10),
            ("Type",             9),
            ("Code",            14),
            ("SONumber",        14),
            ("LineItem",        10),
            ("CustomerName",    18),
            ("ProcessSeq",      12),
            ("ProcessName",     16),
            ("RoomCode",        12),
            ("PlanDate",        13),
            ("Shift",            7),
            ("QtyPlanned",      12),
            ("QtyProduced",     13),
            ("IsLocked",        10),
            ("IsFinalStep",     12),
            ("DueDate",         13),
            ("DaysUntilDue",    14),
            ("ConsolGroup",     14),
            ("Memo",            30),
        ]
        headers = [c[0] for c in COLS]
        widths  = [c[1] for c in COLS]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w
        ws.freeze_panes = "A2"

        today = datetime.today().date()

        # Status fills
        fill_locked   = PatternFill("solid", fgColor="FFF3CD")
        fill_final    = PatternFill("solid", fgColor="D4EDDA")
        fill_material = PatternFill("solid", fgColor="EDE0FF")

        for plan in sorted(plans,
                           key=lambda p: (p["plan_date"], p["sku_code"],
                                          p.get("process_seq", 1))):
            is_mat = plan.get("entity_type") == "MATERIAL"
            so_key = (plan["so_number"], plan["sku_code"], plan["line_item"])
            so     = sos_dict.get(so_key, {})

            due_date_str = so.get("due_date", "") if not is_mat else ""
            if due_date_str:
                try:
                    due = datetime.strptime(due_date_str, "%Y-%m-%d").date()
                    days_until = (due - today).days
                except ValueError:
                    days_until = ""
            else:
                days_until = ""

            # For MATERIAL plans, collect linked SOs from demand group
            if is_mat:
                gid     = plan.get("material_group_id", "")
                members = mat_groups.get(gid, [])
                so_nos  = ", ".join({m["so_number"] for m in members})
                customer = ""
            else:
                so_nos   = plan["so_number"]
                customer = so.get("customer_name", "") or ""

            row = [
                plan["plan_id"],
                plan.get("entity_type", "SKU"),
                plan.get("entity_code") or plan["sku_code"],
                so_nos,
                plan["line_item"],
                customer,
                plan.get("process_seq", 1),
                plan["process_name"],
                plan["room_code"],
                plan["plan_date"],
                plan["shift_no"],
                plan["qty_planned"],
                plan.get("qty_produced", 0),
                "Y" if plan["is_locked"] else "",
                "Y" if plan.get("is_final_seq") else "",
                due_date_str,
                days_until,
                plan.get("consolidation_group", "") or "",
                plan.get("memo", "") or "",
            ]
            ws.append(row)
            ri = ws.max_row

            # Row highlight
            row_fill = None
            if is_mat:
                row_fill = fill_material
            elif plan["is_locked"]:
                row_fill = fill_locked
            elif plan.get("is_final_seq"):
                row_fill = fill_final

            if row_fill:
                for ci in range(1, len(COLS) + 1):
                    ws.cell(row=ri, column=ci).fill = row_fill

            # Days until due: color
            if isinstance(days_until, int):
                cell = ws.cell(row=ri, column=headers.index("DaysUntilDue") + 1)
                if days_until < 0:
                    cell.font = Font(color="CC0000", bold=True)
                elif days_until <= 3:
                    cell.font = Font(color="996600")

        # Auto-filter
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLS)).coordinate}"

        # ── Sheet 2: Material Demand (demand group members) ───────────────────
        if mat_groups:
            ws2 = wb.create_sheet("MaterialDemand")
            hdr2 = ["GroupID", "MaterialCode", "SONumber", "SKUCode",
                    "LineItem", "QtyRequired", "DueDate"]
            for ci, h in enumerate(hdr2, 1):
                cell = ws2.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions["A"].width = 14
            ws2.column_dimensions["B"].width = 16
            ws2.column_dimensions["C"].width = 14
            ws2.column_dimensions["D"].width = 14
            ws2.column_dimensions["E"].width = 10
            ws2.column_dimensions["F"].width = 12
            ws2.column_dimensions["G"].width = 13
            ws2.freeze_panes = "A2"

            for gid, members in sorted(mat_groups.items()):
                for m in members:
                    ws2.append([gid, m["material_code"], m["so_number"],
                                m["sku_code"], m["line_item"],
                                m["qty_required"], m["due_date"]])

        wb.save(path)
        n = len(plans)
        return True, f"Exported {n} plan rows → {os.path.basename(path)}"
    except Exception as e:
        return False, str(e)


# ─── SKU Process Routing Template & Upload ────────────────────────────────────

SKU_PROC_HEADERS = ["SKUCode", "ProcessSeq", "ProcessName", "AllowedRoomTypes", "IsFinalSeq", "Note"]
SKU_PROC_COL_WIDTHS = [14, 12, 20, 25, 12, 30]


def download_sku_process_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook()
    ws = wb.active
    ws.title = "SKUProcess"
    _write_header(ws, SKU_PROC_HEADERS)
    _col_widths(ws, SKU_PROC_COL_WIDTHS)
    ws.append(["SKU-A", 1, "FILL-STD",  "TYPE-A,TYPE-B", 0, "First process"])
    ws.append(["SKU-A", 2, "SEAL-AUTO", "TYPE-A",        0, "Second process"])
    ws.append(["SKU-A", 3, "PACK-STD",  "TYPE-A",        1, "Final process — MRP trigger"])
    # Add a note about IsFinalSeq
    ws["F1"].comment = None
    ws.append([])
    ws.append(["# IsFinalSeq: 1 = final step (MRP), 0 = intermediate step"])
    ws.append(["# AllowedRoomTypes: comma-separated, must match RoomType in Room Master"])
    wb.save(path)
    return True, path


def upload_sku_process(path: str) -> Tuple[bool, str]:
    """
    Upload SKU process routing from Excel.
    Replaces all routing for each SKU found in the file.
    """
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        from data.repositories import SKUProcessRepo
        wb = load_workbook(path, data_only=True)
        ws = wb["SKUProcess"] if "SKUProcess" in wb.sheetnames else wb.active
        count = 0
        seen_skus = set()
        rows_to_insert = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or str(row[0]).startswith("#"):
                continue
            sku_code     = str(row[0]).strip()
            process_seq  = int(row[1]) if row[1] else None
            process_name = str(row[2]).strip() if row[2] else None
            allowed      = str(row[3]).strip() if row[3] else None
            is_final     = int(row[4]) if row[4] is not None else 0
            note         = str(row[5]) if row[5] else None

            if not all([sku_code, process_seq, process_name, allowed]):
                continue

            seen_skus.add(sku_code)
            rows_to_insert.append({
                "sku_code":           sku_code,
                "process_seq":        process_seq,
                "process_name":       process_name,
                "allowed_room_types": allowed,
                "is_final_seq":       is_final,
                "note":               note,
            })
            count += 1

        # Delete existing routing for affected SKUs before re-inserting
        for sku in seen_skus:
            SKUProcessRepo.delete_all_for_sku(sku)
        SKUProcessRepo.bulk_upsert(rows_to_insert)

        # Validate all affected SKUs
        errors = []
        for sku in seen_skus:
            ok, msg = SKUProcessRepo.validate_routing(sku)
            if not ok:
                errors.append(msg)

        result_msg = f"Imported {count} process steps for {len(seen_skus)} SKUs"
        if errors:
            result_msg += f"\nValidation warnings:\n" + "\n".join(errors)
        return True, result_msg
    except Exception as e:
        return False, str(e)


# ─── Material Master Template & Upload ───────────────────────────────────────

MAT_HEADERS    = ["MaterialCode", "MaterialName", "UoM", "PostLeadDays", "Note"]
MAT_COL_WIDTHS = [16, 30, 8, 14, 30]


def download_material_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook(); ws = wb.active; ws.title = "MaterialMaster"
    _write_header(ws, MAT_HEADERS); _col_widths(ws, MAT_COL_WIDTHS)
    ws.append(["MAT-001", "Semi-finished A", 1, 1, "QC lead 1 day"])
    wb.save(path)
    return True, path


def upload_material(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        from data.repositories import MaterialRepo
        wb = load_workbook(path, data_only=True)
        ws = wb["MaterialMaster"] if "MaterialMaster" in wb.sheetnames else wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            MaterialRepo.upsert({
                "material_code":  str(row[0]).strip(),
                "material_name":  str(row[1]).strip() if row[1] else "",
                "uom":            int(row[2]) if row[2] else 1,
                "post_lead_days": int(row[3]) if row[3] else 0,
                "note":           str(row[4]) if row[4] else None,
            })
            count += 1
        return True, f"Imported {count} materials"
    except Exception as e:
        return False, str(e)


# ─── Process Routing Template & Upload ───────────────────────────────────────

PR_HEADERS    = ["EntityType", "EntityCode", "ProcessSeq", "ProcessName",
                 "AllowedRoomTypes", "RequiresMaterialCode", "IsFinalSeq", "Note",
                 "MinGapShifts"]
PR_COL_WIDTHS = [12, 16, 10, 20, 22, 20, 12, 30, 14]


def download_process_routing_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook(); ws = wb.active; ws.title = "ProcessRouting"
    _write_header(ws, PR_HEADERS); _col_widths(ws, PR_COL_WIDTHS)
    ws.append(["SKU",      "SKU-A",   1, "FILL-STD",  "TYPE-A",      "MAT-001", 0, "Step 1 needs MAT-001", 0])
    ws.append(["SKU",      "SKU-A",   2, "PACK-STD",  "TYPE-A",      "",        1, "Final step",            0])
    ws.append(["MATERIAL", "MAT-001", 1, "SEMI-PROC", "TYPE-B",      "",        1, "Material final step",   0])
    ws.append([])
    ws.append(["# EntityType: SKU or MATERIAL"])
    ws.append(["# IsFinalSeq: 1 = final (MRP trigger), 0 = intermediate"])
    ws.append(["# RequiresMaterialCode: leave blank if not needed"])
    ws.append(["# MinGapShifts: empty shifts required between previous step end and this step start"])
    ws.append(["#   0 = adjacent shift OK  |  1 = 1 shift gap  |  2 = 2 shifts gap (≈1 day in 2-shift)"])
    wb.save(path)
    return True, path


def upload_process_routing(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        from data.repositories import ProcessRoutingRepo
        wb = load_workbook(path, data_only=True)
        ws = (wb["ProcessRouting"] if "ProcessRouting" in wb.sheetnames
              else wb.active)
        rows_to_insert, seen_entities, count = [], set(), 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or str(row[0]).startswith("#"): continue
            et   = str(row[0]).strip().upper()
            code = str(row[1]).strip()
            seq  = int(row[2]) if row[2] else None
            proc = str(row[3]).strip() if row[3] else None
            alwd = str(row[4]).strip() if row[4] else None
            rmat = str(row[5]).strip() if row[5] else None
            fin  = int(row[6]) if row[6] is not None else 0
            note = str(row[7]) if row[7] else None
            gap  = int(row[8]) if len(row) > 8 and row[8] is not None else 0
            if not all([et, code, seq, proc, alwd]): continue
            seen_entities.add((et, code))
            rows_to_insert.append({
                "entity_type": et, "entity_code": code,
                "process_seq": seq, "process_name": proc,
                "allowed_room_types": alwd,
                "requires_material_code": rmat or None,
                "is_final_seq": fin, "note": note,
                "min_gap_shifts": gap,
            })
            count += 1

        for et, code in seen_entities:
            ProcessRoutingRepo.delete_all_for_entity(et, code)
        ProcessRoutingRepo.bulk_upsert(rows_to_insert)

        errors = []
        for et, code in seen_entities:
            v, m = ProcessRoutingRepo.validate(et, code)
            if not v: errors.append(m)

        msg = f"Imported {count} routing steps for {len(seen_entities)} entities"
        if errors: msg += "\nWarnings:\n" + "\n".join(errors)
        return True, msg
    except Exception as e:
        return False, str(e)


# ─── Inventory Template & Upload ─────────────────────────────────────────────

INV_HEADERS    = ["SKUCode", "LotNumber", "QtyAvailable",
                  "ProductionDate", "ExpiryDate", "Note"]
INV_COL_WIDTHS = [14, 16, 14, 14, 14, 30]


def download_inventory_template(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    wb = Workbook(); ws = wb.active; ws.title = "Inventory"
    _write_header(ws, INV_HEADERS); _col_widths(ws, INV_COL_WIDTHS)
    ws.append(["SKU-A30", "LOT-2026-001", 200,
               "2026-05-01", "2027-05-01", "Excess from SO-099"])
    ws.append(["SKU-A30", "LOT-2026-002", 150,
               "2026-06-01", "2027-06-01", "Cancelled order stock"])
    ws.append([])
    ws.append(["# QtyAvailable: 가용 수량 (QC 완료, 납품 가능 수량)"])
    ws.append(["# ExpiryDate: FEFO 정렬 기준. 공백이면 후순위 배정"])
    wb.save(path)
    return True, path


def upload_inventory(path: str) -> Tuple[bool, str]:
    if not HAS_OPENPYXL:
        return False, "openpyxl not installed"
    try:
        from data.repositories import InventoryRepo
        wb = load_workbook(path, data_only=True)
        ws = wb["Inventory"] if "Inventory" in wb.sheetnames else wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or str(row[0]).startswith("#"):
                continue
            InventoryRepo.upsert({
                "sku_code":        str(row[0]).strip(),
                "lot_number":      str(row[1]).strip(),
                "qty_available":   int(row[2]) if row[2] else 0,
                "production_date": _parse_date(row[3]) if row[3] else None,
                "expiry_date":     _parse_date(row[4]) if row[4] else None,
                "note":            str(row[5]) if row[5] else None,
            })
            count += 1
        return True, f"Imported {count} inventory lots"
    except Exception as e:
        return False, str(e)
