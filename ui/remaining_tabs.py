"""
CRP Tab, Actuals Entry Tab, Alerts Tab, Dashboard Tab
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from typing import Dict, List

import math

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QAbstractItemView, QHeaderView,
    QFileDialog, QMessageBox, QScrollArea, QGroupBox, QComboBox,
    QSpinBox, QLineEdit, QDateEdit, QFormLayout, QDialog,
    QDialogButtonBox, QTextEdit, QSplitter, QTabWidget, QCheckBox,
    QRadioButton, QStackedWidget, QFrame, QListWidget, QListWidgetItem
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor, QBrush, QFont, QPen

from data.repositories import (
    PlanRepo, SORepo, SKURepo, ShiftRepo, RoomRepo, ActualRepo, ConfigRepo,
    ProcessRoutingRepo, MaterialRepo, InventoryRepo, CompanyHolidayRepo
)
from data.crp_excel import crp_manager
from core.scheduler import scheduler
from utils.korean_holidays import is_holiday, holiday_name


# ════════════════════════════════════════════════════════════════════════════
#  GAP BAR CHART
# ════════════════════════════════════════════════════════════════════════════

class GapBarChart(QWidget):
    """Grouped vertical bar chart drawn with QPainter. Scrollable horizontally."""
    BAR_W      = 14
    BAR_GAP    = 3
    GROUP_GAP  = 16
    PAD_LEFT   = 44
    PAD_BOTTOM = 28
    LEGEND_H   = 22

    def __init__(self, parent=None):
        super().__init__(parent)
        self._data: list   = []   # [(group_label, [val, ...]), ...]
        self._series: list = []
        self._colors: list = []
        self._max_val = 1.0
        self._holiday_lbls: set = set()
        self.setMinimumHeight(160)

    def set_data(self, data, series_labels, colors, holiday_labels: set | None = None):
        self._data         = data
        self._series       = series_labels
        self._colors       = [QColor(c) for c in colors]
        self._holiday_lbls = holiday_labels or set()
        all_vals = [v for _, vals in data for v in vals]
        self._max_val = max((v for v in all_vals if v > 0), default=1.0)
        n  = len(series_labels)
        gw = n * self.BAR_W + max(0, n - 1) * self.BAR_GAP + self.GROUP_GAP
        self.setMinimumWidth(max(300, self.PAD_LEFT + len(data) * gw + 20))
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QPen
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        chart_h  = H - self.LEGEND_H - self.PAD_BOTTOM - 4
        bottom_y = self.LEGEND_H + chart_h
        if not self._data or chart_h < 20:
            p.end(); return
        n = len(self._series)
        gw = n * self.BAR_W + max(0, n - 1) * self.BAR_GAP + self.GROUP_GAP

        # Legend
        lx = self.PAD_LEFT
        from PyQt6.QtGui import QFont
        p.setFont(QFont("Segoe UI", 8))
        for lbl, col in zip(self._series, self._colors):
            p.fillRect(lx, 4, 10, 10, col)
            p.setPen(QPen(QColor("#374151")))
            p.drawText(lx + 13, 13, lbl)
            lx += 13 + len(lbl) * 6 + 12

        # Y gridlines
        for frac in (0.25, 0.5, 0.75, 1.0):
            y = int(bottom_y - frac * chart_h)
            p.setPen(QPen(QColor("#e5e7eb"), 1))
            p.drawLine(self.PAD_LEFT, y, W - 4, y)
            p.setPen(QPen(QColor("#6b7280")))
            p.setFont(QFont("Segoe UI", 7))
            p.drawText(2, y + 4, f"{self._max_val * frac:.0f}")

        # Bars
        x = self.PAD_LEFT
        p.setFont(QFont("Segoe UI", 7))
        for lbl, vals in self._data:
            for si, (v, col) in enumerate(zip(vals, self._colors)):
                bx = x + si * (self.BAR_W + self.BAR_GAP)
                bh = max(1, int((v / self._max_val) * chart_h)) if self._max_val > 0 and v > 0 else 0
                p.fillRect(bx, bottom_y - bh, self.BAR_W, bh, col)
            cx = x + (n * (self.BAR_W + self.BAR_GAP)) // 2 - len(lbl) * 3
            lbl_col = QColor("#c62828") if lbl in self._holiday_lbls else QColor("#374151")
            p.setPen(QPen(lbl_col))
            p.drawText(cx, bottom_y + 14, lbl)
            x += gw
        p.end()


# ════════════════════════════════════════════════════════════════════════════
#  LINE CHART
# ════════════════════════════════════════════════════════════════════════════

class LineChart(QWidget):
    """Multi-series percentage line chart (0-100%) drawn with QPainter."""
    PAD_LEFT   = 48
    PAD_BOTTOM = 28
    PAD_TOP    = 8
    LEGEND_H   = 22

    def __init__(self, parent=None):
        super().__init__(parent)
        self._series: list   = []
        self._x_labels: list = []
        self._holiday_lbls: set = set()
        self.setMinimumHeight(180)

    def set_data(self, x_labels: list, series: list,
                 holiday_labels=None):
        """
        x_labels: ['W22 06/01', ...] or ['06/01', ...]
        series: [{'label': str, 'color': str, 'values': [float, ...]}]
        values aligned with x_labels (0-100 range, % values)
        """
        self._x_labels = x_labels
        self._series = series
        self._holiday_lbls = holiday_labels or set()
        n = len(x_labels)
        self.setMinimumWidth(max(300, self.PAD_LEFT + n * 40 + 20))
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QPen, QFont, QFontMetrics
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        chart_h = H - self.LEGEND_H - self.PAD_BOTTOM - self.PAD_TOP
        chart_w = W - self.PAD_LEFT - 8
        bottom_y = self.LEGEND_H + self.PAD_TOP + chart_h
        left_x = self.PAD_LEFT

        if not self._x_labels or chart_h < 20:
            p.end(); return

        n = len(self._x_labels)

        # Legend
        lx = left_x
        p.setFont(QFont("Segoe UI", 8))
        for s in self._series:
            col = QColor(s['color'])
            p.fillRect(lx, 4, 12, 3, col)
            p.setPen(QPen(QColor("#374151")))
            p.drawText(lx + 15, 13, s['label'])
            lx += 15 + len(s['label']) * 6 + 14

        # Y gridlines + labels (0%, 25%, 50%, 75%, 100%)
        for frac in (0.0, 0.25, 0.5, 0.75, 1.0):
            y = int(bottom_y - frac * chart_h)
            p.setPen(QPen(QColor("#e5e7eb"), 1))
            p.drawLine(left_x, y, W - 8, y)
            p.setPen(QPen(QColor("#6b7280")))
            p.setFont(QFont("Segoe UI", 7))
            p.drawText(2, y + 4, f"{int(frac*100)}%")

        # X labels + vertical guidelines
        if n > 1:
            step_w = chart_w / (n - 1)
        else:
            step_w = chart_w

        p.setFont(QFont("Segoe UI", 7))
        for i, lbl in enumerate(self._x_labels):
            x = int(left_x + i * step_w) if n > 1 else left_x + chart_w // 2
            lbl_col = QColor("#c62828") if lbl in self._holiday_lbls else QColor("#374151")
            p.setPen(QPen(lbl_col))
            p.drawText(x - 14, bottom_y + 14, lbl)
            p.setPen(QPen(QColor("#f3f4f6"), 1))
            p.drawLine(x, self.LEGEND_H + self.PAD_TOP, x, bottom_y)

        # Lines + dots
        for s in self._series:
            vals = s.get('values', [])
            col = QColor(s['color'])
            p.setPen(QPen(col, 2))
            pts = []
            for i, v in enumerate(vals):
                x = int(left_x + i * step_w) if n > 1 else left_x + chart_w // 2
                y = int(bottom_y - (v / 100.0) * chart_h) if v is not None else None
                if y is not None:
                    pts.append((x, y))
            for j in range(len(pts) - 1):
                p.drawLine(pts[j][0], pts[j][1], pts[j+1][0], pts[j+1][1])
            p.setBrush(QBrush(col))
            p.setPen(Qt.PenStyle.NoPen)
            for x, y in pts:
                p.drawEllipse(x - 3, y - 3, 6, 6)

        p.end()


# ════════════════════════════════════════════════════════════════════════════
#  CAPACITY ANALYSIS WIDGET
# ════════════════════════════════════════════════════════════════════════════

class CapacityAnalysisWidget(QWidget):
    """HC shortage / Line Gap analysis + multi-scenario simulation."""
    _MAX_SCEN  = 3
    _SCEN_COLS = ["#2563eb", "#16a34a", "#d97706"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._scenarios: List[Dict] = []
        self._analysis:  Dict       = {}
        self._build_ui()

    # ── UI build ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(6)

        # Controls
        ctrl = QHBoxLayout()
        ctrl.addWidget(QLabel("From:"))
        self.d_from = QDateEdit(QDate.currentDate())
        self.d_from.setDisplayFormat("yyyy-MM-dd")
        ctrl.addWidget(self.d_from)
        ctrl.addWidget(QLabel("To:"))
        self.d_to = QDateEdit(QDate.currentDate().addDays(83))
        self.d_to.setDisplayFormat("yyyy-MM-dd")
        ctrl.addWidget(self.d_to)
        ctrl.addWidget(QLabel("View:"))
        self.gran = QComboBox()
        self.gran.addItems(["Weekly", "Daily"])
        ctrl.addWidget(self.gran)
        btn_analyze = QPushButton("▶ Analyze")
        btn_analyze.setStyleSheet(
            "background:#2563EB;color:white;font-weight:bold;"
            "border:none;border-radius:5px;padding:5px 14px;")
        btn_analyze.clicked.connect(self._analyze)
        ctrl.addWidget(btn_analyze)
        ctrl.addStretch()
        lay.addLayout(ctrl)

        # Scenario panel
        scen_grp = QGroupBox("HC Scenarios")
        scen_vlay = QVBoxLayout(scen_grp)
        btn_add = QPushButton("+ Add Scenario")
        btn_add.clicked.connect(self._add_scenario)
        scen_vlay.addWidget(btn_add, alignment=Qt.AlignmentFlag.AlignLeft)
        self._scen_row = QHBoxLayout()
        self._scen_row.setSpacing(8)
        self._scen_row.addStretch()
        scen_vlay.addLayout(self._scen_row)
        lay.addWidget(scen_grp)

        # Result tabs
        self._rtabs = QTabWidget()

        # HC Gap tab
        hc_w = QWidget(); hcl = QVBoxLayout(hc_w); hcl.setContentsMargins(4, 4, 4, 4)
        lbl_hc = QLabel("HC Utilization (%)")
        lbl_hc.setStyleSheet("font-size:11px;font-weight:bold;color:#374151;padding:2px 0;")
        hcl.addWidget(lbl_hc)
        self._hc_chart = LineChart()
        hc_sa = QScrollArea(); hc_sa.setWidget(self._hc_chart)
        hc_sa.setWidgetResizable(False); hc_sa.setFixedHeight(200)
        hc_sa.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        hcl.addWidget(hc_sa)
        self._hc_tbl = self._make_table()
        hcl.addWidget(self._hc_tbl, 1)
        self._rtabs.addTab(hc_w, "HC Gap")

        # Line Gap tab
        ln_w = QWidget(); lnl = QVBoxLayout(ln_w); lnl.setContentsMargins(4, 4, 4, 4)
        lbl_ln = QLabel("Line Utilization (%)")
        lbl_ln.setStyleSheet("font-size:11px;font-weight:bold;color:#374151;padding:2px 0;")
        lnl.addWidget(lbl_ln)
        self._ln_chart = LineChart()
        ln_sa = QScrollArea(); ln_sa.setWidget(self._ln_chart)
        ln_sa.setWidgetResizable(False); ln_sa.setFixedHeight(200)
        ln_sa.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        lnl.addWidget(ln_sa)
        self._ln_tbl = self._make_table()
        lnl.addWidget(self._ln_tbl, 1)
        self._rtabs.addTab(ln_w, "Line Gap")

        # Backlog Compare tab
        bl_w = QWidget(); bll = QVBoxLayout(bl_w); bll.setContentsMargins(4, 4, 4, 4)
        self._bl_note = QLabel(
            "Run '▶ Analyze' then '▶ Run Simulation' on each scenario to compare HC coverage.")
        self._bl_note.setStyleSheet("color:#6b7280;font-size:10px;padding:2px 4px;")
        self._bl_note.setWordWrap(True)
        bll.addWidget(self._bl_note)
        self._bl_tbl = self._make_table()
        bll.addWidget(self._bl_tbl, 1)
        self._rtabs.addTab(bl_w, "Backlog Compare")

        lay.addWidget(self._rtabs, 1)

    @staticmethod
    def _make_table() -> QTableWidget:
        t = QTableWidget()
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setAlternatingRowColors(True)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        t.verticalHeader().setVisible(False)
        return t

    # ── Scenario management ───────────────────────────────────────────────────

    def _add_scenario(self):
        if len(self._scenarios) >= self._MAX_SCEN:
            QMessageBox.information(self, "Limit", "Maximum 3 scenarios.")
            return
        idx = len(self._scenarios)
        self._scenarios.append({"name": f"Scenario {chr(65+idx)}", "periods": [], "result": None})
        self._rebuild_cards()

    def _rebuild_cards(self):
        while self._scen_row.count() > 1:
            item = self._scen_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        for i, s in enumerate(self._scenarios):
            self._scen_row.insertWidget(i, self._make_card(i, s))

    def _make_card(self, idx: int, scen: Dict) -> QGroupBox:
        col = self._SCEN_COLS[idx]
        card = QGroupBox(scen["name"])
        card.setFixedWidth(295)
        card.setStyleSheet(f"QGroupBox::title{{color:{col};font-weight:bold;}}")
        lay = QVBoxLayout(card)

        hdr = QHBoxLayout()
        btn_ren = QPushButton("Rename")
        btn_ren.clicked.connect(lambda _, i=idx: self._rename_scenario(i))
        btn_x = QPushButton("✕")
        btn_x.setFixedWidth(26)
        btn_x.setStyleSheet("background:#dc2626;color:white;border:none;border-radius:3px;")
        btn_x.clicked.connect(lambda _, i=idx: self._del_scenario(i))
        hdr.addWidget(btn_ren); hdr.addStretch(); hdr.addWidget(btn_x)
        lay.addLayout(hdr)

        tbl = QTableWidget(len(scen["periods"]), 4)
        tbl.setHorizontalHeaderLabels(["From", "To", "HC/Day", ""])
        tbl.verticalHeader().setVisible(False)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setFixedHeight(max(52, min(140, 28 + len(scen["periods"]) * 24)))
        for ri, (d0, d1, hc) in enumerate(scen["periods"]):
            for ci, txt in enumerate([d0, d1, f"{hc} ppl/day"]):
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                tbl.setItem(ri, ci, it)
            bx = QPushButton("✕")
            bx.setFixedSize(22, 20)
            bx.setStyleSheet("background:#dc2626;color:white;border:none;font-size:10px;")
            bx.clicked.connect(lambda _, i=idx, r=ri: self._del_period(i, r))
            tbl.setCellWidget(ri, 3, bx)
        lay.addWidget(tbl)

        btn_add_p = QPushButton("+ Add Period")
        btn_add_p.clicked.connect(lambda _, i=idx: self._add_period(i))
        lay.addWidget(btn_add_p)

        btn_run = QPushButton("▶ Run Simulation")
        btn_run.setStyleSheet(
            f"background:{col};color:white;font-weight:bold;"
            "border:none;border-radius:4px;padding:4px;")
        btn_run.clicked.connect(lambda _, i=idx: self._run_scenario(i))
        lay.addWidget(btn_run)
        return card

    def _add_period(self, sidx: int):
        dlg = QDialog(self)
        dlg.setWindowTitle("Add HC Period")
        form = QFormLayout(dlg)
        d0e = QDateEdit(QDate.currentDate()); d0e.setDisplayFormat("yyyy-MM-dd")
        d1e = QDateEdit(QDate.currentDate().addDays(30)); d1e.setDisplayFormat("yyyy-MM-dd")
        sp  = QSpinBox(); sp.setRange(0, 9999); sp.setValue(40); sp.setSuffix(" ppl/day")
        form.addRow("From:", d0e)
        form.addRow("To:",   d1e)
        form.addRow("HC/Day:", sp)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(dlg.accept); bb.rejected.connect(dlg.reject)
        form.addRow(bb)
        if dlg.exec():
            self._scenarios[sidx]["periods"].append((
                d0e.date().toString("yyyy-MM-dd"),
                d1e.date().toString("yyyy-MM-dd"),
                sp.value()))
            self._scenarios[sidx]["periods"].sort(key=lambda x: x[0])
            self._rebuild_cards()

    def _del_period(self, sidx: int, pidx: int):
        del self._scenarios[sidx]["periods"][pidx]
        self._rebuild_cards()

    def _rename_scenario(self, sidx: int):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(
            self, "Rename", "Name:", text=self._scenarios[sidx]["name"])
        if ok and name.strip():
            self._scenarios[sidx]["name"] = name.strip()
            self._rebuild_cards()

    def _del_scenario(self, sidx: int):
        del self._scenarios[sidx]
        self._rebuild_cards()
        self._refresh_backlog_table()

    # ── Data helpers ──────────────────────────────────────────────────────────

    def _date_list(self):
        from datetime import date as dt, timedelta
        d0 = self.d_from.date().toString("yyyy-MM-dd")
        d1 = self.d_to.date().toString("yyyy-MM-dd")
        cur, end = dt.fromisoformat(d0), dt.fromisoformat(d1)
        dates = []
        while cur <= end:
            dates.append(cur.isoformat()); cur += timedelta(days=1)
        return d0, d1, dates

    def _buckets(self, dates):
        """Returns [(label, [dates])]."""
        if self.gran.currentText() == "Daily":
            return [(d[5:], [d]) for d in dates]
        from datetime import date as dt
        result, seen = [], {}
        for d in dates:
            yr, wk, _ = dt.fromisoformat(d).isocalendar()
            key = (yr, wk)
            if key not in seen:
                seen[key] = len(result)
                result.append((f"W{wk:02d} {d[5:]}", []))
            result[seen[key]][1].append(d)
        return result

    def _holiday_labels_for_buckets(self, bkts) -> set:
        """Return set of bucket labels that contain a public or company holiday."""
        from datetime import date as dt
        company_hdays = CompanyHolidayRepo.date_set()
        hday_lbls = set()
        for lbl, ds in bkts:
            if any(is_holiday(dt.fromisoformat(d)) or d in company_hdays for d in ds):
                hday_lbls.add(lbl)
        return hday_lbls

    def _sum_shifts(self, hc_by_date_shift, dates):
        """Sum all shifts per day → {date: total}."""
        return {d: sum(hc_by_date_shift.get(d, {}).values()) for d in dates}

    def _scen_hc(self, scenario, dates):
        """{date: hc_per_day} from scenario periods."""
        out = {}
        for d in dates:
            for d0p, d1p, hc in scenario["periods"]:
                if d0p <= d <= d1p:
                    out[d] = hc; break
            else:
                out[d] = 0
        return out

    # ── Analysis ─────────────────────────────────────────────────────────────

    def _analyze(self):
        from core.scheduler import scheduler
        d0, d1, dates = self._date_list()
        self._analysis = {
            "d0": d0, "d1": d1, "dates": dates,
            "req_hc":   scheduler.compute_required_hc_from_plans(d0, d1),
            "avail_hc": scheduler.get_available_hc_by_date(d0, d1),
            "line":     scheduler.compute_line_utilization(d0, d1),
        }
        self._update_hc_tab()
        self._update_line_tab()
        self._refresh_backlog_table()

    def _update_hc_tab(self):
        a = self._analysis
        dates   = a["dates"]
        bkts    = self._buckets(dates)
        req_d   = self._sum_shifts(a["req_hc"],   dates)
        avl_d   = self._sum_shifts(a["avail_hc"], dates)
        scen_ds = [self._scen_hc(s, dates) for s in self._scenarios]
        snames  = [s["name"] for s in self._scenarios]

        hdrs = (["Period", "Req HC", "CRP HC", "Gap"] +
                snames + [f"{n} Gap" for n in snames])
        self._hc_tbl.setColumnCount(len(hdrs))
        self._hc_tbl.setHorizontalHeaderLabels(hdrs)
        self._hc_tbl.setRowCount(len(bkts))

        x_labels = []
        util_vals = []

        for ri, (lbl, bd) in enumerate(bkts):
            req  = sum(req_d.get(d, 0) for d in bd)
            avl  = sum(avl_d.get(d, 0) for d in bd)
            gap  = avl - req
            shcs = [sum(sd.get(d, 0) for d in bd) for sd in scen_ds]
            sgps = [sh - req for sh in shcs]

            def _it(txt, bg=None):
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if bg: it.setBackground(QBrush(QColor(bg)))
                return it

            self._hc_tbl.setItem(ri, 0, QTableWidgetItem(lbl))
            self._hc_tbl.setItem(ri, 1, _it(f"{req:.1f}", "#fee2e2"))
            self._hc_tbl.setItem(ri, 2, _it(f"{avl:.0f}"))
            self._hc_tbl.setItem(ri, 3, _it(f"{gap:+.1f}",
                                             "#dcfce7" if gap >= 0 else "#fee2e2"))
            for ci, (sh, sg) in enumerate(zip(shcs, sgps)):
                self._hc_tbl.setItem(ri, 4 + ci, _it(f"{sh:.0f}"))
                self._hc_tbl.setItem(ri, 4 + len(self._scenarios) + ci,
                                     _it(f"{sg:+.1f}", "#dcfce7" if sg >= 0 else "#fee2e2"))

            x_labels.append(lbl)
            util_pct = (req / avl * 100) if avl > 0 else 0.0
            util_vals.append(min(util_pct, 100.0))

        self._hc_tbl.resizeColumnsToContents()
        hday_lbls = self._holiday_labels_for_buckets(bkts)
        self._hc_chart.set_data(
            x_labels,
            [{'label': 'HC Util%', 'color': '#2563eb', 'values': util_vals}],
            holiday_labels=hday_lbls)

    def _update_line_tab(self):
        a    = self._analysis
        line = a["line"]
        bkts = self._buckets(a["dates"])

        # Get active line capacity (rooms that have plans) for Operating Util%
        try:
            from core.scheduler import scheduler as _sched
            active_cap = _sched.compute_active_line_capacity(a["d0"], a["d1"])
        except Exception:
            active_cap = {}

        hdrs = ["Period", "Avail Slots", "Used Slots", "Slot Gap",
                "Max Cap (units)", "Planned (units)", "Total Util%", "Operating Util%"]
        self._ln_tbl.setColumnCount(len(hdrs))
        self._ln_tbl.setHorizontalHeaderLabels(hdrs)
        self._ln_tbl.setRowCount(len(bkts))

        x_labels = []
        total_util_vals = []
        operating_util_vals = []

        for ri, (lbl, bd) in enumerate(bkts):
            avl  = sum(line["available"].get(d, 0) for d in bd)
            used = sum(line["used"].get(d, 0)      for d in bd)
            gap  = avl - used
            mx   = sum(line["max_cap"].get(d, 0)   for d in bd)
            pln  = sum(line["planned"].get(d, 0)   for d in bd)
            act_cap = sum(active_cap.get(d, 0)     for d in bd)

            total_util = (pln / mx * 100) if mx > 0 else 0.0
            op_util    = (pln / act_cap * 100) if act_cap > 0 else 0.0

            def _it(txt, bg=None):
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if bg: it.setBackground(QBrush(QColor(bg)))
                return it

            self._ln_tbl.setItem(ri, 0, QTableWidgetItem(lbl))
            self._ln_tbl.setItem(ri, 1, _it(str(avl)))
            self._ln_tbl.setItem(ri, 2, _it(str(used)))
            self._ln_tbl.setItem(ri, 3, _it(f"{gap:+d}",
                                             "#dcfce7" if gap >= 0 else "#fee2e2"))
            self._ln_tbl.setItem(ri, 4, _it(f"{mx:,.0f}"))
            self._ln_tbl.setItem(ri, 5, _it(f"{pln:,.0f}"))
            self._ln_tbl.setItem(ri, 6, _it(f"{total_util:.1f}%"))
            self._ln_tbl.setItem(ri, 7, _it(f"{op_util:.1f}%"))

            x_labels.append(lbl)
            total_util_vals.append(min(total_util, 100.0))
            operating_util_vals.append(min(op_util, 100.0))

        self._ln_tbl.resizeColumnsToContents()
        hday_lbls = self._holiday_labels_for_buckets(bkts)
        self._ln_chart.set_data(
            x_labels,
            [
                {'label': 'Total Line Util%',     'color': '#2563eb', 'values': total_util_vals},
                {'label': 'Operating Line Util%',  'color': '#16a34a', 'values': operating_util_vals},
            ],
            holiday_labels=hday_lbls)

    def _run_scenario(self, sidx: int):
        if not self._analysis:
            QMessageBox.information(
                self, "Analyze First", "Click '▶ Analyze' first.")
            return
        scen   = self._scenarios[sidx]
        a      = self._analysis
        dates  = a["dates"]
        req_d  = self._sum_shifts(a["req_hc"],   dates)
        avl_d  = self._sum_shifts(a["avail_hc"], dates)
        scen_d = self._scen_hc(scen, dates)
        bkts   = self._buckets(dates)

        bucket_results = []
        for lbl, bd in bkts:
            req     = sum(req_d.get(d, 0)   for d in bd)
            avl     = sum(avl_d.get(d, 0)   for d in bd)
            scen_hc = sum(scen_d.get(d, 0)  for d in bd)
            curr_cov = min(1.0, avl / req)    if req > 0 else 1.0
            scen_cov = min(1.0, scen_hc / req) if req > 0 else 1.0
            bucket_results.append((lbl, req, avl, scen_hc, curr_cov, scen_cov))

        scen["result"] = {"buckets": bucket_results}
        self._refresh_backlog_table()
        self._rtabs.setCurrentIndex(2)  # jump to Backlog Compare

    def _refresh_backlog_table(self):
        run = [(i, s) for i, s in enumerate(self._scenarios) if s.get("result")]
        if not run or not self._analysis:
            self._bl_tbl.setRowCount(0)
            self._bl_tbl.setColumnCount(0)
            return

        bkts = self._buckets(self._analysis["dates"])
        hdrs = (["Period", "Req HC", "CRP Cov%"] +
                [f"{self._scenarios[i]['name']} HC"   for i, _ in run] +
                [f"{self._scenarios[i]['name']} Cov%" for i, _ in run])
        self._bl_tbl.setColumnCount(len(hdrs))
        self._bl_tbl.setHorizontalHeaderLabels(hdrs)
        self._bl_tbl.setRowCount(len(bkts))

        n_run = len(run)
        for ri, (lbl, _) in enumerate(bkts):
            # Req HC + current coverage from first run scenario
            _, _, bkts0 = run[0][0], run[0][1], run[0][1]["result"]["buckets"]
            row0 = bkts0[ri] if ri < len(bkts0) else (lbl, 0, 0, 0, 1.0, 1.0)
            _, req, _, _, curr_cov, _ = row0

            def _it(txt, bg=None):
                it = QTableWidgetItem(txt)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if bg: it.setBackground(QBrush(QColor(bg)))
                return it

            self._bl_tbl.setItem(ri, 0, QTableWidgetItem(lbl))
            self._bl_tbl.setItem(ri, 1, _it(f"{req:.1f}"))
            self._bl_tbl.setItem(ri, 2, _it(
                f"{curr_cov*100:.0f}%",
                "#dcfce7" if curr_cov >= 1.0 else "#fee2e2"))

            for ci, (sidx, scen) in enumerate(run):
                bkts_s = scen["result"]["buckets"]
                row_s  = bkts_s[ri] if ri < len(bkts_s) else (lbl, 0, 0, 0, 1.0, 1.0)
                _, _, _, sh, _, cov = row_s
                self._bl_tbl.setItem(ri, 3 + ci, _it(f"{sh:.0f}"))
                cov_bg = ("#dcfce7" if cov >= 1.0 else
                          "#fff3cd" if cov >= 0.8 else "#fee2e2")
                self._bl_tbl.setItem(ri, 3 + n_run + ci,
                                     _it(f"{cov*100:.0f}%", cov_bg))

        self._bl_tbl.resizeColumnsToContents()
        self._bl_note.setText(
            "Coverage % = Scenario HC / Required HC.  "
            "✅ ≥100%: HC sufficient for planned demand.  "
            "⚠ Line constraints are shown separately in the 'Line Gap' tab.  "
            "For exact backlog, run 'Execute Plan' after adjusting CRP.")


# ════════════════════════════════════════════════════════════════════════════
#  CRP TAB
# ════════════════════════════════════════════════════════════════════════════

class CRPTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._build_ui()

    def _build_ui(self):
        self._crp_edit_mode    = False
        self._crp_changed_cells: set = set()
        self._crp_loading      = False
        self._crp_dates:  list = []
        self._crp_shifts: list = []

        # ── Outer: left sub-menu + right content stack ────────────────────────
        outer = QHBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Sub-menu panel
        sub_panel = QWidget()
        sub_panel.setFixedWidth(168)
        sub_panel.setStyleSheet(
            "QWidget { background:#f4f5f8; border-right:1px solid #dde3ed; }")
        sub_layout = QVBoxLayout(sub_panel)
        sub_layout.setContentsMargins(0, 10, 0, 10)
        sub_layout.setSpacing(2)

        grp_lbl = QLabel("CRP / CAPACITY")
        grp_lbl.setStyleSheet(
            "color:#6b7aa3; font-size:10px; font-weight:bold;"
            " padding:6px 16px 4px 16px; background:transparent;"
            " letter-spacing:0.08em;")
        sub_layout.addWidget(grp_lbl)

        _btn_css = (
            "QPushButton { background:transparent; border:none;"
            " border-left:3px solid transparent;"
            " text-align:left; padding:8px 14px 8px 14px;"
            " font-size:12px; color:#374151; }"
            "QPushButton:hover { background:rgba(0,0,0,0.05); }"
            "QPushButton:checked { background:#dbeafe;"
            " border-left-color:#2563eb; color:#1e40af; font-weight:600; }"
        )
        self._btn_labor = QPushButton("  Labor")
        self._btn_labor.setCheckable(True)
        self._btn_labor.setStyleSheet(_btn_css)
        self._btn_labor.clicked.connect(lambda: self._switch_page(0))

        self._btn_prod = QPushButton("  Production Lines")
        self._btn_prod.setCheckable(True)
        self._btn_prod.setStyleSheet(_btn_css)
        self._btn_prod.clicked.connect(lambda: self._switch_page(1))

        self._btn_cap = QPushButton("  Capacity Analysis")
        self._btn_cap.setCheckable(True)
        self._btn_cap.setStyleSheet(_btn_css)
        self._btn_cap.clicked.connect(lambda: self._switch_page(2))

        sub_layout.addWidget(self._btn_labor)
        sub_layout.addWidget(self._btn_prod)
        sub_layout.addWidget(self._btn_cap)
        sub_layout.addStretch()
        outer.addWidget(sub_panel)

        # Content stack
        self._content_stack = QStackedWidget()
        self._content_stack.addWidget(self._build_labor_page())

        from ui.master_tab import CalendarWidget
        self._cal_widget = CalendarWidget(self)
        self._content_stack.addWidget(self._cal_widget)

        self._cap_widget = CapacityAnalysisWidget(self)
        self._content_stack.addWidget(self._cap_widget)

        outer.addWidget(self._content_stack, stretch=1)
        self._switch_page(0)
        self.refresh()

    def _switch_page(self, idx: int):
        self._content_stack.setCurrentIndex(idx)
        self._btn_labor.setChecked(idx == 0)
        self._btn_prod.setChecked(idx == 1)
        self._btn_cap.setChecked(idx == 2)
        if idx == 1:  # Calendar page — reload room list in case rooms were added
            self._cal_widget.refresh()

    def _build_labor_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)

        bar = QHBoxLayout()
        bar.addWidget(QLabel("CRP Excel:"))
        self.path_label = QLabel(ConfigRepo.get("crp_excel_path", "(not set)"))
        bar.addWidget(self.path_label)

        btn_refresh  = QPushButton("♻ Refresh from Excel")
        btn_template = QPushButton("⬇ Download CRP Template")
        btn_hc       = QPushButton("📊 HC Recommendation")
        btn_refresh.clicked.connect(self._refresh)
        btn_template.clicked.connect(self._template)
        btn_hc.clicked.connect(self._auto_fill_hc)
        bar.addWidget(btn_refresh)
        bar.addWidget(btn_template)
        bar.addWidget(btn_hc)

        self._btn_crp_edit = QPushButton("✏ Edit Mode")
        self._btn_crp_edit.setCheckable(True)
        self._btn_crp_edit.clicked.connect(self._toggle_crp_edit_mode)
        bar.addWidget(self._btn_crp_edit)

        self._btn_crp_save = QPushButton("💾 Save to Excel")
        self._btn_crp_save.setVisible(False)
        self._btn_crp_save.clicked.connect(self._save_crp_changes)
        bar.addWidget(self._btn_crp_save)

        bar.addStretch()
        layout.addLayout(bar)

        self.status_label = QLabel("")
        layout.addWidget(self.status_label)

        self.crp_guide = QLabel(
            "⚠  CRP Excel path is not configured. "
            "Go to Masters > App Config > CRP Excel Path and set the path, then click Refresh.")
        self.crp_guide.setStyleSheet(
            "background:#fff3cd; color:#856404; padding:6px 10px; "
            "font-size:12px; border:1px solid #ffc107;")
        self.crp_guide.setWordWrap(True)
        self.crp_guide.setVisible(False)
        layout.addWidget(self.crp_guide)

        layout.addWidget(QLabel("Total Headcount per Shift (from CRP Excel):"))
        self.table = QTableWidget()
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.itemChanged.connect(self._on_crp_cell_changed)
        layout.addWidget(self.table, stretch=1)

        return page

    def refresh(self):
        if self._crp_edit_mode:
            return  # preserve in-progress edits
        path = ConfigRepo.get("crp_excel_path", "")
        self.path_label.setText(path or "(not set)")
        no_path = not path or not __import__("os").path.exists(path)
        self.crp_guide.setVisible(no_path)
        crp_manager.refresh()
        self._load_table()

    def _refresh(self):
        ok, msg = crp_manager.refresh()
        self.status_label.setText(msg)
        self.status_label.setStyleSheet("color: green;" if ok else "color: red;")
        self._load_table()

    def _load_table(self):
        self._crp_loading = True
        self._crp_changed_cells.clear()
        data = crp_manager.get_all()  # {(date_str, shift_no): total_hc}
        if not data:
            self._crp_dates = []
            self._crp_shifts = []
            self.table.setRowCount(0)
            self.table.setColumnCount(2)
            self.table.setHorizontalHeaderLabels(["Date", "HC"])
            self._crp_loading = False
            return

        dates  = sorted({k[0] for k in data.keys()})
        shifts = sorted({k[1] for k in data.keys()})
        self._crp_dates  = dates
        self._crp_shifts = shifts

        # Transposed: rows = dates, columns = shifts
        headers = ["Date"] + [f"Shift {s}" for s in shifts]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(dates))

        for ri, d in enumerate(dates):
            lbl = QTableWidgetItem(d[5:].replace("-", "/"))  # MM/DD
            lbl.setFlags(lbl.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(ri, 0, lbl)
            for si, shift in enumerate(shifts):
                hc   = data.get((d, shift), 0)
                item = QTableWidgetItem(str(hc))
                if hc == 0:
                    item.setBackground(QBrush(QColor("#ffe0e0")))
                if not self._crp_edit_mode:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(ri, 1 + si, item)

        self.table.resizeColumnsToContents()
        self._crp_loading = False

    def _toggle_crp_edit_mode(self, checked: bool):
        self._crp_edit_mode = checked
        self._btn_crp_edit.setText("🔒 Exit Edit Mode" if checked else "✏ Edit Mode")
        self._btn_crp_edit.setStyleSheet(
            "background:#e65100; color:white; font-weight:bold;" if checked else "")
        self._btn_crp_save.setVisible(checked)
        if checked:
            self.table.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked |
                QAbstractItemView.EditTrigger.EditKeyPressed)
            # unlock all data cells
            for ri in range(self.table.rowCount()):
                for ci in range(1, self.table.columnCount()):
                    item = self.table.item(ri, ci)
                    if item:
                        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        else:
            self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            self._load_table()

    def _on_crp_cell_changed(self, item):
        if self._crp_loading or not self._crp_edit_mode:
            return
        if item.column() == 0:
            return
        item.setBackground(QBrush(QColor("#fef08a")))
        self._crp_changed_cells.add((item.row(), item.column()))
        n = len(self._crp_changed_cells)
        self.status_label.setText(f"✏ Edit mode — {n} cell(s) modified")
        self.status_label.setStyleSheet("color:#92400e;")

    def _save_crp_changes(self):
        if not self._crp_changed_cells:
            self._toggle_crp_edit_mode(False)
            self._btn_crp_edit.setChecked(False)
            return

        updates: dict = {}
        errors  = []
        for ri, ci in self._crp_changed_cells:
            item = self.table.item(ri, ci)
            if not item:
                continue
            try:
                hc = int(item.text())
            except ValueError:
                errors.append(f"Row {ri+1}, Col {ci}: '{item.text()}' is not a valid integer")
                continue
            if hc < 0:
                errors.append(f"Row {ri+1}, Col {ci}: HC cannot be negative (got {hc})")
                continue
            if ri >= len(self._crp_dates) or ci - 1 >= len(self._crp_shifts):
                continue
            date_str = self._crp_dates[ri]
            shift_no = self._crp_shifts[ci - 1]
            updates[(date_str, shift_no)] = hc

        if errors:
            QMessageBox.warning(self, "Invalid Values", "\n".join(errors))
            return

        ok, msg = crp_manager.write_total_hc(updates)
        if ok:
            self.status_label.setText(f"✅ Saved {len(updates)} cell(s) to CRP Excel")
            self.status_label.setStyleSheet("color:green;")
            self._crp_edit_mode = False
            self._btn_crp_edit.setChecked(False)
            self._btn_crp_edit.setText("✏ Edit Mode")
            self._btn_crp_edit.setStyleSheet("")
            self._btn_crp_save.setVisible(False)
            self._load_table()
        else:
            QMessageBox.warning(self, "Save Failed", msg)

    def _template(self):
        from datetime import date as dt
        path, _ = QFileDialog.getSaveFileName(self, "Save CRP Template", "CRP_template.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        shifts = [s["shift_no"] for s in ShiftRepo.all()]
        today = dt.today()
        d0 = today.strftime("%Y-%m-%d")
        d1 = (today + timedelta(days=56)).strftime("%Y-%m-%d")
        ok, msg = crp_manager.create_template(path, shifts, d0, d1)
        (QMessageBox.information if ok else QMessageBox.warning)(self, "Template", msg)

    def _auto_fill_hc(self):
        dlg = HCDemandDialog(self)
        if dlg.exec() and dlg.to_apply:
            ok, msg = crp_manager.write_total_hc(dlg.to_apply)
            (QMessageBox.information if ok else QMessageBox.warning)(self, "HC Recommendation", msg)
            if ok:
                self._load_table()


# ════════════════════════════════════════════════════════════════════════════
#  ACTUALS TAB
# ════════════════════════════════════════════════════════════════════════════

class ActualsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._alloc_rows: list = []
        self._current_lot: dict = {}
        self._excess: int = 0
        self._loading_alloc: bool = False
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── Zone 1: Lot Entry ─────────────────────────────────────────────
        entry_group = QGroupBox("Lot Entry")
        el = QVBoxLayout(entry_group)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("Date:"))
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        row1.addWidget(self.date_edit)
        row1.addSpacing(24)
        row1.addWidget(QLabel("Type:"))
        self.radio_sku = QRadioButton("SKU")
        self.radio_mat = QRadioButton("Material")
        self.radio_sku.setChecked(True)
        self.radio_sku.toggled.connect(self._on_type_changed)
        row1.addWidget(self.radio_sku)
        row1.addWidget(self.radio_mat)
        row1.addStretch()
        el.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Code:"))
        self.code_combo = QComboBox()
        self.code_combo.setEditable(True)
        self.code_combo.setMinimumWidth(200)
        row2.addWidget(self.code_combo)
        row2.addSpacing(12)
        row2.addWidget(QLabel("Lot Number:"))
        self.lot_edit = QLineEdit()
        self.lot_edit.setMinimumWidth(150)
        row2.addWidget(self.lot_edit)
        row2.addSpacing(12)
        row2.addWidget(QLabel("Qty:"))
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 9999999)
        self.qty_spin.setMinimumWidth(100)
        row2.addWidget(self.qty_spin)
        row2.addStretch()
        el.addLayout(row2)

        row3 = QHBoxLayout()
        row3.addWidget(QLabel("Note:"))
        self.note_edit = QLineEdit()
        self.note_edit.setPlaceholderText("Optional")
        row3.addWidget(self.note_edit, stretch=1)
        self.btn_preview = QPushButton("🔍 Preview Allocation")
        self.btn_preview.setStyleSheet(
            "background:#2e6fd8; color:white; font-weight:bold; padding:5px 14px;")
        self.btn_preview.clicked.connect(self._on_entry_action)
        row3.addWidget(self.btn_preview)
        el.addLayout(row3)

        layout.addWidget(entry_group)

        # ── Zone 2: Allocation Preview (hidden until previewed) ───────────
        self.preview_group = QGroupBox("Allocation Preview")
        self.preview_group.setVisible(False)
        pl = QVBoxLayout(self.preview_group)

        self.info_label = QLabel("")
        self.info_label.setStyleSheet(
            "background:#e8f0fe; color:#1a3a7a; font-weight:bold; "
            "padding:6px 10px; border-radius:4px;")
        pl.addWidget(self.info_label)

        self.alloc_table = QTableWidget()
        self.alloc_table.setColumnCount(8)
        self.alloc_table.setHorizontalHeaderLabels([
            "SO#", "Line", "Customer", "Priority",
            "Due Date", "SO Qty", "Actual So Far", "Allocate Qty",
        ])
        hh = self.alloc_table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.alloc_table.setAlternatingRowColors(True)
        self.alloc_table.itemChanged.connect(self._on_alloc_changed)
        pl.addWidget(self.alloc_table)

        excess_row = QHBoxLayout()
        self.excess_label = QLabel("")
        self.excess_label.setStyleSheet("color:#c0392b; font-weight:bold;")
        self.excess_label.setVisible(False)
        excess_row.addWidget(self.excess_label)
        self.radio_inv = QRadioButton("Save as Inventory")
        self.radio_discard = QRadioButton("Discard / Note only")
        self.radio_discard.setChecked(True)
        self.radio_inv.setVisible(False)
        self.radio_discard.setVisible(False)
        excess_row.addWidget(self.radio_inv)
        excess_row.addWidget(self.radio_discard)
        excess_row.addStretch()
        pl.addLayout(excess_row)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel_preview = QPushButton("Cancel")
        btn_cancel_preview.clicked.connect(self._cancel_preview)
        self.btn_confirm = QPushButton("✅ Confirm & Save")
        self.btn_confirm.setStyleSheet(
            "background:#27ae60; color:white; font-weight:bold; padding:6px 18px;")
        self.btn_confirm.clicked.connect(self._confirm_save)
        btn_row.addWidget(btn_cancel_preview)
        btn_row.addWidget(self.btn_confirm)
        pl.addLayout(btn_row)

        layout.addWidget(self.preview_group)

        # ── Zone 3: Today's Lots Log ──────────────────────────────────────
        log_group = QGroupBox("Actuals Log")
        ll = QVBoxLayout(log_group)

        log_bar = QHBoxLayout()
        btn_sample = QPushButton("🧪 LOT Sample Qty")
        btn_sample.clicked.connect(self._open_sample_entry)
        btn_replan = QPushButton("🔄 Replan after Actuals")
        btn_replan.clicked.connect(self._replan)
        log_bar.addWidget(btn_sample)
        log_bar.addWidget(btn_replan)
        log_bar.addStretch()
        log_bar.addWidget(QLabel("Log Date:"))
        self.log_date = QDateEdit(QDate.currentDate())
        self.log_date.setDisplayFormat("yyyy-MM-dd")
        self.log_date.dateChanged.connect(self._refresh_log)
        log_bar.addWidget(self.log_date)
        btn_refresh = QPushButton("Refresh")
        btn_refresh.clicked.connect(self._refresh_log)
        log_bar.addWidget(btn_refresh)
        ll.addLayout(log_bar)

        self.log_table = QTableWidget()
        self.log_table.setColumnCount(8)
        self.log_table.setHorizontalHeaderLabels([
            "Time", "Type", "Code", "Lot Number",
            "SO#", "Line", "Qty", "Note",
        ])
        self.log_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lhh = self.log_table.horizontalHeader()
        lhh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        lhh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.log_table.setAlternatingRowColors(True)
        ll.addWidget(self.log_table, stretch=1)

        self.log_status = QLabel("")
        self.log_status.setStyleSheet("color:#555; font-size:11px; padding:2px 4px;")
        ll.addWidget(self.log_status)

        layout.addWidget(log_group, stretch=1)

        self._reload_codes()
        self._refresh_log()

    # ── Type toggle ───────────────────────────────────────────────────────

    def _on_type_changed(self):
        self._reload_codes()
        if self.radio_sku.isChecked():
            self.btn_preview.setText("🔍 Preview Allocation")
        else:
            self.btn_preview.setText("💾 Save Material Actual")
        self.preview_group.setVisible(False)

    def _reload_codes(self):
        self.code_combo.clear()
        if self.radio_sku.isChecked():
            for sku in SKURepo.all():
                self.code_combo.addItem(sku["sku_code"])
        else:
            for mat in MaterialRepo.all():
                self.code_combo.addItem(mat["material_code"])
        self.code_combo.setCurrentIndex(-1)
        if self.code_combo.lineEdit():
            self.code_combo.lineEdit().setPlaceholderText("Type or select…")

    # ── Entry action ──────────────────────────────────────────────────────

    def _on_entry_action(self):
        code = self.code_combo.currentText().strip()
        lot  = self.lot_edit.text().strip()
        qty  = self.qty_spin.value()

        if not code:
            QMessageBox.warning(self, "Input Required", "Please select a code.")
            return
        if not lot:
            QMessageBox.warning(self, "Input Required", "Please enter a Lot Number.")
            return

        if self.radio_mat.isChecked():
            self._save_material_actual(code, lot, qty)
            return

        # SKU: compute allocation preview
        actual_date = self.date_edit.date().toString("yyyy-MM-dd")
        self._current_lot = {
            "entity_type": "SKU",
            "sku_code": code,
            "lot_number": lot,
            "total_qty": qty,
            "actual_date": actual_date,
            "note": self.note_edit.text().strip(),
        }

        demands = SORepo.open_demand_for_sku(code)
        remaining = qty
        self._alloc_rows = []
        for d in demands:
            alloc = min(d["remaining_needed"], remaining)
            self._alloc_rows.append({
                "so_number":       d["so_number"],
                "sku_code":        d["sku_code"],
                "line_item":       d["line_item"],
                "customer_name":   d.get("customer_name") or "",
                "priority":        d.get("priority", ""),
                "due_date":        d.get("due_date", ""),
                "so_qty":          d["qty"],
                "qty_actual_so_far": d["qty_actual_total"],
                "alloc_qty":       alloc,
            })
            remaining -= alloc
            if remaining <= 0:
                break
        self._excess = remaining
        self._render_alloc_table()
        self.preview_group.setVisible(True)

    # ── Allocation preview ────────────────────────────────────────────────

    def _render_alloc_table(self):
        self._loading_alloc = True
        hdr = self.alloc_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.alloc_table.setUpdatesEnabled(False)
        self.alloc_table.setRowCount(len(self._alloc_rows))
        today = date.today()
        for ri, row in enumerate(self._alloc_rows):
            due_str = row.get("due_date", "")
            vals = [
                row["so_number"], row["line_item"],
                row["customer_name"],
                str(row.get("priority", "")),
                due_str,
                str(row["so_qty"]),
                str(row["qty_actual_so_far"]),
                str(row["alloc_qty"]),
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(val)
                if ci == 7:
                    item.setBackground(QBrush(QColor("#fffde7")))
                else:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if ci == 4 and due_str:
                    try:
                        due = date.fromisoformat(due_str)
                        if due < today:
                            item.setForeground(QBrush(QColor("#c0392b")))
                        elif (due - today).days <= 3:
                            item.setForeground(QBrush(QColor("#e67e22")))
                    except ValueError:
                        pass
                self.alloc_table.setItem(ri, ci, item)
        self.alloc_table.setUpdatesEnabled(True)
        self.alloc_table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._loading_alloc = False
        self._update_alloc_summary()

    def _update_alloc_summary(self):
        total_alloc = sum(r["alloc_qty"] for r in self._alloc_rows)
        total_qty   = self._current_lot.get("total_qty", 0)
        self._excess = total_qty - total_alloc

        self.info_label.setText(
            f"Lot: {self._current_lot.get('lot_number', '')}  ·  "
            f"SKU: {self._current_lot.get('sku_code', '')}  ·  "
            f"Total Qty: {total_qty}  ·  "
            f"Allocated: {total_alloc}  ·  "
            f"Excess: {self._excess}"
        )
        has_excess = self._excess > 0
        self.excess_label.setVisible(has_excess)
        self.radio_inv.setVisible(has_excess)
        self.radio_discard.setVisible(has_excess)
        if has_excess:
            self.excess_label.setText(
                f"⚠  {self._excess} unit(s) unallocated — disposition:")
        self.btn_confirm.setEnabled(total_alloc > 0 or has_excess)

    def _on_alloc_changed(self, item):
        if self._loading_alloc or item.column() != 7:
            return
        ri = item.row()
        if ri >= len(self._alloc_rows):
            return
        try:
            new_qty = max(0, int(item.text()))
        except ValueError:
            new_qty = self._alloc_rows[ri]["alloc_qty"]
        self._alloc_rows[ri]["alloc_qty"] = new_qty
        self._update_alloc_summary()

    def _cancel_preview(self):
        self.preview_group.setVisible(False)
        self._alloc_rows = []
        self._current_lot = {}

    def _confirm_save(self):
        actual_date = self._current_lot["actual_date"]
        lot         = self._current_lot["lot_number"]
        sku         = self._current_lot["sku_code"]
        note        = self._current_lot.get("note", "")

        saved = 0
        for row in self._alloc_rows:
            if row["alloc_qty"] <= 0:
                continue
            ActualRepo.insert({
                "entity_type":  "SKU",
                "entity_code":  sku,
                "so_number":    row["so_number"],
                "sku_code":     sku,
                "line_item":    row["line_item"],
                "lot_number":   lot,
                "actual_date":  actual_date,
                "qty_actual":   row["alloc_qty"],
                "note":         note,
            })
            saved += 1

        excess_msg = ""
        if self._excess > 0:
            if self.radio_inv.isChecked():
                InventoryRepo.add_excess(sku, lot, self._excess, actual_date)
                excess_msg = f", {self._excess} unit(s) saved to inventory"
            else:
                ActualRepo.insert({
                    "entity_type": "SKU",
                    "entity_code": sku,
                    "so_number":   "",
                    "sku_code":    sku,
                    "line_item":   "",
                    "lot_number":  lot,
                    "actual_date": actual_date,
                    "qty_actual":  self._excess,
                    "note":        f"EXCESS{' — ' + note if note else ''}",
                })
                excess_msg = f", {self._excess} noted as excess"

        self.preview_group.setVisible(False)
        self._alloc_rows = []
        self._current_lot = {}
        self.lot_edit.clear()
        self.qty_spin.setValue(1)
        self.note_edit.clear()
        self._refresh_log()

        if self.main_window:
            self.main_window.notify(
                f"Saved {saved} actual record(s){excess_msg}.")

    # ── Material direct save ──────────────────────────────────────────────

    def _save_material_actual(self, code: str, lot: str, qty: int):
        ActualRepo.insert({
            "entity_type": "MATERIAL",
            "entity_code": code,
            "lot_number":  lot,
            "actual_date": self.date_edit.date().toString("yyyy-MM-dd"),
            "qty_actual":  qty,
            "note":        self.note_edit.text().strip(),
        })
        self.lot_edit.clear()
        self.qty_spin.setValue(1)
        self.note_edit.clear()
        self._refresh_log()
        if self.main_window:
            self.main_window.notify(
                f"Material actual saved: {code} · Lot {lot} · {qty}")

    # ── Log ───────────────────────────────────────────────────────────────

    def _refresh_log(self):
        d = self.log_date.date().toString("yyyy-MM-dd")
        actuals = ActualRepo.for_date(d)
        lhdr = self.log_table.horizontalHeader()
        lhdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.log_table.setUpdatesEnabled(False)
        self.log_table.setRowCount(len(actuals))
        for ri, a in enumerate(actuals):
            ts = a.get("entered_at", "")
            time_str = ts[11:16] if len(ts) >= 16 else ""
            note = a.get("note") or ""
            tag  = "EXCESS" if note.startswith("EXCESS") else ""
            vals = [
                time_str,
                a.get("entity_type", ""),
                a.get("entity_code", ""),
                a.get("lot_number", ""),
                a.get("so_number", "") or tag,
                a.get("line_item", ""),
                str(a.get("qty_actual", "")),
                note,
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(val)
                if tag and ci in (4, 7):
                    item.setForeground(QBrush(QColor("#e67e22")))
                self.log_table.setItem(ri, ci, item)
        self.log_table.setUpdatesEnabled(True)
        self.log_table.resizeColumnsToContents()
        lhdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        lhdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)

        total_sku = sum(
            a.get("qty_actual", 0) for a in actuals
            if a.get("entity_type") == "SKU" and a.get("so_number")
        )
        excess_sku = sum(
            a.get("qty_actual", 0) for a in actuals
            if a.get("entity_type") == "SKU" and not a.get("so_number")
        )
        self.log_status.setText(
            f"{len(actuals)} record(s) on {d}  ·  "
            f"SKU allocated: {total_sku}  ·  Excess: {excess_sku}"
        )

    # ── Public ────────────────────────────────────────────────────────────

    def refresh(self):
        self._reload_codes()
        self._refresh_log()

    def _open_sample_entry(self):
        dlg = LotSampleDialog(self)
        dlg.exec()

    def _replan(self):
        d0 = date.today().strftime("%Y-%m-%d")
        d1 = (date.today() + timedelta(weeks=8)).strftime("%Y-%m-%d")
        result = scheduler.replan_after_actuals(d0, d1)
        dlg = ReplanReportDialog(result, self)
        dlg.exec()
        if self.main_window:
            self.main_window.gantt_tab.refresh()
            self.main_window.notify(
                f"Replan: {len(result['deleted'])} deleted, "
                f"{len(result['replanned'])} re-planned, "
                f"{len(result['errors'])} errors."
            )


class ReplanReportDialog(QDialog):
    """리플래닝 실행 결과를 삭제/재계획/오류로 나눠 보여주는 다이얼로그."""

    def __init__(self, result: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Replan from Actuals — Report")
        self.setMinimumSize(740, 480)
        self._result = result
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        deleted   = self._result.get("deleted", [])
        replanned = self._result.get("replanned", [])
        errors    = self._result.get("errors", [])

        summary = QLabel(
            f"Deleted: {len(deleted)} plan(s)   |   "
            f"Re-planned: {len(replanned)} SO line(s)   |   "
            f"Errors: {len(errors)}"
        )
        summary.setStyleSheet(
            "font-weight:bold; font-size:12px; padding:8px; "
            "background:#e8f0fe; border-radius:4px;")
        layout.addWidget(summary)

        tabs = QTabWidget()

        # ── Deleted ──
        t_del = QTableWidget(len(deleted), 4)
        t_del.setHorizontalHeaderLabels(["Plan ID", "SO", "SKU / Line", "Reason"])
        t_del.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        t_del.horizontalHeader().setStretchLastSection(True)
        t_del.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        for ri, d in enumerate(deleted):
            t_del.setItem(ri, 0, QTableWidgetItem(str(d["plan_id"])))
            t_del.setItem(ri, 1, QTableWidgetItem(d["so_number"]))
            t_del.setItem(ri, 2, QTableWidgetItem(f"{d['sku_code']} / {d['line_item']}"))
            t_del.setItem(ri, 3, QTableWidgetItem(d["reason"]))
            for ci in range(4):
                t_del.item(ri, ci).setBackground(QBrush(QColor("#fff9c4")))
        t_del.resizeColumnsToContents()
        tabs.addTab(t_del, f"Deleted ({len(deleted)})")

        # ── Re-planned ──
        t_rep = QTableWidget(len(replanned), 5)
        t_rep.setHorizontalHeaderLabels(["SO", "SKU", "Line", "Actual Qty", "Remaining Qty"])
        t_rep.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        t_rep.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        for ri, r in enumerate(replanned):
            t_rep.setItem(ri, 0, QTableWidgetItem(r["so_number"]))
            t_rep.setItem(ri, 1, QTableWidgetItem(r["sku_code"]))
            t_rep.setItem(ri, 2, QTableWidgetItem(r["line_item"]))
            t_rep.setItem(ri, 3, QTableWidgetItem(str(r["actual_qty"])))
            t_rep.setItem(ri, 4, QTableWidgetItem(str(r["remaining_qty"])))
            for ci in range(5):
                t_rep.item(ri, ci).setBackground(QBrush(QColor("#c8e6c9")))
        t_rep.resizeColumnsToContents()
        tabs.addTab(t_rep, f"Re-planned ({len(replanned)})")

        # ── Errors ──
        t_err = QTableWidget(len(errors), 2)
        t_err.setHorizontalHeaderLabels(["SO", "Reason"])
        t_err.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        t_err.horizontalHeader().setStretchLastSection(True)
        t_err.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        for ri, e in enumerate(errors):
            t_err.setItem(ri, 0, QTableWidgetItem(str(e.get("so", ""))))
            t_err.setItem(ri, 1, QTableWidgetItem(str(e.get("reason", ""))))
            for ci in range(2):
                t_err.item(ri, ci).setBackground(QBrush(QColor("#ffcdd2")))
        t_err.resizeColumnsToContents()
        tabs.addTab(t_err, f"Errors ({len(errors)})")

        layout.addWidget(tabs, stretch=1)

        btn = QPushButton("Close")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn, alignment=Qt.AlignmentFlag.AlignRight)


# ════════════════════════════════════════════════════════════════════════════
#  ALERTS TAB
# ════════════════════════════════════════════════════════════════════════════

class AlertsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        bar = QHBoxLayout()
        btn_refresh = QPushButton("🔍 Refresh Alerts")
        btn_refresh.clicked.connect(self.refresh)
        bar.addWidget(btn_refresh); bar.addStretch()
        layout.addLayout(bar)

        self.conflict_grp = QGroupBox("⚠ Capacity Conflicts")
        cg_l = QVBoxLayout(self.conflict_grp)
        self.conflict_table = _alerts_table(
            ["Date", "Room", "Process", "Shift", "Planned", "Capacity", "Overrun"]
        )
        cg_l.addWidget(self.conflict_table)
        layout.addWidget(self.conflict_grp)

        self.late_grp = QGroupBox("🔴 Late / At-Risk SOs")
        lg_l = QVBoxLayout(self.late_grp)
        self.late_table = _alerts_table(
            ["SO", "SKU", "Line", "Due Date", "Planned Complete", "Remaining Qty", "Reason"]
        )
        lg_l.addWidget(self.late_table)
        layout.addWidget(self.late_grp)

        self.refresh()

    def refresh(self):
        self._load_conflicts()
        self._load_late()

    def _load_conflicts(self):
        d0 = date.today().strftime("%Y-%m-%d")
        d1 = (date.today() + timedelta(weeks=4)).strftime("%Y-%m-%d")
        conflicts = scheduler.detect_conflicts(d0, d1)
        self.conflict_table.setRowCount(0)
        if not conflicts:
            self.conflict_grp.setTitle("✅ Capacity Conflicts — None detected")
            self.conflict_table.setRowCount(1)
            item = QTableWidgetItem("✅  No capacity conflicts in the next 4 weeks")
            item.setForeground(QBrush(QColor("#2d7a2d")))
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.conflict_table.setItem(0, 0, item)
            self.conflict_table.setSpan(0, 0, 1, self.conflict_table.columnCount())
            return
        self.conflict_grp.setTitle(f"⚠ Capacity Conflicts ({len(conflicts)})")
        red_bg = QBrush(QColor("#ffe0e0"))
        hdr = self.conflict_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.conflict_table.setUpdatesEnabled(False)
        self.conflict_table.setRowCount(len(conflicts))
        for ri, c in enumerate(conflicts):
            if c.get("conflict_type") == "multi_process":
                procs = ", ".join(c.get("processes", []))
                row_vals = [c["plan_date"], c["room_code"], procs, c["shift_no"],
                            "multi-process", "—", "—"]
            else:
                row_vals = [c["plan_date"], c["room_code"], c["process_name"], c["shift_no"],
                            c["planned_inner"], c["capacity_inner"], c["overrun_inner"]]
            for ci, val in enumerate(row_vals):
                item = QTableWidgetItem(str(val))
                item.setBackground(red_bg)
                self.conflict_table.setItem(ri, ci, item)
        self.conflict_table.setUpdatesEnabled(True)
        self.conflict_table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

    def _load_late(self):
        today = date.today()
        sos = SORepo.all("OPEN")
        # Batch queries — avoids N+1 (2 queries total regardless of SO count)
        actual_map    = ActualRepo.actual_qty_bulk()
        last_plan_map = PlanRepo.last_plan_info_bulk()
        late_rows = []
        for so in sos:
            key = (so["so_number"], so["sku_code"], so["line_item"])
            due = datetime.strptime(so["due_date"], "%Y-%m-%d").date()
            actual = actual_map.get(key, 0)
            remaining = so["qty"] - actual
            if remaining <= 0:
                continue
            last_info = last_plan_map.get(key)
            if last_info:
                plan_complete = f"{last_info[0]} S{last_info[1]}"
                plan_complete_date = datetime.strptime(last_info[0], "%Y-%m-%d").date()
                is_late = plan_complete_date > due
            else:
                plan_complete = "Not planned"
                is_late = True

            if is_late or due < today:
                late_rows.append({
                    "so": so["so_number"], "sku": so["sku_code"],
                    "li": so["line_item"], "due": so["due_date"],
                    "complete": plan_complete, "remaining": remaining,
                    "reason": "Late" if due < today else "Plan exceeds due date"
                })

        self.late_table.setRowCount(0)
        if not late_rows:
            self.late_grp.setTitle("✅ Late / At-Risk SOs — None")
            self.late_table.setRowCount(1)
            item = QTableWidgetItem("✅  All open SOs are on track")
            item.setForeground(QBrush(QColor("#2d7a2d")))
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.late_table.setItem(0, 0, item)
            self.late_table.setSpan(0, 0, 1, self.late_table.columnCount())
        else:
            self.late_grp.setTitle(f"🔴 Late / At-Risk SOs ({len(late_rows)})")
            late_bg = QBrush(QColor("#ffcccc"))
            hdr = self.late_table.horizontalHeader()
            hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
            self.late_table.setUpdatesEnabled(False)
            self.late_table.setRowCount(len(late_rows))
            for ri, r in enumerate(late_rows):
                for ci, val in enumerate([
                    r["so"], r["sku"], r["li"], r["due"],
                    r["complete"], r["remaining"], r["reason"]
                ]):
                    item = QTableWidgetItem(str(val))
                    item.setBackground(late_bg)
                    self.late_table.setItem(ri, ci, item)
            self.late_table.setUpdatesEnabled(True)
            self.late_table.resizeColumnsToContents()
            hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # QC shortfall section — pass pre-fetched maps to avoid re-querying
        self._load_qc_shortfall(sos, actual_map)

    def _load_qc_shortfall(self, sos=None, actual_map=None):
        """Show SOs where actual - sample - reject < SO qty."""
        from data.repositories import LotSampleRepo
        if sos is None:
            sos = SORepo.all("OPEN")
        if actual_map is None:
            actual_map = ActualRepo.actual_qty_bulk()
        # Batch sample+reject query (1 query)
        sr_map = LotSampleRepo.sample_reject_bulk()
        shortfall_rows = []
        for so in sos:
            key = (so["so_number"], so["sku_code"], so["line_item"])
            actual_total = actual_map.get(key, 0)
            if actual_total == 0:
                continue   # nothing produced yet — skip
            sample_total, reject_total = sr_map.get(key, (0, 0))
            net = actual_total - sample_total - reject_total
            if net < so["qty"]:
                shortfall_rows.append({
                    "so":     so["so_number"],
                    "sku":    so["sku_code"],
                    "li":     so["line_item"],
                    "due":    so["due_date"],
                    "actual": actual_total,
                    "sample": sample_total,
                    "reject": reject_total,
                    "net":    net,
                    "needed": so["qty"],
                    "short":  so["qty"] - net,
                })

        # Re-use or create QC shortfall table
        if not hasattr(self, "qc_table"):
            self.qc_grp = QGroupBox("🔴 QC Shortfall (net qty < SO qty)")
            qg_l = QVBoxLayout(self.qc_grp)
            self.qc_table = _alerts_table([
                "SO", "SKU", "Line", "Due Date",
                "Actual", "Sample", "Reject", "Net Qty",
                "SO Qty", "Short By"
            ])
            qg_l.addWidget(self.qc_table)
            # Insert after late_grp in parent layout
            parent_layout = self.layout()
            parent_layout.addWidget(self.qc_grp)

        self.qc_table.setRowCount(0)
        if not shortfall_rows:
            self.qc_grp.setTitle("✅ QC Shortfall — None")
            self.qc_table.setRowCount(1)
            item = QTableWidgetItem("✅  No QC shortfalls detected")
            item.setForeground(QBrush(QColor("#2d7a2d")))
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.qc_table.setItem(0, 0, item)
            self.qc_table.setSpan(0, 0, 1, self.qc_table.columnCount())
            return
        self.qc_grp.setTitle(f"🔴 QC Shortfall — {len(shortfall_rows)} SO(s)")
        red_bg    = QBrush(QColor("#ffcccc"))
        yellow_bg = QBrush(QColor("#fffbe6"))
        orange_bg = QBrush(QColor("#fde8d8"))
        hdr = self.qc_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.qc_table.setUpdatesEnabled(False)
        self.qc_table.setRowCount(len(shortfall_rows))
        for ri, r in enumerate(shortfall_rows):
            for ci, val in enumerate([
                r["so"], r["sku"], r["li"], r["due"],
                r["actual"], r["sample"], r["reject"],
                r["net"], r["needed"], r["short"]
            ]):
                item = QTableWidgetItem(str(val))
                if ci in (6, 9):
                    item.setBackground(red_bg)
                elif ci == 5:
                    item.setBackground(yellow_bg)
                elif ci == 7:
                    item.setBackground(orange_bg)
                self.qc_table.setItem(ri, ci, item)
        self.qc_table.setUpdatesEnabled(True)
        self.qc_table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)


def _alerts_table(cols: list) -> QTableWidget:
    t = QTableWidget()
    t.setColumnCount(len(cols))
    t.setHorizontalHeaderLabels(cols)
    t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    t.setAlternatingRowColors(True)
    t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
    t.horizontalHeader().setStretchLastSection(True)
    return t


def _fill_alerts_table(table: QTableWidget, rows: list, fill_fn):
    """Fill an alerts table with Fixed mode during fill then Interactive after."""
    hdr = table.horizontalHeader()
    hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
    table.setUpdatesEnabled(False)
    table.setRowCount(len(rows))
    for ri, r in enumerate(rows):
        fill_fn(table, ri, r)
    table.setUpdatesEnabled(True)
    table.resizeColumnsToContents()
    hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)


# ════════════════════════════════════════════════════════════════════════════
#  DASHBOARD TAB
# ════════════════════════════════════════════════════════════════════════════

class DashboardTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("📊 Production Dashboard"))

        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.refresh)
        layout.addWidget(btn_refresh)

        self.summary_label = QLabel()
        self.summary_label.setWordWrap(True)
        self.summary_label.setStyleSheet("font-size:13px; padding:8px;")
        layout.addWidget(self.summary_label)

        # SO status table by due week
        self.week_table = QTableWidget()
        self.week_table.setColumnCount(6)
        self.week_table.setHorizontalHeaderLabels(
            ["Week", "Total SOs", "Fully Planned", "Partially Planned",
             "Not Planned", "Actual Started"])
        self.week_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.week_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.week_table.setAlternatingRowColors(True)
        layout.addWidget(self.week_table, stretch=1)

        self.refresh()

    def refresh(self):
        self._compute_summary()

    def _compute_summary(self):
        sos = SORepo.all("OPEN")
        total = len(sos)
        # Batch queries — 2 queries regardless of SO count
        planned_map = PlanRepo.planned_qty_bulk()
        actual_map  = ActualRepo.actual_qty_bulk()

        fully, partial, none_ = 0, 0, 0
        from collections import defaultdict
        weeks: Dict[str, Dict] = defaultdict(lambda: {"total": 0, "full": 0, "part": 0, "none": 0, "started": 0})

        for so in sos:
            key      = (so["so_number"], so["sku_code"], so["line_item"])
            planned  = planned_map.get(key, 0)
            actual   = actual_map.get(key, 0)
            remaining = so["qty"] - actual

            if planned >= remaining:
                fully += 1
            elif planned > 0:
                partial += 1
            else:
                none_ += 1

            try:
                due = datetime.strptime(so["due_date"], "%Y-%m-%d").date()
                week_key = due.strftime("%Y-W%V")
            except Exception:
                week_key = "Unknown"
            weeks[week_key]["total"] += 1
            if actual > 0:
                weeks[week_key]["started"] += 1
            if planned >= remaining:
                weeks[week_key]["full"] += 1
            elif planned > 0:
                weeks[week_key]["part"] += 1
            else:
                weeks[week_key]["none"] += 1

        self.summary_label.setText(
            f"Total Open SOs: <b>{total}</b>  |  "
            f"Fully Planned: <b style='color:green'>{fully}</b>  |  "
            f"Partially Planned: <b style='color:orange'>{partial}</b>  |  "
            f"Not Planned: <b style='color:red'>{none_}</b>"
        )

        week_keys = sorted(weeks.keys())
        none_bg = QBrush(QColor("#ffcccc"))
        total_bg = QBrush(QColor("#e8f0fe"))
        hdr = self.week_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.week_table.setUpdatesEnabled(False)
        self.week_table.setRowCount(len(week_keys))
        for ri, wk in enumerate(week_keys):
            w = weeks[wk]
            for ci, val in enumerate([wk, w["total"], w["full"], w["part"], w["none"], w["started"]]):
                item = QTableWidgetItem(str(val))
                if ci == 4 and val > 0:
                    item.setBackground(none_bg)
                elif ci == 1:
                    item.setBackground(total_bg)
                self.week_table.setItem(ri, ci, item)
        self.week_table.setUpdatesEnabled(True)
        self.week_table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)


# ════════════════════════════════════════════════════════════════════════════
#  LOT SAMPLE DIALOG
# ════════════════════════════════════════════════════════════════════════════

class LotSampleDialog(QDialog):
    """
    Planner enters QC sample and reject quantities per LOT after production.
    net_qty = actual_qty - sample_qty - reject_qty (deliverable qty).
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("LOT QC Quantity Entry")
        self.resize(1050, 600)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Filter bar
        fbar = QHBoxLayout()
        fbar.addWidget(QLabel("Filter SO/SKU:"))
        self.search = QLineEdit()
        self.search.setPlaceholderText("SO number or SKU code…")
        self.search.textChanged.connect(self._load)
        fbar.addWidget(self.search)
        fbar.addStretch()
        layout.addLayout(fbar)

        # Summary label
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("padding:4px; font-weight:bold;")
        layout.addWidget(self.summary_label)

        # Legend
        legend = QLabel(
            "  🟡 Sample Qty — consumed for QC sampling   "
            "  🔴 Reject Qty — QC-rejected units   "
            "  Net Qty = Actual − Sample − Reject")
        legend.setStyleSheet(
            "font-size:11px; color:var(--secondary); "
            "padding:4px 8px; background:#f8f8f8;")
        layout.addWidget(legend)

        # Table columns:
        # 0:actual_id | 1:type | 2:code | 3:SO/Line | 4:LOT
        # 5:actual_qty | 6:sample_qty✏ | 7:reject_qty✏ | 8:net_qty | 9:note✏
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Actual ID", "Type", "Code", "SO / Line",
            "LOT Number", "Actual Qty",
            "Sample Qty ✏", "Reject Qty ✏", "Net Qty", "Note ✏"
        ])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table, stretch=1)

        # Buttons
        bbar = QHBoxLayout()
        btn_save = QPushButton("💾 Save QC Entries")
        btn_save.clicked.connect(self._save)
        bbar.addWidget(btn_save)
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.accept)
        bbar.addWidget(btn_close)
        bbar.addStretch()
        layout.addLayout(bbar)

        self._load()

    def _load(self):
        from data.repositories import ActualRepo, LotSampleRepo
        actuals = ActualRepo.recent(300)
        f = self.search.text().lower()
        if f:
            actuals = [a for a in actuals
                       if f in (a.get("so_number", "") +
                                a.get("entity_code", "")).lower()]

        lot_hdr = self.table.horizontalHeader()
        lot_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(actuals))
        self._actual_rows = actuals

        SAMPLE_BG = QColor("#fffbe6")   # yellow tint — sample
        REJECT_BG = QColor("#fde8e8")   # red tint   — reject
        NET_OK    = QColor("#d4f0c0")   # green      — sufficient
        NET_SHORT = QColor("#ffcccc")   # red        — short

        short_count = 0
        for ri, a in enumerate(actuals):
            samples     = LotSampleRepo.for_actual(a["actual_id"])
            sample_qty  = sum(s["sample_qty"] for s in samples)
            reject_qty  = sum(s.get("reject_qty", 0) for s in samples)
            net_qty     = a["qty_actual"] - sample_qty - reject_qty

            so_line = (f"{a.get('so_number','')}/{a.get('line_item','')}"
                       if a.get("so_number") else "—")
            so_needed = 0
            if a.get("so_number") and a.get("sku_code"):
                so = SORepo.get(
                    a["so_number"], a["sku_code"],
                    a.get("line_item", ""))
                if so:
                    so_needed = so["qty"]

            row_vals = [
                a["actual_id"],
                a["entity_type"],
                a["entity_code"],
                so_line,
                a.get("lot_number") or "—",
                a["qty_actual"],
                sample_qty,    # col 6 — editable
                reject_qty,    # col 7 — editable
                net_qty,       # col 8 — computed
                ""             # col 9 — note editable
            ]

            for ci, val in enumerate(row_vals):
                item = QTableWidgetItem(str(val))

                if ci == 6:    # sample qty — editable, yellow
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(QBrush(SAMPLE_BG))
                elif ci == 7:  # reject qty — editable, red tint
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(QBrush(REJECT_BG))
                elif ci == 8:  # net qty — colour coded
                    is_short = (so_needed > 0 and net_qty < so_needed)
                    item.setBackground(QBrush(NET_SHORT if is_short else NET_OK))
                    if is_short:
                        short_count += 1
                        item.setText(
                            f"{net_qty}  ⚠ short {so_needed - net_qty}")
                        item.setFlags(
                            item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    else:
                        item.setFlags(
                            item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                elif ci == 9:  # note — editable
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                else:
                    item.setFlags(
                        item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                self.table.setItem(ri, ci, item)

        self.table.setUpdatesEnabled(True)
        self.table.resizeColumnsToContents()
        lot_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        if short_count:
            self.summary_label.setText(
                f"⚠  {short_count} SO-LOT(s) short after sample/reject deduction")
            self.summary_label.setStyleSheet(
                "color:red; font-weight:bold; padding:4px;")
        else:
            self.summary_label.setText(
                f"✅  {len(actuals)} actuals — all SOs appear sufficient")
            self.summary_label.setStyleSheet(
                "color:green; font-weight:bold; padding:4px;")

    def _save(self):
        from data.repositories import ActualRepo, LotSampleRepo
        saved = 0
        for ri in range(self.table.rowCount()):
            actual_id = int(self.table.item(ri, 0).text())
            try:
                sample_qty = int(self.table.item(ri, 6).text())
                reject_qty = int(self.table.item(ri, 7).text())
            except (ValueError, AttributeError):
                continue
            if sample_qty < 0 or reject_qty < 0:
                continue

            actual = next((a for a in self._actual_rows
                           if a["actual_id"] == actual_id), None)
            if not actual:
                continue

            note_item = self.table.item(ri, 9)
            note = note_item.text() if note_item else None

            # Delete existing, re-insert with latest values
            for s in LotSampleRepo.for_actual(actual_id):
                LotSampleRepo.delete(s["sample_id"])

            if sample_qty > 0 or reject_qty > 0:
                LotSampleRepo.insert({
                    "actual_id":   actual_id,
                    "entity_type": actual["entity_type"],
                    "entity_code": actual["entity_code"],
                    "lot_number":  actual.get("lot_number") or "",
                    "so_number":   actual.get("so_number") or "",
                    "sku_code":    actual.get("sku_code") or "",
                    "line_item":   actual.get("line_item") or "",
                    "sample_qty":  sample_qty,
                    "reject_qty":  reject_qty,
                    "note":        note or None,
                })
                saved += 1

        QMessageBox.information(
            self, "Saved",
            f"QC quantities saved for {saved} lot(s).")
        self._load()


# ════════════════════════════════════════════════════════════════════════════
#  INVENTORY TAB
# ════════════════════════════════════════════════════════════════════════════

class InventoryTab(QWidget):
    """
    재고 관리 탭.
    - 재고 Excel 업로드 / 템플릿 다운로드
    - 재고 목록 (FEFO 정렬, 잔여 수량 표시)
    - SO별 재고 배정: FEFO 자동 제안 → 플래너 컨펌 or 수동 변경
    """
    _EDITABLE_INV_COLS = {3, 6, 7, 8}   # Qty Available, Prod Date, Expiry Date, Status
    _READONLY_INV_COLS = {0, 1, 2, 4, 5}

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._inv_edit_mode = False
        self._inv_changed_cells: set = set()
        self._inv_loading = False
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── Toolbar ──
        bar = QHBoxLayout()
        btn_upload = QPushButton("📤 Upload Inventory")
        btn_tmpl   = QPushButton("⬇ Template")
        btn_alloc  = QPushButton("🔗 Allocate to SO")
        btn_mb51   = QPushButton("📥 Upload MB51")
        btn_mb51.setStyleSheet(
            "background:#1d4ed8; color:white; font-weight:bold; "
            "border:none; border-radius:4px; padding:4px 12px;")
        btn_upload.clicked.connect(self._upload)
        btn_tmpl.clicked.connect(self._template)
        btn_alloc.clicked.connect(self._open_allocate)
        btn_mb51.clicked.connect(self._upload_mb51)
        for b in (btn_upload, btn_tmpl, btn_alloc, btn_mb51):
            bar.addWidget(b)

        bar.addWidget(QLabel("  Filter SKU:"))
        self.sku_filter = QLineEdit()
        self.sku_filter.setPlaceholderText("SKU code…")
        self.sku_filter.textChanged.connect(self.refresh)
        bar.addWidget(self.sku_filter)

        bar.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["ALL", "AVAILABLE", "ALLOCATED", "CONSUMED", "BLOCKED"])
        self.status_filter.currentTextChanged.connect(self.refresh)
        bar.addWidget(self.status_filter)

        self._btn_inv_edit = QPushButton("✏ Edit Mode")
        self._btn_inv_edit.clicked.connect(self._toggle_inv_edit_mode)
        self._btn_inv_save = QPushButton("💾 Save Changes")
        self._btn_inv_save.clicked.connect(self._save_inv_changes)
        self._btn_inv_save.setEnabled(False)
        bar.addWidget(self._btn_inv_edit)
        bar.addWidget(self._btn_inv_save)
        bar.addStretch()
        layout.addLayout(bar)

        # ── Summary ──
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("padding:4px; font-size:12px;")
        layout.addWidget(self.summary_label)

        self._inv_status = QLabel()
        self._inv_status.setWordWrap(True)
        layout.addWidget(self._inv_status)

        # ── Inventory table ──
        inv_grp = QGroupBox("Inventory Lots (FEFO order)")
        ig_l = QVBoxLayout(inv_grp)
        self.inv_table = QTableWidget()
        self.inv_table.setColumnCount(9)
        self.inv_table.setHorizontalHeaderLabels([
            "Inv ID", "SKU", "LOT Number", "Qty Available",
            "Qty Allocated", "Qty Remaining",
            "Production Date", "Expiry Date", "Status"
        ])
        self.inv_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.inv_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.inv_table.setAlternatingRowColors(True)
        self.inv_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.inv_table.itemChanged.connect(self._on_inv_cell_changed)
        self.inv_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.inv_table.customContextMenuRequested.connect(self._inv_context_menu)
        ig_l.addWidget(self.inv_table)
        layout.addWidget(inv_grp, stretch=1)

        # ── Allocation history table ──
        alloc_grp = QGroupBox("SO Allocation History")
        ag_l = QVBoxLayout(alloc_grp)
        self.alloc_table = QTableWidget()
        self.alloc_table.setColumnCount(8)
        self.alloc_table.setHorizontalHeaderLabels([
            "Alloc ID", "SO", "SKU", "Line",
            "LOT", "Qty Allocated", "Allocated At", "Note"
        ])
        self.alloc_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.alloc_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.alloc_table.setAlternatingRowColors(True)

        btn_dealloc = QPushButton("🗑 Remove Selected Allocation")
        btn_dealloc.clicked.connect(self._deallocate)
        ag_l.addWidget(self.alloc_table)
        ag_l.addWidget(btn_dealloc)
        layout.addWidget(alloc_grp, stretch=1)

        self.refresh()

    def refresh(self):
        self._load_inventory()
        self._load_allocations()

    def _load_inventory(self):
        self._inv_loading = True
        self._inv_changed_cells.clear()
        self._btn_inv_save.setEnabled(False)
        from data.repositories import InventoryRepo
        sku_f    = self.sku_filter.text().strip()
        status_f = self.status_filter.currentText()
        rows = InventoryRepo.all(
            sku_code=sku_f if sku_f else None,
            status=status_f if status_f != "ALL" else None)

        total_avail = sum(r["qty_available"] for r in rows)
        total_rem   = sum(r["qty_remaining"]  for r in rows)
        self.summary_label.setText(
            f"Total lots: {len(rows)}  |  "
            f"Total available: {total_avail:,}  |  "
            f"Total remaining: {total_rem:,}")

        STATUS_COLORS = {
            "AVAILABLE": QColor("#d4f0c0"),
            "ALLOCATED": QColor("#ffe0a0"),
            "CONSUMED":  QColor("#d0d0d0"),
            "EXPIRED":   QColor("#ffcccc"),
            "BLOCKED":   QColor("#f87171"),
        }
        inv_hdr = self.inv_table.horizontalHeader()
        inv_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.inv_table.setUpdatesEnabled(False)
        self.inv_table.setRowCount(len(rows))
        for ri, r in enumerate(rows):
            vals = [
                r["inv_id"], r["sku_code"], r["lot_number"],
                r["qty_available"], r["qty_allocated"], r["qty_remaining"],
                r.get("production_date") or "—",
                r.get("expiry_date") or "—",
                r["status"],
            ]
            bg = STATUS_COLORS.get(r["status"], QColor("white"))
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                if ci in self._READONLY_INV_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if ci == 0:
                    item.setData(Qt.ItemDataRole.UserRole, r)
                if ci == 5 and r["qty_remaining"] == 0:
                    item.setBackground(QBrush(QColor("#d0d0d0")))
                elif ci == 8:
                    item.setBackground(QBrush(bg))
                self.inv_table.setItem(ri, ci, item)
        self.inv_table.setUpdatesEnabled(True)
        self.inv_table.resizeColumnsToContents()
        inv_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._inv_loading = False
        self._inv_status.setText("")

    def _toggle_inv_edit_mode(self):
        self._inv_edit_mode = not self._inv_edit_mode
        if self._inv_edit_mode:
            self._btn_inv_edit.setText("🔒 Exit Edit Mode")
            self._btn_inv_edit.setStyleSheet(
                "background:#e65100; color:white; font-weight:bold;")
            self.inv_table.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked |
                QAbstractItemView.EditTrigger.EditKeyPressed)
            self._inv_status.setText(
                "✏ Edit mode — Qty Available, Production Date, Expiry Date, Status are editable")
            self._inv_status.setStyleSheet(
                "color:#7a5800; background:#fff9c4; padding:4px; border-radius:4px;")
        else:
            self._btn_inv_edit.setText("✏ Edit Mode")
            self._btn_inv_edit.setStyleSheet("")
            self.inv_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            if not self._inv_changed_cells:
                self._inv_status.setText("")
                self._inv_status.setStyleSheet("")

    def _on_inv_cell_changed(self, item):
        if self._inv_loading:
            return
        if item.column() not in self._EDITABLE_INV_COLS:
            return
        item.setBackground(QBrush(QColor("#fff9c4")))
        self._inv_changed_cells.add((item.row(), item.column()))
        n = len({r for r, _ in self._inv_changed_cells})
        self._inv_status.setText(f"✏ {n} row(s) modified — unsaved")
        self._inv_status.setStyleSheet(
            "color:#7a5800; background:#fff9c4; padding:4px; border-radius:4px;")
        self._btn_inv_save.setEnabled(True)

    def _save_inv_changes(self):
        if not self._inv_changed_cells:
            return
        from data.repositories import InventoryRepo
        changed_rows = {r for r, _ in self._inv_changed_cells}
        saved, errors = 0, []
        for ri in sorted(changed_rows):
            try:
                orig = self.inv_table.item(ri, 0).data(Qt.ItemDataRole.UserRole)
                if not orig:
                    continue
                qty_text     = self.inv_table.item(ri, 3).text().strip()
                prod_date    = self.inv_table.item(ri, 6).text().strip()
                expiry_date  = self.inv_table.item(ri, 7).text().strip()
                status       = self.inv_table.item(ri, 8).text().strip().upper()

                qty = int(qty_text) if qty_text else orig["qty_available"]
                if qty < 0:
                    errors.append(f"Row {ri+1}: Qty Available must be >= 0")
                    continue
                valid_statuses = ("AVAILABLE", "ALLOCATED", "CONSUMED", "EXPIRED")
                if status not in valid_statuses:
                    errors.append(f"Row {ri+1}: Status must be one of {valid_statuses}")
                    continue

                prod_date   = None if prod_date in ("—", "") else prod_date
                expiry_date = None if expiry_date in ("—", "") else expiry_date

                updated = dict(orig)
                updated.update({
                    "qty_available":   qty,
                    "production_date": prod_date,
                    "expiry_date":     expiry_date,
                    "status":          status,
                })
                InventoryRepo.upsert(updated)
                saved += 1
            except Exception as e:
                errors.append(f"Row {ri + 1}: {e}")
        self._inv_changed_cells.clear()
        self._btn_inv_save.setEnabled(False)
        if errors:
            QMessageBox.warning(self, "Save Errors", "\n".join(errors))
        self._inv_status.setText(f"✅ Saved {saved} row(s)")
        self._inv_status.setStyleSheet("color:green; padding:4px;")
        self._load_inventory()

    def _load_allocations(self):
        from data.repositories import AllocationRepo
        rows = AllocationRepo.all_allocations()
        alloc_hdr = self.alloc_table.horizontalHeader()
        alloc_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.alloc_table.setUpdatesEnabled(False)
        self.alloc_table.setRowCount(len(rows))
        for ri, r in enumerate(rows):
            for ci, v in enumerate([
                r["alloc_id"], r["so_number"], r["sku_code"],
                r["line_item"], r["lot_number"], r["qty_allocated"],
                r["allocated_at"][:16] if r.get("allocated_at") else "—",
                r.get("note") or ""
            ]):
                self.alloc_table.setItem(ri, ci, QTableWidgetItem(str(v)))
        self.alloc_table.setUpdatesEnabled(True)
        self.alloc_table.resizeColumnsToContents()
        alloc_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

    def _upload(self):
        from utils.excel_io import upload_inventory, parse_inventory_preview
        from ui.master_tab import UploadPreviewDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "Upload Inventory Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        ok, err, headers, rows = parse_inventory_preview(path)
        if not ok:
            QMessageBox.warning(self, "Parse Error", err)
            return
        dlg = UploadPreviewDialog("Inventory — Upload Preview", headers, rows, self)
        if not dlg.exec() or not dlg._confirmed:
            return
        ok, msg = upload_inventory(path)
        (QMessageBox.information if ok else QMessageBox.warning)(self, "Upload", msg)
        if ok:
            self.refresh()

    def _template(self):
        from utils.excel_io import download_inventory_template
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Template", "Inventory_template.xlsx", "Excel (*.xlsx)")
        if path:
            ok, msg = download_inventory_template(path)
            (QMessageBox.information if ok else QMessageBox.warning)(
                self, "Template", msg)

    def _open_allocate(self):
        dlg = InventoryAllocationDialog(self)
        if dlg.exec():
            self.refresh()
            if self.main_window:
                self.main_window.notify("Inventory allocation saved.")

    def _deallocate(self):
        row = self.alloc_table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Deallocate",
                                    "Select an allocation row first.")
            return
        alloc_id = int(self.alloc_table.item(row, 0).text())
        if QMessageBox.question(
            self, "Remove Allocation",
            f"Remove allocation #{alloc_id}?"
        ) == QMessageBox.StandardButton.Yes:
            from data.repositories import AllocationRepo
            AllocationRepo.deallocate(alloc_id)
            self.refresh()

    def _inv_context_menu(self, pos):
        row = self.inv_table.rowAt(pos.y())
        if row < 0:
            return
        item = self.inv_table.item(row, 0)
        if not item:
            return
        inv_data = item.data(Qt.ItemDataRole.UserRole)
        if not inv_data:
            return
        inv_id = inv_data["inv_id"]
        status = inv_data["status"]
        lot    = inv_data["lot_number"]
        sku    = inv_data["sku_code"]

        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        if status != "BLOCKED":
            act_block = menu.addAction(f"🔒 Block LOT {lot}")
        else:
            act_block = None
        if status == "BLOCKED":
            act_unblock = menu.addAction(f"🔓 Unblock LOT {lot}")
        else:
            act_unblock = None
        menu.addSeparator()
        act_realloc = menu.addAction(f"♻ Re-allocate SKU {sku}")

        chosen = menu.exec(self.inv_table.viewport().mapToGlobal(pos))
        if chosen is None:
            return

        from data.repositories import InventoryRepo, AllocationRepo
        if chosen == act_block:
            if QMessageBox.question(
                self, "Block LOT",
                f"Block LOT {lot}?\n\nAll SO allocations for this lot will be removed."
            ) != QMessageBox.StandardButton.Yes:
                return
            AllocationRepo.deallocate_lot(inv_id)
            InventoryRepo.block_lot(inv_id)
            # Re-allocate the affected SKU from remaining available lots
            AllocationRepo.reallocate_sku(sku)
            self.refresh()
            if self.main_window:
                self.main_window.notify(f"LOT {lot} blocked and deallocated.")

        elif chosen == act_unblock:
            InventoryRepo.unblock_lot(inv_id)
            AllocationRepo.reallocate_sku(sku)
            self.refresh()
            if self.main_window:
                self.main_window.notify(f"LOT {lot} unblocked and re-allocated.")

        elif chosen == act_realloc:
            stats = AllocationRepo.reallocate_sku(sku)
            self.refresh()
            short = stats.get("sos_short", [])
            msg = (f"Re-allocation complete for {sku}.\n"
                   f"Allocated: {stats.get('allocated_total', 0):,} units.\n")
            if short:
                msg += f"Short SOs: {len(short)} (need additional production)."
            QMessageBox.information(self, "Re-allocate", msg)

    def _upload_mb51(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Upload MB51 Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        from utils.excel_io import parse_mb51
        ok, msg, rows = parse_mb51(path)
        if not ok:
            QMessageBox.warning(self, "MB51 Parse Error", msg)
            return
        if not rows:
            QMessageBox.information(self, "MB51", "No data rows found.")
            return
        dlg = MB51UploadDialog(rows, self)
        if dlg.exec():
            self.refresh()
            if self.main_window:
                self.main_window.notify(
                    f"MB51 processed: {dlg.result.get('new_docs', 0)} new documents, "
                    f"{len(dlg.result.get('affected_skus', []))} SKUs updated.")


# ─── Inventory Allocation Dialog ──────────────────────────────────────────────

class InventoryAllocationDialog(QDialog):
    """
    FEFO 자동 제안 후 플래너 컨펌 / 수동 변경 다이얼로그.
    1. SO 선택 → 필요 수량 표시
    2. FEFO 제안 자동 생성
    3. 플래너가 수량 조정 가능
    4. 확인 → AllocationRepo.confirm_fefo_suggestion()
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Allocate Inventory to SO")
        self.resize(820, 580)
        self._suggestion: list = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── SO 선택 ──
        so_grp = QGroupBox("1. Select SO-LineItem")
        sg_l   = QFormLayout(so_grp)

        self.so_combo = QComboBox()
        self._reload_sos()
        sg_l.addRow("SO / SKU / Line:", self.so_combo)

        self.so_info = QLabel()
        self.so_info.setStyleSheet("color:gray; font-size:11px;")
        sg_l.addRow("", self.so_info)

        btn_suggest = QPushButton("🔍 Generate FEFO Suggestion")
        btn_suggest.clicked.connect(self._suggest)
        sg_l.addRow("", btn_suggest)
        layout.addWidget(so_grp)

        # ── FEFO 제안 테이블 ──
        sug_grp = QGroupBox("2. FEFO Suggestion (editable)")
        sug_l   = QVBoxLayout(sug_grp)

        self.sug_table = QTableWidget()
        self.sug_table.setColumnCount(6)
        self.sug_table.setHorizontalHeaderLabels([
            "Inv ID", "LOT Number", "Expiry Date",
            "Lot Remaining", "Allocate Qty ✏", "Note ✏"
        ])
        self.sug_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.sug_table.setAlternatingRowColors(True)
        sug_l.addWidget(self.sug_table)

        self.sug_total_label = QLabel()
        self.sug_total_label.setStyleSheet("font-size:12px; padding:4px;")
        sug_l.addWidget(self.sug_total_label)

        # Manual override — add any lot
        add_bar = QHBoxLayout()
        add_bar.addWidget(QLabel("Add lot manually:"))
        self.manual_inv_combo = QComboBox()
        self.manual_inv_combo.setMinimumWidth(200)
        add_bar.addWidget(self.manual_inv_combo)
        self.manual_qty = QSpinBox()
        self.manual_qty.setRange(1, 999999)
        add_bar.addWidget(self.manual_qty)
        btn_add_manual = QPushButton("➕ Add")
        btn_add_manual.clicked.connect(self._add_manual_lot)
        add_bar.addWidget(btn_add_manual)
        add_bar.addStretch()
        sug_l.addLayout(add_bar)
        layout.addWidget(sug_grp, stretch=1)

        # ── Buttons ──
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Ok).setText(
            "✅ Confirm Allocation")
        btns.accepted.connect(self._confirm)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _reload_sos(self):
        from data.repositories import SORepo
        self.so_combo.clear()
        sos = SORepo.all("OPEN")
        for so in sos:
            label = (f"{so['so_number']} / {so['sku_code']} / "
                     f"{so['line_item']}  (due: {so['due_date']})")
            self.so_combo.addItem(
                label,
                (so["so_number"], so["sku_code"], so["line_item"]))

    def _current_so(self):
        return self.so_combo.currentData()

    def _suggest(self):
        so_key = self._current_so()
        if not so_key:
            return
        so_no, sku, li = so_key
        from data.repositories import AllocationRepo, InventoryRepo, SORepo
        so = SORepo.get(so_no, sku, li)
        if not so:
            return

        prod_needed = AllocationRepo.production_needed(so_no, sku, li)
        already_alloc = AllocationRepo.total_allocated(so_no, sku, li)
        total_inv     = InventoryRepo.total_available(sku)

        self.so_info.setText(
            f"SO qty: {so['qty']}  |  Already allocated: {already_alloc}  |  "
            f"Prod needed: {prod_needed}  |  Inventory available: {total_inv}")

        # How much can we cover from inventory?
        cover = min(prod_needed, total_inv)
        if cover <= 0:
            QMessageBox.information(
                self, "No Inventory",
                f"No available inventory for {sku}.")
            return

        suggestion = InventoryRepo.fefo_suggestion(sku, cover)
        self._suggestion = suggestion
        self._load_suggestion_table(suggestion)

        # Populate manual override combo
        self.manual_inv_combo.clear()
        for lot in InventoryRepo.available_for_sku(sku):
            self.manual_inv_combo.addItem(
                f"{lot['lot_number']}  (rem: {lot['qty_remaining']}, "
                f"exp: {lot.get('expiry_date') or '—'})",
                lot)

    def _load_suggestion_table(self, suggestion):
        self.sug_table.setRowCount(len(suggestion))
        for ri, s in enumerate(suggestion):
            for ci, v in enumerate([
                s["inv_id"], s["lot_number"],
                s.get("expiry_date") or "—",
                s["qty_remaining"],
                s["qty_to_allocate"],  # editable
                ""                     # note editable
            ]):
                item = QTableWidgetItem(str(v))
                if ci == 4:
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(QBrush(QColor("#fffbe6")))
                elif ci == 5:
                    item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                else:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.sug_table.setItem(ri, ci, item)
        self._update_total_label()

    def _update_total_label(self):
        total = 0
        for ri in range(self.sug_table.rowCount()):
            try:
                total += int(self.sug_table.item(ri, 4).text())
            except (ValueError, AttributeError):
                pass
        so_key = self._current_so()
        if so_key:
            from data.repositories import AllocationRepo
            needed = AllocationRepo.production_needed(*so_key)
            color = "green" if total >= needed else "orange"
            self.sug_total_label.setText(
                f"Total to allocate: <b style='color:{color}'>{total}</b>  "
                f"/ Prod needed: {needed}")
        else:
            self.sug_total_label.setText(f"Total to allocate: {total}")

    def _add_manual_lot(self):
        lot_data = self.manual_inv_combo.currentData()
        if not lot_data:
            return
        qty = self.manual_qty.value()
        # Check if already in table
        for ri in range(self.sug_table.rowCount()):
            if self.sug_table.item(ri, 0).text() == str(lot_data["inv_id"]):
                # Update existing row qty
                existing = int(self.sug_table.item(ri, 4).text() or 0)
                self.sug_table.item(ri, 4).setText(str(existing + qty))
                self._update_total_label()
                return
        # Add new row
        ri = self.sug_table.rowCount()
        self.sug_table.setRowCount(ri + 1)
        for ci, v in enumerate([
            lot_data["inv_id"], lot_data["lot_number"],
            lot_data.get("expiry_date") or "—",
            lot_data["qty_remaining"], qty, ""
        ]):
            item = QTableWidgetItem(str(v))
            if ci in (4, 5):
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                if ci == 4:
                    item.setBackground(QBrush(QColor("#fffbe6")))
            else:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.sug_table.setItem(ri, ci, item)
        self._update_total_label()

    def _confirm(self):
        so_key = self._current_so()
        if not so_key:
            QMessageBox.warning(self, "Error", "Select an SO first.")
            return
        so_no, sku, li = so_key

        allocations = []
        for ri in range(self.sug_table.rowCount()):
            try:
                inv_id = int(self.sug_table.item(ri, 0).text())
                lot    = self.sug_table.item(ri, 1).text()
                qty    = int(self.sug_table.item(ri, 4).text())
                note_i = self.sug_table.item(ri, 5)
                note   = note_i.text() if note_i else None
            except (ValueError, AttributeError):
                continue
            if qty <= 0:
                continue
            # Validate against lot remaining
            from data.repositories import InventoryRepo
            lot_data = InventoryRepo.get(inv_id)
            if not lot_data:
                continue
            if qty > lot_data["qty_remaining"]:
                QMessageBox.warning(
                    self, "Qty Error",
                    f"LOT {lot}: allocate qty {qty} exceeds "
                    f"remaining {lot_data['qty_remaining']}.")
                return
            allocations.append({
                "inv_id":          inv_id,
                "lot_number":      lot,
                "qty_to_allocate": qty,
                "note":            note,
            })

        if not allocations:
            QMessageBox.warning(self, "Error",
                                "No valid allocation rows.")
            return

        from data.repositories import AllocationRepo
        AllocationRepo.confirm_fefo_suggestion(so_no, sku, li, allocations)
        QMessageBox.information(
            self, "Allocated",
            f"Allocated {len(allocations)} lot(s) to "
            f"{so_no}/{sku}/{li}.")
        self.accept()


# ════════════════════════════════════════════════════════════════════════════
#  RELEASE REPORT TAB
# ════════════════════════════════════════════════════════════════════════════

class ReleaseReportTab(QWidget):
    """
    SO / SKU / LineItem별 예상 릴리즈 일정 리포트.

    계산 로직:
      - 생산계획의 마지막 공정(is_final_seq=1) 완료 예정 Shift 기준
      - Release Date = 마지막 공정 완료일 + SKU.post_lead_days
      - Due Date와 비교해서 ON TIME / AT RISK / LATE 상태 표시

    컬럼:
      SO | SKU | Line | SO Qty | Prod Needed | Planned Qty |
      Last Process Date | Last Shift | Release Date | Due Date |
      Days to Due | Status | Note
    """

    STATUS_ON_TIME = "ON TIME"
    STATUS_AT_RISK = "AT RISK"   # release within 3 days of due
    STATUS_LATE    = "LATE"
    STATUS_NO_PLAN = "NOT PLANNED"

    AT_RISK_DAYS = 3             # configurable per instance

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._rows: list = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── Toolbar ──
        bar = QHBoxLayout()

        bar.addWidget(QLabel("Filter:"))
        self.search = QLineEdit()
        self.search.setPlaceholderText("SO / SKU / line item…")
        self.search.textChanged.connect(self._apply_filter)
        bar.addWidget(self.search)

        bar.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(
            ["ALL", "LATE", "AT RISK", "ON TIME", "NOT PLANNED"])
        self.status_filter.currentTextChanged.connect(self._apply_filter)
        bar.addWidget(self.status_filter)

        bar.addWidget(QLabel("Due within (days):"))
        self.due_days_spin = QSpinBox()
        self.due_days_spin.setRange(0, 365)
        self.due_days_spin.setValue(0)
        self.due_days_spin.setSpecialValueText("All")
        self.due_days_spin.valueChanged.connect(self._apply_filter)
        bar.addWidget(self.due_days_spin)

        bar.addStretch()

        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.refresh)
        bar.addWidget(btn_refresh)

        btn_export = QPushButton("📥 Export")
        btn_export.clicked.connect(self._export)
        bar.addWidget(btn_export)

        layout.addLayout(bar)

        # ── Summary strip ──
        self.summary = QLabel()
        self.summary.setStyleSheet(
            "padding:5px 8px; font-size:12px; "
            "background:var(--surface); border-bottom:1px solid #ddd;")
        layout.addWidget(self.summary)

        # ── Table ──
        self.table = QTableWidget()
        self.table.setColumnCount(13)
        self.table.setHorizontalHeaderLabels([
            "SO Number", "SKU Code", "Line Item",
            "SO Qty", "Prod Needed", "Planned Qty",
            "Last Process Date", "Shift",
            "Release Date", "Due Date",
            "Days to Due", "Status", "Note"
        ])
        _hdr = self.table.horizontalHeader()
        _hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _hdr.setStretchLastSection(True)
        for i, w in enumerate([120, 90, 55, 60, 80, 80, 110, 45, 100, 100, 75, 80, 120]):
            _hdr.resizeSection(i, w)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table, stretch=1)

        self.refresh()

    # ── Data computation ─────────────────────────────────────────────────────

    def refresh(self):
        self._rows = self._compute_rows()
        self._apply_filter()

    def _compute_rows(self) -> list:
        from datetime import date as dt, datetime as dtt, timedelta
        from data.repositories import (
            SORepo, PlanRepo, SKURepo, ActualRepo, AllocationRepo
        )

        today = dt.today()
        sos   = SORepo.all()
        skus  = {s["sku_code"]: s for s in SKURepo.all()}
        # Batch queries — avoids N+1 (4 queries regardless of SO count)
        alloc_map     = AllocationRepo.allocation_summary_for_open_sos()
        actual_map    = ActualRepo.actual_qty_bulk()
        planned_map   = PlanRepo.planned_qty_bulk()
        last_plan_map = PlanRepo.last_plan_info_bulk()
        rows  = []

        for so in sos:
            so_no = so["so_number"]
            sku_c = so["sku_code"]
            li    = so["line_item"]
            key   = (so_no, sku_c, li)
            sku   = skus.get(sku_c, {})
            post_lead = int(sku.get("post_lead_days") or 0)

            # Quantities (computed inline — no per-SO DB calls)
            allocated   = alloc_map.get(key, 0)
            actual_qty  = actual_map.get(key, 0)
            prod_needed = max(0, so["qty"] - allocated - actual_qty)
            planned_qty = planned_map.get(key, 0)

            # Last plan slot (last_plan_info_bulk returns max date+shift across all plans;
            # backward scheduling places the final step latest, so this matches FINAL intent)
            last_info = last_plan_map.get(key)
            if last_info:
                last_date_str = last_info[0]
                last_shift    = last_info[1]
                last_dt       = dtt.strptime(last_date_str, "%Y-%m-%d").date()
                from utils.workdays import add_workdays as _awdays
                release_dt    = _awdays(last_dt, post_lead)
                release_str   = release_dt.strftime("%Y-%m-%d")
            else:
                last_date_str = "—"
                last_shift    = "—"
                release_dt    = None
                release_str   = "—"

            due_dt  = dtt.strptime(so["due_date"], "%Y-%m-%d").date()

            # Status
            if release_dt is None:
                status     = self.STATUS_NO_PLAN
                days_to_due = ""
            else:
                days_to_due = (due_dt - release_dt).days
                if release_dt > due_dt:
                    status = self.STATUS_LATE
                elif days_to_due <= self.AT_RISK_DAYS:
                    status = self.STATUS_AT_RISK
                else:
                    status = self.STATUS_ON_TIME

            # Note: flag closed/hold
            note = ""
            if so["status"] == "HOLD":
                note = "⏸ HOLD"
            elif so["status"] == "CLOSED":
                note = "✔ CLOSED"
            elif prod_needed == 0 and planned_qty == 0:
                note = "Covered by inventory"

            rows.append({
                "so_number":    so_no,
                "sku_code":     sku_c,
                "line_item":    li,
                "so_qty":       so["qty"],
                "prod_needed":  prod_needed,
                "planned_qty":  planned_qty,
                "last_date":    last_date_str,
                "last_shift":   str(last_shift),
                "release_date": release_str,
                "due_date":     so["due_date"],
                "days_to_due":  days_to_due,
                "status":       status,
                "note":         note,
                "so_status":    so["status"],
                "release_dt":   release_dt,
                "due_dt":       due_dt,
            })

        return rows

    # ── Filter & render ──────────────────────────────────────────────────────

    def _apply_filter(self):
        from datetime import date as dt, timedelta
        search     = self.search.text().lower()
        status_f   = self.status_filter.currentText()
        due_days   = self.due_days_spin.value()
        today      = dt.today()

        filtered = self._rows
        if search:
            filtered = [r for r in filtered if search in (
                r["so_number"] + r["sku_code"] + r["line_item"]).lower()]
        if status_f != "ALL":
            filtered = [r for r in filtered if r["status"] == status_f]
        if due_days > 0:
            cutoff = today + timedelta(days=due_days)
            filtered = [r for r in filtered
                        if r["due_dt"] <= cutoff]

        self._render(filtered)
        self._update_summary(filtered)
        if not filtered:
            self.summary.setText(
                "  <span style='color:#9aa1b3;'>No records match the current filter.</span>")

    def _render(self, rows: list):
        STATUS_BG = {
            self.STATUS_LATE:     QColor("#ffcccc"),
            self.STATUS_AT_RISK:  QColor("#fff0cc"),
            self.STATUS_ON_TIME:  QColor("#d4f0c0"),
            self.STATUS_NO_PLAN:  QColor("#e8e8e8"),
        }
        STATUS_FG = {
            self.STATUS_LATE:     QColor("#cc0000"),
            self.STATUS_AT_RISK:  QColor("#996600"),
            self.STATUS_ON_TIME:  QColor("#2d7a2d"),
            self.STATUS_NO_PLAN:  QColor("#666666"),
        }
        bold_font = QFont(); bold_font.setBold(True)
        gray_fg   = QBrush(QColor("#999999"))
        late_bg   = QBrush(QColor("#ffcccc"))
        at_risk_bg = QBrush(QColor("#fff0cc"))

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        self.table.setRowCount(len(rows))

        for ri, r in enumerate(rows):
            bg = QBrush(STATUS_BG.get(r["status"], QColor("white")))
            fg = QBrush(STATUS_FG.get(r["status"], QColor("black")))
            is_dim = r["so_status"] in ("CLOSED", "HOLD")

            vals = [
                r["so_number"], r["sku_code"], r["line_item"],
                r["so_qty"], r["prod_needed"], r["planned_qty"],
                r["last_date"], r["last_shift"],
                r["release_date"], r["due_date"],
                str(r["days_to_due"]) if r["days_to_due"] != "" else "—",
                r["status"], r["note"],
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if ci == 11:
                    item.setBackground(bg)
                    item.setForeground(fg)
                    item.setFont(bold_font)
                elif ci == 10 and r["days_to_due"] != "":
                    try:
                        d = int(r["days_to_due"])
                        if d < 0:
                            item.setBackground(late_bg)
                        elif d <= self.AT_RISK_DAYS:
                            item.setBackground(at_risk_bg)
                    except ValueError:
                        pass
                elif ci == 8 and r["status"] == self.STATUS_LATE:
                    item.setBackground(late_bg)
                    item.setFont(bold_font)
                if is_dim and ci != 11:
                    item.setForeground(gray_fg)
                self.table.setItem(ri, ci, item)

        self.table.setUpdatesEnabled(True)
        hdr.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setSortingEnabled(True)

    def _update_summary(self, rows: list):
        total    = len(rows)
        late     = sum(1 for r in rows if r["status"] == self.STATUS_LATE)
        at_risk  = sum(1 for r in rows if r["status"] == self.STATUS_AT_RISK)
        on_time  = sum(1 for r in rows if r["status"] == self.STATUS_ON_TIME)
        no_plan  = sum(1 for r in rows if r["status"] == self.STATUS_NO_PLAN)

        parts = [f"Total: <b>{total}</b>"]
        if late:
            parts.append(
                f"<span style='color:#cc0000'>🔴 LATE: <b>{late}</b></span>")
        if at_risk:
            parts.append(
                f"<span style='color:#996600'>🟡 AT RISK: <b>{at_risk}</b></span>")
        if on_time:
            parts.append(
                f"<span style='color:#2d7a2d'>🟢 ON TIME: <b>{on_time}</b></span>")
        if no_plan:
            parts.append(
                f"<span style='color:#666'>⚪ NOT PLANNED: <b>{no_plan}</b></span>")

        self.summary.setText("  |  ".join(parts))

    # ── Export ───────────────────────────────────────────────────────────────

    def _export(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from PyQt6.QtWidgets import QFileDialog
            path, _ = QFileDialog.getSaveFileName(
                self, "Export Release Report",
                "Release_Report.xlsx", "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook(); ws = wb.active
            ws.title = "Release Report"

            HDR_FILL = PatternFill("solid", fgColor="4472C4")
            HDR_FONT = Font(color="FFFFFF", bold=True)
            STATUS_FILLS = {
                "LATE":        PatternFill("solid", fgColor="FFCCCC"),
                "AT RISK":     PatternFill("solid", fgColor="FFF0CC"),
                "ON TIME":     PatternFill("solid", fgColor="D4F0C0"),
                "NOT PLANNED": PatternFill("solid", fgColor="E8E8E8"),
            }

            headers = [
                "SO Number", "SKU Code", "Line Item",
                "SO Qty", "Prod Needed", "Planned Qty",
                "Last Process Date", "Shift",
                "Release Date", "Due Date",
                "Days to Due", "Status", "Note"
            ]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = HDR_FILL; c.font = HDR_FONT
                c.alignment = Alignment(horizontal="center")

            for ri, r in enumerate(self._rows, 2):
                vals = [
                    r["so_number"], r["sku_code"], r["line_item"],
                    r["so_qty"], r["prod_needed"], r["planned_qty"],
                    r["last_date"], r["last_shift"],
                    r["release_date"], r["due_date"],
                    r["days_to_due"] if r["days_to_due"] != "" else None,
                    r["status"], r["note"],
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    if r["status"] in STATUS_FILLS:
                        cell.fill = STATUS_FILLS[r["status"]]

            # Auto column width
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[col[0].column_letter].width = \
                    min(max_len + 2, 30)

            wb.save(path)
            QMessageBox.information(
                self, "Export", f"Saved to {path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", str(e))


# ════════════════════════════════════════════════════════════════════════════
#  FULFILLMENT STATUS TAB  (merged Release Report + Dispatch List)
# ════════════════════════════════════════════════════════════════════════════

class FulfillmentStatusTab(QWidget):
    """
    Per-SO fulfillment view: release-date adherence + Critical Ratio priority.

    CR = open shifts to due ÷ shifts needed (worst-case across all routing steps).
    Status priority: CRITICAL → AT RISK → ON TIME → COMPLETE → NOT PLANNED
    """

    STATUS_CRITICAL = "CRITICAL"
    STATUS_AT_RISK  = "AT RISK"
    STATUS_ON_TIME  = "ON TIME"
    STATUS_COMPLETE = "COMPLETE"
    STATUS_NO_PLAN  = "NOT PLANNED"
    STATUS_NO_ROUTE = "NO ROUTING"

    CR_CRITICAL  = 1.0
    CR_AT_RISK   = 1.5
    AT_RISK_DAYS = 3

    _COLS = [
        "Rank", "SO", "Customer", "SKU", "Line",
        "SO Qty", "Prod Needed", "Planned",
        "Prod. Complete", "Release Date", "Req. Due", "Committed",
        "Days to Due", "CR", "Status", "Note",
    ]

    _STATUS_COLORS = {
        "CRITICAL":    ("#FFEBEE", "#B71C1C"),
        "AT RISK":     ("#FFF8E1", "#E65100"),
        "ON TIME":     ("#E8F5E9", "#1B5E20"),
        "COMPLETE":    ("#F5F5F5", "#757575"),
        "NOT PLANNED": ("#F5F5F5", "#9E9E9E"),
        "NO ROUTING":  ("#F5F5F5", "#BDBDBD"),
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._rows: List[Dict] = []
        self._build_ui()
        from PyQt6.QtCore import QTimer
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(30_000)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(6)

        # ── Toolbar ──
        bar = QHBoxLayout()
        bar.setSpacing(8)

        bar.addWidget(QLabel("Search:"))
        self._search = QLineEdit()
        self._search.setPlaceholderText("SO / SKU / Customer…")
        self._search.setFixedWidth(180)
        self._search.textChanged.connect(self._apply_filter)
        bar.addWidget(self._search)

        bar.addWidget(QLabel("Status:"))
        self._status_combo = QComboBox()
        self._status_combo.addItems([
            "ALL", self.STATUS_CRITICAL, self.STATUS_AT_RISK,
            self.STATUS_ON_TIME, self.STATUS_COMPLETE,
            self.STATUS_NO_PLAN, self.STATUS_NO_ROUTE,
        ])
        self._status_combo.currentTextChanged.connect(self._apply_filter)
        bar.addWidget(self._status_combo)

        bar.addWidget(QLabel("Due within (days):"))
        self._due_spin = QSpinBox()
        self._due_spin.setRange(0, 365)
        self._due_spin.setValue(0)
        self._due_spin.setSpecialValueText("All")
        self._due_spin.valueChanged.connect(self._apply_filter)
        bar.addWidget(self._due_spin)

        bar.addStretch()

        self._summary_lbl = QLabel()
        self._summary_lbl.setStyleSheet("color:#555; font-size:11px;")
        bar.addWidget(self._summary_lbl)

        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.refresh)
        bar.addWidget(btn_refresh)

        btn_export = QPushButton("📥 Export")
        btn_export.clicked.connect(self._export)
        bar.addWidget(btn_export)

        layout.addLayout(bar)

        # ── Legend ──
        legend = QHBoxLayout()
        legend.setSpacing(16)
        for status, (_, fg) in self._STATUS_COLORS.items():
            dot = QLabel(f"● {status}")
            dot.setStyleSheet(f"color:{fg}; font-size:10px; font-weight:600;")
            legend.addWidget(dot)
        legend.addStretch()
        note_lbl = QLabel("CR = Open shifts to due ÷ Shifts needed (worst-case process)")
        note_lbl.setStyleSheet("color:#888; font-size:10px; font-style:italic;")
        legend.addWidget(note_lbl)
        layout.addLayout(legend)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("background:#dde3ed; border:none;")
        sep.setFixedHeight(1)
        layout.addWidget(sep)

        # ── Table ──
        self._table = QTableWidget()
        self._table.setColumnCount(len(self._COLS))
        self._table.setHorizontalHeaderLabels(self._COLS)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(False)
        self._table.setSortingEnabled(True)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        for i, w in enumerate([45, 110, 110, 80, 45, 60, 80, 60, 90, 90, 90, 90, 65, 50, 75, 110]):
            hdr.resizeSection(i, w)
        self._table.verticalHeader().setDefaultSectionSize(26)
        self._table.setStyleSheet(
            "QTableWidget { border:none; font-size:11px; }"
            "QHeaderView::section { background:#f0f2f7; font-weight:600;"
            " padding:4px; border:none; border-bottom:1px solid #dde3ed; }")
        layout.addWidget(self._table, stretch=1)

        self.refresh()

    # ── Data ─────────────────────────────────────────────────────────────────

    def refresh(self):
        self._rows = self._compute_rows()
        self._apply_filter()

    @staticmethod
    def _shift_h(shift: Dict) -> float:
        from datetime import datetime as _dt
        fmt = "%H:%M"
        t0 = _dt.strptime(shift["start_time"], fmt)
        t1 = _dt.strptime(shift["end_time"], fmt)
        h = (t1 - t0).seconds / 3600
        return h if h > 0 else 24 + h

    def _compute_rows(self) -> List[Dict]:
        from data.repositories import AllocationRepo, CalendarRepo
        from core.scheduler import calc_uph

        today     = date.today()
        today_str = today.strftime("%Y-%m-%d")

        sos = SORepo.all()
        if not sos:
            return []

        skus          = {s["sku_code"]: s for s in SKURepo.all()}
        alloc_map     = AllocationRepo.allocation_summary_for_open_sos()
        actual_map    = ActualRepo.actual_qty_bulk()
        planned_map   = PlanRepo.planned_qty_bulk()
        last_plan_map = PlanRepo.last_plan_info_bulk()

        # Pre-load all routings (avoids N+1 per-SO DB calls)
        routing_map: Dict = {}
        for row in ProcessRoutingRepo.all():
            routing_map.setdefault((row["entity_type"], row["entity_code"]), []).append(row)
        for k in routing_map:
            routing_map[k].sort(key=lambda r: r["process_seq"])

        # CR infrastructure
        all_rooms   = RoomRepo.all()
        shifts      = ShiftRepo.all()
        avg_shift_h = (sum(self._shift_h(s) for s in shifts) / len(shifts)
                       if shifts else 12.0)
        n_shifts    = len(shifts) if shifts else 2

        process_rooms: Dict[str, list] = {}
        for rp in all_rooms:
            process_rooms.setdefault(rp["process_name"], []).append(rp)

        proc_room_set: Dict[str, set] = {
            p: {rp["room_code"] for rp in rps}
            for p, rps in process_rooms.items()
        }

        max_due = max(
            (so.get("committed_due_date") or so["due_date"] for so in sos),
            default=today_str)

        raw_slots  = CalendarRepo.get_open_slots(today_str, max_due)
        open_set   = {(s["cal_date"], s["room_code"], s["shift_no"])
                      for s in raw_slots if not s.get("is_hold", 0)}
        calendared = {(s["cal_date"], s["shift_no"]) for s in raw_slots}

        all_procs = list(process_rooms.keys())
        proc_cum: Dict[str, Dict[str, int]] = {}
        running = {p: 0 for p in all_procs}
        d     = today
        until = datetime.strptime(max_due, "%Y-%m-%d").date()
        while d <= until:
            ds   = d.strftime("%Y-%m-%d")
            is_wd = d.weekday() < 5
            for proc_name in all_procs:
                added = 0
                for sno_idx in range(1, n_shifts + 1):
                    if (ds, sno_idx) in calendared:
                        if any((ds, rc, sno_idx) in open_set
                               for rc in proc_room_set[proc_name]):
                            added += 1
                    elif is_wd:
                        added += 1
                running[proc_name] += added
            for proc_name in all_procs:
                proc_cum.setdefault(proc_name, {})[ds] = running[proc_name]
            d += timedelta(days=1)

        rows: List[Dict] = []
        for so in sos:
            so_no = so["so_number"]
            sku_c = so["sku_code"]
            li    = so["line_item"]
            key   = (so_no, sku_c, li)
            sku   = skus.get(sku_c, {})
            post_lead = int(sku.get("post_lead_days") or 0)
            uom       = int(sku.get("uom") or 1)

            allocated   = alloc_map.get(key, 0)
            actual_qty  = actual_map.get(key, 0)
            prod_needed = max(0, so["qty"] - allocated - actual_qty)
            planned_qty = planned_map.get(key, 0)

            # Production complete date = last plan slot; release adds post_lead
            last_info = last_plan_map.get(key)
            if last_info:
                from utils.workdays import add_workdays
                last_dt         = datetime.strptime(last_info[0], "%Y-%m-%d").date()
                prod_complete   = last_info[0]          # YYYY-MM-DD string
                release_dt      = add_workdays(last_dt, post_lead)
                release_str     = release_dt.strftime("%Y-%m-%d")
            else:
                prod_complete   = "—"
                release_dt      = None
                release_str     = "—"

            req_due_str       = so["due_date"]
            committed_due_str = so.get("committed_due_date") or ""
            eff_due_str       = committed_due_str or req_due_str
            due_dt  = datetime.strptime(req_due_str, "%Y-%m-%d").date()
            eff_due = datetime.strptime(eff_due_str, "%Y-%m-%d").date()
            days_to_due = (eff_due - today).days

            # CR computation — worst-case across all routing steps
            routing    = routing_map.get(("SKU", sku_c), [])
            min_cr     = None
            cr_display = "—"

            if prod_needed == 0:
                status = self.STATUS_COMPLETE
            elif not routing:
                status = self.STATUS_NO_ROUTE
            else:
                for step in routing:
                    proc_name = step["process_name"]
                    best_cap  = 0.0
                    for rp in process_rooms.get(proc_name, []):
                        hc = (int(rp.get("hc_fixed") or 1)
                              if rp.get("process_type") == "AUTO"
                              else int(rp.get("hc_max") or rp.get("hc_min") or 1))
                        cap = calc_uph(rp, hc) * avg_shift_h
                        if cap > best_cap:
                            best_cap = cap

                    rem_inner = prod_needed * uom
                    if rem_inner == 0:
                        step_cr: float | None = 999.0
                    elif best_cap <= 0:
                        step_cr = None
                    else:
                        shifts_needed = math.ceil(rem_inner / best_cap)
                        open_shifts   = proc_cum.get(proc_name, {}).get(eff_due_str, 0)
                        step_cr = (open_shifts / shifts_needed
                                   if shifts_needed > 0 else 999.0)

                    if step_cr is not None and (min_cr is None or step_cr < min_cr):
                        min_cr = step_cr

                if min_cr is None:
                    status = self.STATUS_NO_ROUTE
                else:
                    cr_display = f"{min_cr:.2f}"
                    if min_cr < self.CR_CRITICAL:
                        status = self.STATUS_CRITICAL
                    elif min_cr < self.CR_AT_RISK:
                        status = self.STATUS_AT_RISK
                    elif release_dt is None:
                        status = self.STATUS_NO_PLAN
                    elif release_dt > due_dt:
                        status = self.STATUS_AT_RISK
                    elif (due_dt - release_dt).days <= self.AT_RISK_DAYS:
                        status = self.STATUS_AT_RISK
                    else:
                        status = self.STATUS_ON_TIME

            note = ""
            if so["status"] == "HOLD":
                note = "⏸ HOLD"
            elif so["status"] == "CLOSED":
                note = "✔ CLOSED"
            elif prod_needed == 0 and planned_qty == 0:
                note = "Covered by inventory"

            rows.append({
                "rank":           "—",
                "so_number":      so_no,
                "customer":       so.get("customer_name", "") or "",
                "sku_code":       sku_c,
                "line_item":      li,
                "so_qty":         so["qty"],
                "prod_needed":    prod_needed,
                "planned_qty":    planned_qty,
                "prod_complete":  prod_complete,
                "release_date":   release_str,
                "req_due":        req_due_str,
                "committed_due":  committed_due_str or "—",
                "days_to_due":    days_to_due,
                "cr":             min_cr,
                "cr_display":     cr_display,
                "status":         status,
                "note":           note,
                "so_status":      so["status"],
                "due_dt":         due_dt,
            })

        _order = {
            self.STATUS_CRITICAL: 0, self.STATUS_AT_RISK:  1,
            self.STATUS_ON_TIME:  2, self.STATUS_COMPLETE: 3,
            self.STATUS_NO_PLAN:  4, self.STATUS_NO_ROUTE: 5,
        }
        rows.sort(key=lambda r: (
            _order.get(r["status"], 9),
            r["cr"] if r["cr"] is not None else 999.0,
            r["days_to_due"],
        ))
        rank = 1
        for r in rows:
            if r["status"] in (self.STATUS_CRITICAL, self.STATUS_AT_RISK, self.STATUS_ON_TIME):
                r["rank"] = rank
                rank += 1

        return rows

    # ── Filter & render ───────────────────────────────────────────────────────

    def _apply_filter(self):
        search   = self._search.text().strip().lower()
        status_f = self._status_combo.currentText()
        due_days = self._due_spin.value()
        today    = date.today()

        filtered = self._rows
        if search:
            filtered = [r for r in filtered if search in (
                r["so_number"] + r["sku_code"] + r["line_item"]
                + r["customer"]).lower()]
        if status_f != "ALL":
            filtered = [r for r in filtered if r["status"] == status_f]
        if due_days > 0:
            cutoff = today + timedelta(days=due_days)
            filtered = [r for r in filtered if r["due_dt"] <= cutoff]

        self._render(filtered)
        self._update_summary(filtered)

    def _render(self, rows: List[Dict]):
        bold = QFont(); bold.setBold(True)

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._table.setSortingEnabled(False)
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(rows))

        for ri, r in enumerate(rows):
            bg_hex, fg_hex = self._STATUS_COLORS.get(r["status"], ("#FFFFFF", "#000000"))

            vals = [
                str(r["rank"]),
                r["so_number"], r["customer"], r["sku_code"], r["line_item"],
                r["so_qty"], r["prod_needed"], r["planned_qty"],
                r["prod_complete"], r["release_date"],
                r["req_due"], r["committed_due"],
                str(r["days_to_due"]),
                r["cr_display"],
                r["status"], r["note"],
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if ci == 14:  # Status
                    item.setBackground(QBrush(QColor(bg_hex)))
                    item.setForeground(QBrush(QColor(fg_hex)))
                    item.setFont(bold)
                elif ci == 13 and r["cr"] is not None:  # CR
                    cr = r["cr"]
                    if cr < self.CR_CRITICAL:
                        item.setForeground(QBrush(QColor("#B71C1C")))
                        item.setFont(bold)
                    elif cr < self.CR_AT_RISK:
                        item.setForeground(QBrush(QColor("#E65100")))
                elif ci == 12:  # Days to Due
                    try:
                        d = int(r["days_to_due"])
                        if d < 0:
                            item.setBackground(QBrush(QColor("#FFCCCC")))
                        elif d <= self.AT_RISK_DAYS:
                            item.setBackground(QBrush(QColor("#FFF0CC")))
                    except (ValueError, TypeError):
                        pass
                self._table.setItem(ri, ci, item)

        self._table.setUpdatesEnabled(True)
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.setSortingEnabled(True)

    def _update_summary(self, rows: List[Dict]):
        total    = len(rows)
        critical = sum(1 for r in rows if r["status"] == self.STATUS_CRITICAL)
        at_risk  = sum(1 for r in rows if r["status"] == self.STATUS_AT_RISK)
        on_time  = sum(1 for r in rows if r["status"] == self.STATUS_ON_TIME)
        complete = sum(1 for r in rows if r["status"] == self.STATUS_COMPLETE)
        no_plan  = sum(1 for r in rows if r["status"] == self.STATUS_NO_PLAN)

        parts = [f"Total: <b>{total}</b>"]
        if critical:
            parts.append(
                f"<span style='color:#B71C1C'>🔴 CRITICAL: <b>{critical}</b></span>")
        if at_risk:
            parts.append(
                f"<span style='color:#E65100'>🟡 AT RISK: <b>{at_risk}</b></span>")
        if on_time:
            parts.append(
                f"<span style='color:#1B5E20'>🟢 ON TIME: <b>{on_time}</b></span>")
        if complete:
            parts.append(
                f"<span style='color:#757575'>✅ COMPLETE: <b>{complete}</b></span>")
        if no_plan:
            parts.append(
                f"<span style='color:#9E9E9E'>⚪ NOT PLANNED: <b>{no_plan}</b></span>")
        self._summary_lbl.setText("  |  ".join(parts))

    def _export(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            path, _ = QFileDialog.getSaveFileName(
                self, "Export Fulfillment Status",
                "Fulfillment_Status.xlsx", "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook(); ws = wb.active
            ws.title = "Fulfillment Status"

            HDR_FILL = PatternFill("solid", fgColor="2563EB")
            HDR_FONT = Font(color="FFFFFF", bold=True)
            STATUS_FILLS = {
                self.STATUS_CRITICAL: PatternFill("solid", fgColor="FFEBEE"),
                self.STATUS_AT_RISK:  PatternFill("solid", fgColor="FFF8E1"),
                self.STATUS_ON_TIME:  PatternFill("solid", fgColor="E8F5E9"),
                self.STATUS_COMPLETE: PatternFill("solid", fgColor="F5F5F5"),
                self.STATUS_NO_PLAN:  PatternFill("solid", fgColor="F5F5F5"),
            }

            for ci, h in enumerate(self._COLS, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = HDR_FILL
                c.font = HDR_FONT
                c.alignment = Alignment(horizontal="center")

            for ri, r in enumerate(self._rows, 2):
                vals = [
                    r["rank"], r["so_number"], r["customer"],
                    r["sku_code"], r["line_item"],
                    r["so_qty"], r["prod_needed"], r["planned_qty"],
                    r["prod_complete"], r["release_date"],
                    r["req_due"], r["committed_due"],
                    r["days_to_due"], r["cr_display"],
                    r["status"], r["note"],
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    if r["status"] in STATUS_FILLS:
                        cell.fill = STATUS_FILLS[r["status"]]

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

            wb.save(path)
            QMessageBox.information(self, "Export", f"Saved to {path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", str(e))


# ════════════════════════════════════════════════════════════════════════════
#  MB51 UPLOAD DIALOG
# ════════════════════════════════════════════════════════════════════════════

class MB51UploadDialog(QDialog):
    """Preview MB51 movements before committing to inventory."""

    def __init__(self, rows, parent=None):
        super().__init__(parent)
        self.rows   = rows
        self.result = {}
        self.setWindowTitle("MB51 Upload Preview")
        self.resize(820, 560)
        self._build_ui()

    def _build_ui(self):
        vl = QVBoxLayout(self)

        # ── Summary header ──
        from collections import Counter
        mv_counts = Counter(
            str(r.get("movement_type") or "").strip()
            for r in self.rows)
        from data.repositories import MB51Repo
        already = MB51Repo.processed_docs()
        new_count = sum(
            1 for r in self.rows
            if (str(r.get("material_document") or "").strip(),
                str(r.get("movement_type") or "").strip()) not in already)

        info = QLabel(
            f"Total rows: {len(self.rows)}  |  "
            f"New documents: {new_count}  |  "
            f"Already processed: {len(self.rows) - new_count}\n"
            f"Movement types: " +
            ", ".join(f"{mv}×{cnt}" for mv, cnt in sorted(mv_counts.items())))
        info.setStyleSheet(
            "background:#eff6ff; color:#1e3a5f; padding:8px; "
            "border-radius:4px; font-size:12px;")
        info.setWordWrap(True)
        vl.addWidget(info)

        # ── Row preview table ──
        grp = QGroupBox("MB51 Rows (preview — first 200)")
        gl = QVBoxLayout(grp)
        tbl = QTableWidget()
        cols = ["Posting Date", "Material", "Batch", "Qty",
                "Move Type", "Mat.Doc", "Status"]
        tbl.setColumnCount(len(cols))
        tbl.setHorizontalHeaderLabels(cols)
        tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setAlternatingRowColors(True)

        show = self.rows[:200]
        tbl.setRowCount(len(show))
        MV_COLORS = {"101": "#d4f0c0", "102": "#ffe0a0",
                     "331": "#fef9c3", "332": "#e0f2fe"}
        for ri, r in enumerate(show):
            mv  = str(r.get("movement_type") or "").strip()
            doc = str(r.get("material_document") or "").strip()
            is_new = (doc, mv) not in already
            vals = [
                str(r.get("posting_date") or ""),
                str(r.get("material")     or ""),
                str(r.get("batch")        or ""),
                str(r.get("quantity")     or ""),
                mv,
                doc,
                "NEW" if is_new else "skip",
            ]
            bg = QColor(MV_COLORS.get(mv, "#ffffff")) if is_new else QColor("#f0f0f0")
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setBackground(QBrush(bg))
                if ci == 6 and not is_new:
                    item.setForeground(QBrush(QColor("#999999")))
                tbl.setItem(ri, ci, item)
        tbl.resizeColumnsToContents()
        gl.addWidget(tbl)
        vl.addWidget(grp, stretch=1)

        # ── Buttons ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("Cancel")
        btn_confirm = QPushButton("✅ Confirm & Process")
        btn_confirm.setStyleSheet(
            "background:#16A34A; color:white; font-weight:bold; "
            "border:none; border-radius:5px; padding:6px 18px;")
        btn_confirm.setEnabled(new_count > 0)
        btn_cancel.clicked.connect(self.reject)
        btn_confirm.clicked.connect(self._confirm)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_confirm)
        vl.addLayout(btn_row)

    def _confirm(self):
        from core.mb51_processor import MB51Processor
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            self.result = MB51Processor().process(self.rows)
        finally:
            QApplication.restoreOverrideCursor()

        stats  = self.result
        skus   = stats.get("affected_skus", [])
        allocs = stats.get("allocation_stats", {})
        short_all = []
        for s in allocs.values():
            short_all.extend(s.get("sos_short", []))

        msg = (f"Processed {stats.get('new_docs', 0)} new documents.\n"
               f"Skipped {stats.get('skipped_docs', 0)} already-processed.\n"
               f"Affected SKUs: {', '.join(skus) or '—'}\n")
        if short_all:
            msg += (f"\n⚠ {len(short_all)} SO(s) still short after allocation "
                    "(additional production may be needed):\n")
            for s in short_all[:5]:
                msg += f"  • {s['so_number']} / {s['line_item']}: -{s['short_qty']:,}\n"
            if len(short_all) > 5:
                msg += f"  … and {len(short_all)-5} more.\n"

        QMessageBox.information(self, "MB51 Result", msg)
        self.accept()


# ════════════════════════════════════════════════════════════════════════════
#  HC DEMAND DIALOG (생산실별 인력 자동배정)
# ════════════════════════════════════════════════════════════════════════════

class HCDemandDialog(QDialog):
    """
    현재 생산계획을 기반으로 Shift별 필요 총 HC를 계산하고
    CRP Excel에 반영할 수 있는 다이얼로그.

    - 각 생산실/공정의 필요 HC를 계산 → Shift별로 합산
    - 현재 CRP 총 HC와 비교, 차이가 있는 행 강조
    - 체크 후 Apply → CRP Excel의 ShiftNo별 총 HC 업데이트
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("HC Recommendation — Shift Total")
        self.setMinimumSize(780, 500)
        self.to_apply: dict = {}
        self._build_ui()

    def _build_ui(self):
        from core.scheduler import scheduler, _shift_hours
        layout = QVBoxLayout(self)

        info = QLabel(
            "Calculates required total HC per Shift based on the current production plan. "
            "Enter only the Shift total in CRP Excel — the system distributes automatically.\n"
            "Check the rows to apply, then click Apply to update CRP Excel.")
        info.setWordWrap(True)
        info.setStyleSheet("padding:6px; background:#e8f0fe; border-radius:4px;")
        layout.addWidget(info)

        # ── Compute recommended total HC per (date, shift) ────────────────────
        plans = PlanRepo.all()
        shifts = {s["shift_no"]: s for s in ShiftRepo.all()}
        room_proc_map = {
            (rp["room_code"], rp["process_name"]): rp
            for rp in RoomRepo.all()
        }

        # Aggregate inner qty per (date, room, proc, shift)
        slot_qty: dict = {}
        for p in plans:
            key = (p["plan_date"], p["room_code"], p["process_name"], p["shift_no"])
            uom = 1
            if p.get("sku_code"):
                sku = SKURepo.get(p["sku_code"])
                if sku:
                    uom = int(sku.get("uom") or 1)
            slot_qty[key] = slot_qty.get(key, 0) + p["qty_planned"] * uom

        # Required HC per room/process/shift → sum to total per (date, shift)
        shift_req: dict = {}   # (date, shift) -> total required HC
        for (d, room, proc, sno), inner_qty in slot_qty.items():
            rp = room_proc_map.get((room, proc))
            sh = shifts.get(sno)
            if not rp or not sh:
                continue
            sh_h = _shift_hours(sh)
            if rp["process_type"] == "AUTO":
                req_hc = int(rp.get("hc_fixed") or 1)
            else:
                upph = float(rp.get("upph") or 0)
                if upph <= 0 or sh_h <= 0:
                    req_hc = int(rp.get("hc_min") or 1)
                else:
                    req_hc = math.ceil(inner_qty / (upph * sh_h))
                    req_hc = max(int(rp.get("hc_min") or 1),
                                 min(int(rp.get("hc_max") or 999), req_hc))
            shift_req[(d, sno)] = shift_req.get((d, sno), 0) + req_hc

        rows = []
        for (d, sno), req_total in sorted(shift_req.items()):
            cur_total = crp_manager.get_total_hc(d, sno)
            rows.append({
                "date": d, "shift": sno,
                "cur_hc": cur_total, "req_hc": req_total,
                "diff": req_total - cur_total,
            })

        # ── Table ─────────────────────────────────────────────────────────────
        self.table = QTableWidget(len(rows), 6)
        self.table.setHorizontalHeaderLabels([
            "Apply", "Date", "Shift",
            "Current CRP Total HC", "Recommended Total HC", "Diff"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)

        self._rows_data = rows
        self._checks: list = []
        for ri, r in enumerate(rows):
            cb = QCheckBox()
            cb.setChecked(r["diff"] != 0)
            self._checks.append(cb)
            self.table.setCellWidget(ri, 0, cb)
            self.table.setItem(ri, 1, QTableWidgetItem(r["date"]))
            self.table.setItem(ri, 2, QTableWidgetItem(f"Shift {r['shift']}"))
            self.table.setItem(ri, 3, QTableWidgetItem(str(r["cur_hc"])))
            self.table.setItem(ri, 4, QTableWidgetItem(str(r["req_hc"])))
            diff_item = QTableWidgetItem(f"{r['diff']:+d}")
            if r["diff"] > 0:
                diff_item.setForeground(QColor("#c62828"))
            elif r["diff"] < 0:
                diff_item.setForeground(QColor("#2e7d32"))
            self.table.setItem(ri, 5, diff_item)

            row_bg = QColor("#fff9c4") if r["diff"] != 0 else QColor("#f5f5f5")
            for ci in range(1, 6):
                if self.table.item(ri, ci):
                    self.table.item(ri, ci).setBackground(QBrush(row_bg))

        self.table.resizeColumnsToContents()
        layout.addWidget(self.table, stretch=1)

        # ── Buttons ───────────────────────────────────────────────────────────
        bbar = QHBoxLayout()
        btn_all  = QPushButton("Check All")
        btn_none = QPushButton("Uncheck All")
        btn_all.clicked.connect(lambda: [cb.setChecked(True) for cb in self._checks])
        btn_none.clicked.connect(lambda: [cb.setChecked(False) for cb in self._checks])
        bbar.addWidget(btn_all)
        bbar.addWidget(btn_none)
        bbar.addStretch()
        btn_cancel = QPushButton("Cancel")
        btn_apply  = QPushButton("💾 Apply to CRP Excel")
        btn_apply.setStyleSheet(
            "background:#2e6fd8; color:white; font-weight:bold; padding:6px 16px;")
        btn_cancel.clicked.connect(self.reject)
        btn_apply.clicked.connect(self._apply)
        bbar.addWidget(btn_cancel)
        bbar.addWidget(btn_apply)
        layout.addLayout(bbar)

        if not rows:
            info.setText("No production plans found. Run Auto-Plan first.")

    def _apply(self):
        self.to_apply = {}
        for r, cb in zip(self._rows_data, self._checks):
            if cb.isChecked():
                self.to_apply[(r["date"], r["shift"])] = r["req_hc"]
        if not self.to_apply:
            QMessageBox.information(self, "Apply", "No items checked.")
            return
        self.accept()


# ════════════════════════════════════════════════════════════════════════════
# LOT Batch Input Dialog (Actuals Tab)
# ════════════════════════════════════════════════════════════════════════════

class _LOTBatchDialog(QDialog):
    """Single dialog for entering LOT numbers for multiple actual entries at once."""

    def __init__(self, pending: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Enter LOT Numbers")
        self.resize(700, 300)
        self.results = []  # [(plan, qty, lot_str), ...]
        self._pending = pending
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("Enter LOT numbers for each actual entry, then click Confirm."))

        self._tbl = QTableWidget(len(self._pending), 5)
        self._tbl.setHorizontalHeaderLabels(
            ["SO", "SKU", "Line", "Qty", "LOT Number"])
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self._tbl.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)

        for ri, (plan, qty) in enumerate(self._pending):
            for ci, val in enumerate([
                plan["so_number"], plan["sku_code"], plan["line_item"], str(qty)
            ]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self._tbl.setItem(ri, ci, item)
            lot_item = QTableWidgetItem("")
            lot_item.setBackground(QBrush(QColor("#fffde7")))
            self._tbl.setItem(ri, 4, lot_item)

        lay.addWidget(self._tbl)

        bbar = QHBoxLayout()
        bbar.addStretch()
        btn_cancel = QPushButton("Cancel")
        btn_ok = QPushButton("✅ Confirm")
        btn_ok.setStyleSheet(
            "background:#2e6fd8; color:white; font-weight:bold; padding:6px 18px;")
        btn_cancel.clicked.connect(self.reject)
        btn_ok.clicked.connect(self._accept)
        bbar.addWidget(btn_cancel)
        bbar.addWidget(btn_ok)
        lay.addLayout(bbar)

    def _accept(self):
        self.results = []
        for ri, (plan, qty) in enumerate(self._pending):
            lot = self._tbl.item(ri, 4).text().strip() if self._tbl.item(ri, 4) else ""
            self.results.append((plan, qty, lot))
        self.accept()


# ════════════════════════════════════════════════════════════════════════════
# Plan Change History Dialog
# ════════════════════════════════════════════════════════════════════════════

class PlanHistoryDialog(QDialog):
    """Full plan change history viewer — moves, locks, deletes."""

    _ACTION_COLOR = {
        "MODIFIED":     QColor(255, 251, 210),
        "DELETED":      QColor(255, 218, 218),
        "LOCKED":       QColor(218, 236, 255),
        "UNLOCKED":     QColor(220, 255, 224),
        "BULK_CLEARED": QColor(255, 234, 200),
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Plan Change History")
        self.resize(1200, 660)
        self._all_rows: List[Dict] = []
        self._build_ui()
        self._load()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        fbar = QHBoxLayout()
        fbar.addWidget(QLabel("Action:"))
        self._act_combo = QComboBox()
        self._act_combo.addItems(
            ["ALL", "MODIFIED", "DELETED", "LOCKED", "UNLOCKED", "BULK_CLEARED"])
        self._act_combo.currentTextChanged.connect(self._apply_filter)
        fbar.addWidget(self._act_combo)
        fbar.addStretch()
        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self._load)
        fbar.addWidget(btn_refresh)
        btn_export = QPushButton("📥 Export")
        btn_export.clicked.connect(self._export)
        fbar.addWidget(btn_export)
        lay.addLayout(fbar)

        self._tbl = QTableWidget()
        self._tbl.setColumnCount(7)
        self._tbl.setHorizontalHeaderLabels(
            ["Timestamp", "Action", "Plan ID", "SKU / Process",
             "Before", "After", "Reason"])
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setStyleSheet("QTableWidget { font-size:12px; }")
        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self._tbl)

        self._count_lbl = QLabel("")
        self._count_lbl.setStyleSheet("color:#6b7280; font-size:11px;")
        lay.addWidget(self._count_lbl)

    def _load(self):
        self._all_rows = PlanRepo.plan_history()
        self._apply_filter()

    def _apply_filter(self):
        f = self._act_combo.currentText()
        rows = (self._all_rows if f == "ALL"
                else [r for r in self._all_rows if r["action"] == f])

        _ph_hdr = self._tbl.horizontalHeader()
        _ph_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._tbl.setUpdatesEnabled(False)
        self._tbl.setRowCount(len(rows))
        for i, r in enumerate(rows):
            old = self._parse(r.get("old_value"))
            new = self._parse(r.get("new_value"))

            ref = old if isinstance(old, dict) else (new if isinstance(new, dict) else {})
            sku_proc = " / ".join(filter(None, [
                ref.get("sku_code") or ref.get("entity_code"),
                ref.get("process_name"),
            ]))

            cells = [
                r.get("changed_at", ""),
                r.get("action", ""),
                str(r.get("plan_id") or ""),
                sku_proc,
                self._fmt(old),
                self._fmt(new),
                r.get("reason") or "",
            ]
            bg = self._ACTION_COLOR.get(r.get("action", ""))
            for j, v in enumerate(cells):
                item = QTableWidgetItem(str(v))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if bg:
                    item.setBackground(QBrush(bg))
                self._tbl.setItem(i, j, item)

        self._tbl.setUpdatesEnabled(True)
        self._tbl.resizeColumnsToContents()
        _ph_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _ph_hdr.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)

        self._count_lbl.setText(
            f"{len(rows)} record(s) shown  (total {len(self._all_rows)})")

    def _parse(self, s):
        if not s:
            return None
        try:
            return json.loads(s)
        except Exception:
            return s

    def _fmt(self, val) -> str:
        if val is None:
            return ""
        if not isinstance(val, dict):
            return str(val)
        parts = []
        if val.get("plan_date"):
            parts.append(val["plan_date"])
        if val.get("shift_no") is not None:
            parts.append(f"Shift {val['shift_no']}")
        if val.get("room_code"):
            parts.append(val["room_code"])
        if val.get("qty_planned") is not None:
            parts.append(f"qty={val['qty_planned']}")
        return " | ".join(parts) if parts else ""

    def _export(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            QMessageBox.warning(self, "Export", "openpyxl is not installed.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Plan History", "plan_history.xlsx",
            "Excel Files (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "PlanHistory"
        ws.append(["Timestamp", "Action", "Plan ID", "SKU / Process",
                   "Before", "After", "Reason"])
        for row in range(self._tbl.rowCount()):
            ws.append([
                self._tbl.item(row, col).text() if self._tbl.item(row, col) else ""
                for col in range(self._tbl.columnCount())
            ])
        wb.save(path)
        QMessageBox.information(self, "Export", f"Saved to:\n{path}")


# ════════════════════════════════════════════════════════════════════════════
#  IMPACT REPORT TAB  — Pull-in / Push-out vs plan
# ════════════════════════════════════════════════════════════════════════════

class ImpactReportTab(QWidget):
    """Shows how actual production has shifted SO completion dates
    relative to original plan dates (pull-in = early, push-out = late)."""

    _STATUS_STYLE = {
        "PULL-IN":     ("#E8F5E9", "#2E7D32"),
        "ON TRACK":    ("#FFFFFF", "#555555"),
        "PARTIAL":     ("#FFF8E1", "#F57F17"),
        "NOT STARTED": ("#F5F5F5", "#9E9E9E"),
        "PUSH-OUT":    ("#FFEBEE", "#C62828"),
        "COMPLETE":    ("#F5F5F5", "#9E9E9E"),
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._all_rows: List[Dict] = []
        self._build_ui()

    # ── Build ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Header
        hbar = QHBoxLayout()
        title = QLabel("Production Impact Report")
        title.setStyleSheet("font-size:14px; font-weight:bold;")
        hbar.addWidget(title)
        hbar.addStretch()
        self._ts_label = QLabel("")
        self._ts_label.setStyleSheet("color:#888; font-size:11px;")
        hbar.addWidget(self._ts_label)
        btn_refresh = QPushButton("Refresh")
        btn_refresh.clicked.connect(self.refresh)
        hbar.addWidget(btn_refresh)
        lay.addLayout(hbar)

        # KPI cards
        kpi_lay = QHBoxLayout()
        self._kpi_pull  = self._make_kpi("PULL-IN",   "#2E7D32", "#E8F5E9")
        self._kpi_push  = self._make_kpi("PUSH-OUT",  "#C62828", "#FFEBEE")
        self._kpi_track = self._make_kpi("ON TRACK",  "#37474F", "#ECEFF1")
        for box, _, _ in [self._kpi_pull, self._kpi_push, self._kpi_track]:
            kpi_lay.addWidget(box)
        lay.addLayout(kpi_lay)

        # Filter bar
        fbar = QHBoxLayout()
        fbar.addWidget(QLabel("SKU:"))
        self._sku_combo = QComboBox(); self._sku_combo.setMinimumWidth(150)
        self._sku_combo.currentIndexChanged.connect(self._apply_filters)
        fbar.addWidget(self._sku_combo)
        fbar.addWidget(QLabel("Status:"))
        self._status_combo = QComboBox()
        self._status_combo.addItems(["All", "PULL-IN", "ON TRACK", "PARTIAL",
                                     "PUSH-OUT", "NOT STARTED"])
        self._status_combo.currentIndexChanged.connect(self._apply_filters)
        fbar.addWidget(self._status_combo)
        fbar.addWidget(QLabel("Customer:"))
        self._cust_combo = QComboBox(); self._cust_combo.setMinimumWidth(150)
        self._cust_combo.currentIndexChanged.connect(self._apply_filters)
        fbar.addWidget(self._cust_combo)
        self._hide_complete_cb = QCheckBox("Hide completed")
        self._hide_complete_cb.setChecked(True)
        self._hide_complete_cb.toggled.connect(self._apply_filters)
        fbar.addWidget(self._hide_complete_cb)
        fbar.addStretch()
        btn_csv = QPushButton("📥 Export CSV")
        btn_csv.clicked.connect(self._export_csv)
        fbar.addWidget(btn_csv)
        lay.addLayout(fbar)

        # Splitter: main table + detail panel
        splitter = QSplitter(Qt.Orientation.Vertical)

        self._table = QTableWidget()
        self._table.setColumnCount(8)
        self._table.setHorizontalHeaderLabels([
            "SO#", "Customer", "SKU / Line",
            "Produced / Total", "Planned Release", "Projected Release",
            "Delta", "Status",
        ])
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(False)
        self._table.verticalHeader().setVisible(False)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.itemSelectionChanged.connect(self._on_row_selected)
        splitter.addWidget(self._table)

        # Detail panel
        self._detail = QGroupBox("SO Detail")
        self._detail.setVisible(False)
        dl = QVBoxLayout(self._detail)
        self._det_header = QLabel("")
        self._det_header.setWordWrap(True)
        self._det_header.setStyleSheet(
            "font-weight:bold; padding:6px; background:#f0f4ff; border-radius:4px;")
        dl.addWidget(self._det_header)

        self._det_table = QTableWidget()
        self._det_table.setColumnCount(6)
        self._det_table.setHorizontalHeaderLabels([
            "Process", "Seq", "Plan Date (max)",
            "Actual Date (max)", "Qty Planned", "Qty Actual",
        ])
        self._det_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._det_table.verticalHeader().setVisible(False)
        dhh = self._det_table.horizontalHeader()
        dhh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        dhh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        dl.addWidget(self._det_table)

        self._det_footer = QLabel("")
        self._det_footer.setStyleSheet("font-size:11px; color:#444; padding:4px;")
        self._det_footer.setWordWrap(True)
        dl.addWidget(self._det_footer)
        splitter.addWidget(self._detail)
        splitter.setSizes([500, 250])
        lay.addWidget(splitter, stretch=1)

        self._footer_lbl = QLabel("")
        self._footer_lbl.setStyleSheet("color:#777; font-size:11px; padding:2px 4px;")
        lay.addWidget(self._footer_lbl)

        self.refresh()

    # ── KPI factory ───────────────────────────────────────────────────────

    def _make_kpi(self, title: str, fg: str, bg: str):
        box = QGroupBox(title)
        box.setStyleSheet(
            f"QGroupBox{{background:{bg};border:1px solid {fg};"
            f"border-radius:6px;margin-top:8px;padding:4px;}}"
            f"QGroupBox::title{{color:{fg};font-weight:bold;"
            f"subcontrol-origin:margin;left:10px;}}")
        vl = QVBoxLayout(box)
        cnt = QLabel("—")
        cnt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cnt.setStyleSheet(f"font-size:26px;font-weight:bold;color:{fg};")
        sub = QLabel("")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet(f"color:{fg};font-size:11px;")
        vl.addWidget(cnt)
        vl.addWidget(sub)
        return box, cnt, sub

    def _set_kpi(self, card, count: int, sub: str = ""):
        _, cnt, sub_lbl = card
        cnt.setText(str(count))
        sub_lbl.setText(sub)

    # ── Data ──────────────────────────────────────────────────────────────

    def refresh(self):
        self._all_rows = PlanRepo.impact_summary()
        self._rebuild_combos()
        self._apply_filters()
        self._ts_label.setText(
            f"Last: {datetime.now().strftime('%H:%M:%S')}")

    def _rebuild_combos(self):
        skus  = sorted({r["sku_code"] for r in self._all_rows})
        custs = sorted({r.get("customer_name") or "" for r in self._all_rows} - {""})
        for combo, header, items in [
            (self._sku_combo,  "All SKUs",       skus),
            (self._cust_combo, "All Customers",  custs),
        ]:
            prev = combo.currentText()
            combo.blockSignals(True)
            combo.clear()
            combo.addItem(header)
            combo.addItems(items)
            idx = combo.findText(prev)
            combo.setCurrentIndex(max(0, idx))
            combo.blockSignals(False)

    # ── Filters ───────────────────────────────────────────────────────────

    def _apply_filters(self):
        sku_f    = self._sku_combo.currentText()
        status_f = self._status_combo.currentText()
        cust_f   = self._cust_combo.currentText()
        hide_done = self._hide_complete_cb.isChecked()

        visible = []
        for r in self._all_rows:
            if hide_done and r["impact_status"] == "COMPLETE":
                continue
            if not sku_f.startswith("All") and r["sku_code"] != sku_f:
                continue
            if status_f != "All" and r["impact_status"] != status_f:
                continue
            if not cust_f.startswith("All"):
                if (r.get("customer_name") or "") != cust_f:
                    continue
            visible.append(r)

        self._render_table(visible)

        pull  = [r for r in self._all_rows if r["impact_status"] == "PULL-IN"]
        push  = [r for r in self._all_rows if r["impact_status"] == "PUSH-OUT"]
        track = [r for r in self._all_rows if r["impact_status"] == "ON TRACK"]
        best  = min((r["delta_days"] for r in pull), default=0)
        worst = max((r["delta_days"] for r in push), default=0)
        self._set_kpi(self._kpi_pull,  len(pull),
                      f"Best: {best:+d}d" if pull else "")
        self._set_kpi(self._kpi_push,  len(push),
                      f"Worst: {worst:+d}d" if push else "")
        self._set_kpi(self._kpi_track, len(track))

    # ── Table rendering ───────────────────────────────────────────────────

    def _render_table(self, rows: List[Dict]):
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(rows))
        self._table.setProperty("_rows", rows)

        for ri, r in enumerate(rows):
            status = r["impact_status"]
            delta  = r.get("delta_days")
            bg, fg = self._STATUS_STYLE.get(status, ("#FFFFFF", "#333"))
            if status == "PUSH-OUT" and delta is not None and delta > 5:
                bg, fg = "#FFCDD2", "#B71C1C"

            prefix = {"PULL-IN": ">> ", "PUSH-OUT": "!! "}.get(status, "")
            delta_str = f"{delta:+d}d" if delta is not None else "—"

            vals = [
                prefix + r["so_number"],
                r.get("customer_name") or "",
                f"{r['sku_code']} / {r['line_item']}",
                f"{r['qty_produced']} / {r['qty']}",
                r.get("planned_release", ""),
                r.get("projected_release", ""),
                delta_str,
                status,
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setBackground(QBrush(QColor(bg)))
                if ci == 6:
                    item.setForeground(QBrush(QColor(fg)))
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    f2 = item.font(); f2.setBold(bool(delta)); item.setFont(f2)
                if ci == 5 and status in ("PARTIAL", "NOT STARTED"):
                    item.setToolTip("Estimated — production not yet complete")
                    item.setForeground(QBrush(QColor("#999")))
                self._table.setItem(ri, ci, item)

        self._table.setUpdatesEnabled(True)
        self._table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._footer_lbl.setText(f"{len(rows)} SO(s) shown")
        self._detail.setVisible(False)

    # ── Detail panel ──────────────────────────────────────────────────────

    def _on_row_selected(self):
        rows = self._table.property("_rows") or []
        sel = self._table.currentRow()
        if sel < 0 or sel >= len(rows):
            self._detail.setVisible(False)
            return
        self._show_detail(rows[sel])

    def _show_detail(self, r: Dict):
        so, sku, li = r["so_number"], r["sku_code"], r["line_item"]
        post = r.get("post_lead_days") or 0

        self._det_header.setText(
            f"SO: {so}  ·  Customer: {r.get('customer_name') or '—'}  ·  "
            f"SKU: {sku} / {li}  ·  "
            f"Priority: {r.get('priority', '—')}  ·  Due: {r.get('due_date', '—')}"
        )

        plans = PlanRepo.for_so(so, sku, li)
        actuals = ActualRepo.for_so(so, sku, li)
        total_actual = sum(a["qty_actual"] for a in actuals)
        max_actual_date = max((a["actual_date"] for a in actuals), default="")

        # Group plans by process_seq
        steps: Dict[int, Dict] = {}
        for p in plans:
            if p.get("entity_type", "SKU") != "SKU":
                continue
            seq = p["process_seq"]
            if seq not in steps:
                steps[seq] = {
                    "name":      p["process_name"],
                    "is_final":  p.get("is_final_seq", 0),
                    "dates":     [],
                    "qty":       0,
                }
            steps[seq]["dates"].append(p["plan_date"])
            steps[seq]["qty"] += p["qty_planned"]

        step_list = sorted(steps.values(), key=lambda s: s["name"])
        self._det_table.setRowCount(len(step_list))
        for ri, s in enumerate(step_list):
            max_plan = max(s["dates"]) if s["dates"] else ""
            is_final = s["is_final"]
            vals = [
                ("⭐ " if is_final else "") + s["name"],
                str(min(
                    p["process_seq"] for p in plans
                    if p["process_name"] == s["name"]
                ) if plans else ""),
                max_plan,
                max_actual_date if is_final else "—",
                str(s["qty"]),
                str(total_actual) if is_final else "—",
            ]
            for ci, val in enumerate(vals):
                item = QTableWidgetItem(val)
                if is_final:
                    item.setBackground(QBrush(QColor("#FFF8E1")))
                self._det_table.setItem(ri, ci, item)

        delta = r.get("delta_days")
        delta_txt = ""
        if delta is not None:
            if delta < -1:
                delta_txt = f" → PULL-IN {abs(delta)}d ahead of plan"
            elif delta > 1:
                delta_txt = f" → PUSH-OUT {delta}d behind plan"
            else:
                delta_txt = " → ON TRACK"

        self._det_footer.setText(
            f"Planned Release: {r.get('planned_release', '—')} "
            f"(final plan + {post} post-lead day(s))  ·  "
            f"Projected Release: {r.get('projected_release', '—')}"
            f"{delta_txt}  ·  "
            f"Inventory Allocated: {r.get('qty_inv_allocated', 0)}  ·  "
            f"Status: {r['impact_status']}"
        )
        self._detail.setVisible(True)

    # ── Export ────────────────────────────────────────────────────────────

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Impact Report", "impact_report.csv",
            "CSV Files (*.csv)")
        if not path:
            return
        import csv
        rows = self._table.property("_rows") or []
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["SO#", "Customer", "SKU", "Line Item",
                        "Produced", "Total Qty", "Planned Release",
                        "Projected Release", "Delta (days)", "Status"])
            for r in rows:
                w.writerow([
                    r["so_number"], r.get("customer_name", ""),
                    r["sku_code"], r["line_item"],
                    r["qty_produced"], r["qty"],
                    r.get("planned_release", ""),
                    r.get("projected_release", ""),
                    r.get("delta_days", ""),
                    r["impact_status"],
                ])
        QMessageBox.information(self, "Export", f"Saved to:\n{path}")


# ════════════════════════════════════════════════════════════════════════════
#  SCENARIO PLANNER TAB
# ════════════════════════════════════════════════════════════════════════════

class ScenarioTab(QWidget):
    """
    Scenario Planner: bottleneck detection + stepped HC scenario simulation.
    Scenarios are saved to DB for decision-making.
    """
    HC_STEP = 5

    def __init__(self, parent=None):
        super().__init__(parent)
        self._bottlenecks: List[Dict] = []
        self._max_hc: int = 0
        self._build_ui()
        self._load_saved()

    def _build_ui(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Left panel: controls + bottlenecks ──────────────────────────────
        left = QWidget()
        left.setFixedWidth(320)
        left.setStyleSheet("background:#f4f5f8; border-right:1px solid #dde3ed;")
        lv = QVBoxLayout(left)
        lv.setContentsMargins(12, 12, 12, 12)
        lv.setSpacing(8)

        lv.addWidget(self._section_label("ANALYSIS PERIOD"))
        dr = QHBoxLayout()
        self._d_from = QDateEdit(QDate.currentDate())
        self._d_from.setDisplayFormat("yyyy-MM-dd")
        self._d_to = QDateEdit(QDate.currentDate().addDays(83))
        self._d_to.setDisplayFormat("yyyy-MM-dd")
        dr.addWidget(QLabel("From:")); dr.addWidget(self._d_from)
        dr.addWidget(QLabel("To:"));   dr.addWidget(self._d_to)
        lv.addLayout(dr)

        self._name_edit = QLineEdit("Scenario")
        self._name_edit.setPlaceholderText("Scenario name")
        lv.addWidget(self._name_edit)

        btn_detect = QPushButton("🔍 Detect Bottlenecks")
        btn_detect.setStyleSheet(
            "background:#2563EB;color:white;font-weight:bold;"
            "border:none;border-radius:5px;padding:6px 12px;")
        btn_detect.clicked.connect(self._detect)
        lv.addWidget(btn_detect)

        lv.addWidget(self._section_label("BOTTLENECKS"))
        self._bn_tbl = QTableWidget(0, 4)
        self._bn_tbl.setHorizontalHeaderLabels(["Room", "Process", "Unmet Qty", "Rec. HC"])
        self._bn_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._bn_tbl.verticalHeader().setVisible(False)
        self._bn_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._bn_tbl.setFixedHeight(140)
        lv.addWidget(self._bn_tbl)

        self._max_hc_lbl = QLabel("Max recommended HC: —")
        self._max_hc_lbl.setStyleSheet("font-size:11px;font-weight:bold;color:#1e293b;")
        lv.addWidget(self._max_hc_lbl)

        btn_sim = QPushButton("▶ Run Stepped Scenarios")
        btn_sim.setStyleSheet(
            "background:#16A34A;color:white;font-weight:bold;"
            "border:none;border-radius:5px;padding:6px 12px;")
        btn_sim.clicked.connect(self._run_scenarios)
        lv.addWidget(btn_sim)

        lv.addWidget(self._section_label("SAVED SCENARIOS"))
        self._saved_list = QListWidget()
        self._saved_list.setFixedHeight(160)
        self._saved_list.currentRowChanged.connect(self._load_scenario)
        lv.addWidget(self._saved_list)

        btn_del = QPushButton("🗑 Delete Selected")
        btn_del.setStyleSheet(
            "background:#DC2626;color:white;border:none;border-radius:4px;padding:4px 10px;")
        btn_del.clicked.connect(self._delete_scenario)
        lv.addWidget(btn_del, alignment=Qt.AlignmentFlag.AlignLeft)
        lv.addStretch()

        # ── Right panel: results ─────────────────────────────────────────────
        right = QWidget()
        rv = QVBoxLayout(right)
        rv.setContentsMargins(12, 12, 12, 12)
        rv.setSpacing(8)

        rv.addWidget(self._section_label("SCENARIO COMPARISON"))

        self._cmp_tbl = QTableWidget(0, 5)
        self._cmp_tbl.setHorizontalHeaderLabels(
            ["HC Added", "LATE Before", "LATE After", "Resolved", "Resolution Rate"])
        self._cmp_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._cmp_tbl.verticalHeader().setVisible(False)
        self._cmp_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._cmp_tbl.setAlternatingRowColors(True)
        self._cmp_tbl.setFixedHeight(160)
        rv.addWidget(self._cmp_tbl)

        rv.addWidget(self._section_label("RESOLVED ORDERS (selected scenario step)"))
        self._detail_tbl = QTableWidget(0, 3)
        self._detail_tbl.setHorizontalHeaderLabels(["SO Number", "Status", "Note"])
        self._detail_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self._detail_tbl.verticalHeader().setVisible(False)
        self._detail_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._detail_tbl.setAlternatingRowColors(True)
        self._cmp_tbl.itemSelectionChanged.connect(self._on_step_selected)
        rv.addWidget(self._detail_tbl, 1)

        self._note_lbl = QLabel("")
        self._note_lbl.setWordWrap(True)
        self._note_lbl.setStyleSheet("color:#6b7280;font-size:10px;padding:4px;")
        rv.addWidget(self._note_lbl)

        outer.addWidget(left)
        outer.addWidget(right, 1)

        self._current_results: List[Dict] = []

    @staticmethod
    def _section_label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "color:#6b7aa3;font-size:10px;font-weight:bold;"
            "padding:4px 0 2px 0;letter-spacing:0.05em;")
        return lbl

    def _detect(self):
        from core.scheduler import scheduler as _sched
        d0 = self._d_from.date().toString("yyyy-MM-dd")
        d1 = self._d_to.date().toString("yyyy-MM-dd")
        try:
            self._bottlenecks = _sched.detect_bottlenecks(d0, d1)
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
            return

        self._bn_tbl.setRowCount(len(self._bottlenecks))
        total_rec = 0
        for ri, bn in enumerate(self._bottlenecks):
            for ci, val in enumerate([
                bn["room_code"], bn["process_name"],
                f"{bn['unmet_qty']:,.0f}", str(bn["recommended_hc"])
            ]):
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._bn_tbl.setItem(ri, ci, it)
            total_rec += bn["recommended_hc"]

        step = self.HC_STEP
        self._max_hc = ((total_rec + step - 1) // step) * step
        if self._max_hc == 0:
            self._max_hc = step
        self._max_hc_lbl.setText(f"Max recommended HC: {self._max_hc} people")

        if not self._bottlenecks:
            self._note_lbl.setText(
                "No bottlenecks found — all MANUAL rooms have sufficient capacity.")

    def _run_scenarios(self):
        if not self._bottlenecks:
            QMessageBox.information(self, "Detect First",
                                    "Click '🔍 Detect Bottlenecks' first.")
            return
        from core.scheduler import scheduler as _sched
        from data.repositories import ScenarioRepo
        d0 = self._d_from.date().toString("yyyy-MM-dd")
        d1 = self._d_to.date().toString("yyyy-MM-dd")
        name = self._name_edit.text().strip() or "Scenario"

        step = self.HC_STEP
        max_hc = self._max_hc
        hc_steps = []
        v = step
        while v <= max_hc:
            hc_steps.append(v)
            v += step
        if not hc_steps or hc_steps[-1] < max_hc:
            hc_steps.append(max_hc)
        hc_steps = sorted(set(hc_steps))

        bn_json = json.dumps(self._bottlenecks)
        scen_id = ScenarioRepo.insert(name, d0, d1, max_hc, step, bn_json)

        results = []
        for hc in hc_steps:
            try:
                res = _sched.simulate_hc_scenario(d0, d1, self._bottlenecks, hc)
            except Exception:
                res = {"late_before": 0, "late_after": 0, "resolved_sos": [], "detail": {}}
            ScenarioRepo.insert_result(
                scen_id, hc,
                res["late_before"], res["late_after"],
                json.dumps(res["resolved_sos"]),
                json.dumps(res["detail"]))
            results.append({"hc_added": hc, **res})

        self._current_results = results
        self._populate_cmp_table(results)
        self._load_saved()
        for i in range(self._saved_list.count()):
            item = self._saved_list.item(i)
            if item and item.data(Qt.ItemDataRole.UserRole) == scen_id:
                self._saved_list.setCurrentRow(i)
                break

    def _populate_cmp_table(self, results: List[Dict]):
        self._cmp_tbl.setRowCount(len(results))
        for ri, r in enumerate(results):
            lb = r["late_before"]
            la = r["late_after"]
            resolved = len(r["resolved_sos"])
            rate = f"{resolved / lb * 100:.0f}%" if lb > 0 else "N/A"
            vals = [f"+{r['hc_added']} ppl", str(lb), str(la), str(resolved), rate]
            colors = [None, "#fee2e2", "#dcfce7" if la < lb else "#fee2e2",
                      "#dcfce7" if resolved > 0 else None,
                      "#dcfce7" if resolved > 0 else None]
            for ci, (val, bg) in enumerate(zip(vals, colors)):
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if bg:
                    it.setBackground(QBrush(QColor(bg)))
                self._cmp_tbl.setItem(ri, ci, it)
        self._cmp_tbl.resizeColumnsToContents()

    def _on_step_selected(self):
        row = self._cmp_tbl.currentRow()
        if row < 0 or row >= len(self._current_results):
            self._detail_tbl.setRowCount(0)
            return
        r = self._current_results[row]
        sos = r["resolved_sos"]
        self._detail_tbl.setRowCount(len(sos))
        for ri, so_no in enumerate(sos):
            for ci, val in enumerate([so_no, "RESOLVED", "Expected to meet due date"]):
                it = QTableWidgetItem(val)
                if ci == 1:
                    it.setForeground(QBrush(QColor("#16A34A")))
                self._detail_tbl.setItem(ri, ci, it)

    def _load_saved(self):
        from data.repositories import ScenarioRepo
        scenarios = ScenarioRepo.all()
        self._saved_list.clear()
        for s in scenarios:
            item = QListWidgetItem(
                f"{s['name']}  ({s['date_from']}~{s['date_to']})  +{s['max_hc_add']}ppl")
            item.setData(Qt.ItemDataRole.UserRole, s["scenario_id"])
            self._saved_list.addItem(item)

    def _load_scenario(self, row: int):
        if row < 0:
            return
        item = self._saved_list.item(row)
        if not item:
            return
        scen_id = item.data(Qt.ItemDataRole.UserRole)
        from data.repositories import ScenarioRepo
        results_raw = ScenarioRepo.results_for(scen_id)
        results = []
        for r in results_raw:
            sos = json.loads(r.get("resolved_sos") or "[]")
            results.append({
                "hc_added":     r["hc_added"],
                "late_before":  r["late_before"],
                "late_after":   r["late_after"],
                "resolved_sos": sos,
                "detail":       json.loads(r.get("detail") or "{}"),
            })
        self._current_results = results
        self._populate_cmp_table(results)

        scen = ScenarioRepo.get(scen_id)
        if scen:
            bns = json.loads(scen.get("bottlenecks") or "[]")
            self._bottlenecks = bns
            self._bn_tbl.setRowCount(len(bns))
            for ri, bn in enumerate(bns):
                for ci, val in enumerate([
                    bn.get("room_code", ""), bn.get("process_name", ""),
                    f"{bn.get('unmet_qty', 0):,.0f}", str(bn.get("recommended_hc", 0))
                ]):
                    it = QTableWidgetItem(val)
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self._bn_tbl.setItem(ri, ci, it)
            self._max_hc_lbl.setText(f"Max recommended HC: {scen['max_hc_add']} people")

    def _delete_scenario(self):
        row = self._saved_list.currentRow()
        if row < 0:
            return
        item = self._saved_list.item(row)
        if not item:
            return
        scen_id = item.data(Qt.ItemDataRole.UserRole)
        if QMessageBox.question(
                self, "Delete", "Delete this scenario?",
                QMessageBox.StandardButton.Yes |
                QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            from data.repositories import ScenarioRepo
            ScenarioRepo.delete(scen_id)
            self._load_saved()
            self._cmp_tbl.setRowCount(0)
            self._detail_tbl.setRowCount(0)
            self._current_results = []

    def refresh(self):
        self._load_saved()


# ─── Labor Utilization Tab ────────────────────────────────────────────────────

class LaborUtilizationTab(QWidget):
    """
    Grid view: X-axis = Date × Shift, Y-axis = Room × Process
    Cell = distributed HC allocated to that room/process in each shift.
    Footer rows show Total HC, CRP HC, and Util %.
    """

    _COL_ZERO  = QColor(241, 245, 249)   # 0 HC — light gray
    _COL_OK    = QColor(209, 250, 229)   # < 80 % — green
    _COL_WARN  = QColor(254, 249, 195)   # 80–99 % — yellow
    _COL_OVER  = QColor(254, 226, 226)   # ≥ 100 % — red
    _COL_ROOM  = QColor(226, 232, 240)   # room header row
    _COL_TOTAL = QColor(219, 234, 254)   # footer rows (blue tint)
    _COL_TOTAL_OVER = QColor(254, 202, 202)   # Util% cell when ≥ 100 %

    def __init__(self, parent=None):
        super().__init__(parent)
        self._loaded = False
        self._last_col_keys = []
        self._last_flat_rows = []
        self._last_dist: Dict = {}
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # ── Controls ─────────────────────────────────────────────────────────
        ctrl = QHBoxLayout()
        ctrl.setSpacing(6)

        ctrl.addWidget(QLabel("From:"))
        self._from_edit = QDateEdit()
        self._from_edit.setCalendarPopup(True)
        self._from_edit.setDate(QDate.currentDate())
        self._from_edit.setFixedWidth(110)
        ctrl.addWidget(self._from_edit)

        ctrl.addWidget(QLabel("To:"))
        self._to_edit = QDateEdit()
        self._to_edit.setCalendarPopup(True)
        self._to_edit.setDate(QDate.currentDate().addDays(13))
        self._to_edit.setFixedWidth(110)
        ctrl.addWidget(self._to_edit)

        btn_refresh = QPushButton("▶ Refresh")
        btn_refresh.setFixedWidth(90)
        btn_refresh.clicked.connect(self._load_table)
        ctrl.addWidget(btn_refresh)

        ctrl.addSpacing(16)

        # Legend chips
        for bg, fg, label in [
            ("#d1fae5", "#166534", "< 80%"),
            ("#fef9c3", "#854d0e", "80–99%"),
            ("#fee2e2", "#991b1b", "≥ 100%"),
            ("#f1f5f9", "#64748b", "0 HC"),
        ]:
            chip = QLabel(f"  {label}  ")
            chip.setStyleSheet(
                f"background:{bg}; color:{fg}; border-radius:3px; "
                f"font-size:11px; padding:2px 4px;")
            ctrl.addWidget(chip)

        ctrl.addStretch()

        btn_export = QPushButton("📥 Export")
        btn_export.setFixedWidth(90)
        btn_export.clicked.connect(self._export_excel)
        btn_export.setStyleSheet(
            "background:#2563EB; color:white; font-weight:bold; "
            "border:none; border-radius:5px; padding:5px 10px;")
        ctrl.addWidget(btn_export)

        root.addLayout(ctrl)

        # ── Status label ─────────────────────────────────────────────────────
        self._status_lbl = QLabel("")
        self._status_lbl.setStyleSheet("color:#64748b; font-size:11px;")
        root.addWidget(self._status_lbl)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setMinimumSectionSize(70)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(False)
        self._table.setStyleSheet(
            "QTableWidget { gridline-color: #CBD5E1; }"
            "QHeaderView::section { background:#334155; color:white; "
            "  font-weight:bold; font-size:11px; padding:4px; border:none; }")
        root.addWidget(self._table)

    # ── Public API ────────────────────────────────────────────────────────────

    def refresh(self):
        self._load_table()

    def invalidate(self):
        """Kept for call-site compatibility; refresh() always reloads now."""
        pass

    # ── Data loading ──────────────────────────────────────────────────────────

    def _load_table(self):
        d_from = self._from_edit.date().toString("yyyy-MM-dd")
        d_to   = self._to_edit.date().toString("yyyy-MM-dd")

        if d_from > d_to:
            QMessageBox.warning(self, "Invalid Range",
                                "From date must be ≤ To date.")
            return

        self._status_lbl.setText("Loading…")
        self.repaint()

        dist = scheduler.compute_hc_distribution_preview(d_from, d_to)

        # ── Column keys: ALL CRP dates in range (HC > 0), not just plan dates ──
        # This ensures dates without plans but with CRP data still appear.
        shifts_all = ShiftRepo.all()
        shift_order = {s["shift_no"]: i for i, s in
                       enumerate(sorted(shifts_all, key=lambda x: x["shift_no"]))}
        avl_hc = scheduler.get_available_hc_by_date(d_from, d_to)
        col_keys = sorted(
            [(ds, sno) for ds, shifts in avl_hc.items()
             for sno, hc in shifts.items() if hc > 0],
            key=lambda k: (k[0], shift_order.get(k[1], k[1])))

        if not col_keys:
            self._table.setRowCount(0)
            self._table.setColumnCount(0)
            self._status_lbl.setText(
                "No HC data for this range. "
                "Check CRP file path (App Config → CRP File Path).")
            self._loaded = True
            return

        # ── Row keys: (room, process) in master-defined order ─────────────────
        rooms = RoomRepo.rooms()
        seen: set = set()
        flat_rows: List = []   # ("room_header", room, "") or ("proc", room, proc)
        for room in rooms:
            procs_for_room = [rp["process_name"]
                              for rp in RoomRepo.processes_for_room(room)]
            if not procs_for_room:
                continue
            flat_rows.append(("room_header", room, ""))
            for proc in procs_for_room:
                key = (room, proc)
                if key not in seen:
                    seen.add(key)
                    flat_rows.append(("proc", room, proc))

        if not flat_rows:
            self._status_lbl.setText("No Room/Process masters found.")
            self._loaded = True
            return

        N_FOOTER = 3
        n_cols = 1 + len(col_keys)
        n_rows = len(flat_rows) + N_FOOTER

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(n_rows)
        self._table.setColumnCount(n_cols)

        # Column headers
        headers = ["Room / Process"] + [
            f"{k[0][5:].replace('-', '/')}  S{k[1]}" for k in col_keys
        ]
        self._table.setHorizontalHeaderLabels(headers)

        f_bold  = QFont("Segoe UI", 9)
        f_bold.setBold(True)
        f_norm  = QFont("Segoe UI", 9)

        def _cell(text: str, font=None, bg=None, fg=None,
                  center=True, selectable=True) -> QTableWidgetItem:
            it = QTableWidgetItem(text)
            if center:
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if font:
                it.setFont(font)
            if bg:
                it.setBackground(QBrush(bg))
            if fg:
                it.setForeground(QBrush(fg))
            if not selectable:
                it.setFlags(Qt.ItemFlag.ItemIsEnabled)
            return it

        # ── Data rows ─────────────────────────────────────────────────────────
        for ri, (rtype, room, proc) in enumerate(flat_rows):
            if rtype == "room_header":
                it = _cell(f"  {room}", font=f_bold, bg=self._COL_ROOM,
                           center=False, selectable=False)
                self._table.setItem(ri, 0, it)
                self._table.setSpan(ri, 0, 1, n_cols)
                self._table.setRowHeight(ri, 22)
            else:
                self._table.setItem(
                    ri, 0,
                    _cell(f"    {proc}", font=f_norm, center=False))
                self._table.setRowHeight(ri, 24)

                for ci, col_key in enumerate(col_keys):
                    hc_map    = dist.get(col_key, {})
                    hc        = hc_map.get((room, proc), 0)
                    crp_total = crp_manager.get_total_hc(col_key[0], col_key[1])

                    text = str(hc) if hc > 0 else "—"
                    bg   = self._COL_ZERO
                    if hc > 0 and crp_total > 0:
                        util = hc / crp_total
                        bg = (self._COL_OVER  if util >= 1.0 else
                              self._COL_WARN  if util >= 0.8 else
                              self._COL_OK)
                    elif hc > 0:
                        bg = self._COL_OK

                    self._table.setItem(ri, ci + 1, _cell(text, bg=bg))

        # ── Footer rows ───────────────────────────────────────────────────────
        footer_labels = ["Total HC", "CRP HC", "Util %"]
        for fi, lbl in enumerate(footer_labels):
            r = len(flat_rows) + fi
            self._table.setItem(
                r, 0,
                _cell(lbl, font=f_bold, bg=self._COL_TOTAL,
                      center=False, selectable=False))
            self._table.setRowHeight(r, 24)

        for ci, col_key in enumerate(col_keys):
            hc_map    = dist.get(col_key, {})
            total_hc  = sum(hc_map.values())
            crp_total = crp_manager.get_total_hc(col_key[0], col_key[1])
            util_pct  = round(total_hc / crp_total * 100) if crp_total > 0 else 0

            values = [
                str(total_hc),
                str(crp_total) if crp_total > 0 else "—",
                f"{util_pct}%" if crp_total > 0 else "—",
            ]
            for fi, txt in enumerate(values):
                r  = len(flat_rows) + fi
                bg = self._COL_TOTAL
                fg = None
                if fi == 2 and crp_total > 0:   # Util% row
                    if util_pct >= 100:
                        bg = self._COL_TOTAL_OVER
                        fg = QColor(153, 27, 27)
                    elif util_pct >= 80:
                        fg = QColor(133, 77, 14)
                self._table.setItem(
                    r, ci + 1,
                    _cell(txt, font=f_bold, bg=bg, fg=fg, selectable=False))

        self._table.setUpdatesEnabled(True)
        self._table.resizeColumnsToContents()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        # Freeze first column width after resize
        self._table.setColumnWidth(0, max(self._table.columnWidth(0), 180))

        # Cache for export
        self._last_col_keys  = col_keys
        self._last_flat_rows = flat_rows
        self._last_dist      = dist

        n_shifts = len(col_keys)
        n_rooms  = sum(1 for r in flat_rows if r[0] == "room_header")
        self._status_lbl.setText(
            f"{n_shifts} shift(s) across {d_from} – {d_to}  |  "
            f"{n_rooms} room(s)  |  "
            f"Color threshold: HC / CRP total  (80% = yellow, 100% = red)")
        self._loaded = True

    # ── Excel Export ──────────────────────────────────────────────────────────

    def _export_excel(self):
        if not self._last_col_keys:
            QMessageBox.warning(self, "No Data",
                                "Click Refresh to load data first.")
            return

        d_from = self._from_edit.date().toString("yyyy-MM-dd")
        d_to   = self._to_edit.date().toString("yyyy-MM-dd")
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Labor Utilization",
            f"LaborUtil_{d_from}_{d_to}.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Labor Utilization"

            def _fill(hex6: str) -> PatternFill:
                return PatternFill("solid", fgColor=hex6)

            FILL_ROOM  = _fill("E2E8F0")
            FILL_TOTAL = _fill("DBEAFE")
            FILL_TOT_OVER = _fill("FECACA")
            FILL_OK    = _fill("D1FAE5")
            FILL_WARN  = _fill("FEF9C3")
            FILL_OVER  = _fill("FEE2E2")
            FILL_ZERO  = _fill("F1F5F9")

            f_bold   = Font(bold=True)
            f_hdr    = Font(bold=True, color="FFFFFF")
            aln_c    = Alignment(horizontal="center", vertical="center")
            aln_l    = Alignment(horizontal="left",   vertical="center")
            thin     = Side(style="thin", color="CBD5E1")
            brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

            col_keys  = self._last_col_keys
            flat_rows = self._last_flat_rows
            dist      = self._last_dist
            n_cols    = len(col_keys)

            # ── Header row ────────────────────────────────────────────────────
            hdr_fill = _fill("334155")
            c = ws.cell(1, 1, "Room / Process")
            c.font = f_hdr; c.fill = hdr_fill; c.alignment = aln_l; c.border = brd
            for ci, col_key in enumerate(col_keys):
                label = f"{col_key[0][5:].replace('-', '/')}  S{col_key[1]}"
                c = ws.cell(1, ci + 2, label)
                c.font = f_hdr; c.fill = hdr_fill; c.alignment = aln_c; c.border = brd
            ws.row_dimensions[1].height = 22

            # ── Data rows ─────────────────────────────────────────────────────
            for ri, (rtype, room, proc) in enumerate(flat_rows):
                r_excel = ri + 2
                if rtype == "room_header":
                    c = ws.cell(r_excel, 1, f"  {room}")
                    c.font = f_bold; c.fill = FILL_ROOM; c.alignment = aln_l
                    ws.merge_cells(start_row=r_excel, start_column=1,
                                   end_row=r_excel, end_column=n_cols + 1)
                    ws.row_dimensions[r_excel].height = 18
                else:
                    c = ws.cell(r_excel, 1, f"    {proc}")
                    c.alignment = aln_l; c.border = brd
                    ws.row_dimensions[r_excel].height = 18

                    for ci, col_key in enumerate(col_keys):
                        hc_map    = dist.get(col_key, {})
                        hc        = hc_map.get((room, proc), 0)
                        crp_total = crp_manager.get_total_hc(col_key[0], col_key[1])

                        c = ws.cell(r_excel, ci + 2,
                                    hc if hc > 0 else "")
                        c.alignment = aln_c; c.border = brd

                        if hc == 0:
                            c.fill = FILL_ZERO
                        elif crp_total > 0:
                            util = hc / crp_total
                            c.fill = (FILL_OVER if util >= 1.0 else
                                      FILL_WARN if util >= 0.8 else FILL_OK)
                        else:
                            c.fill = FILL_OK

            # ── Footer rows ───────────────────────────────────────────────────
            footer_labels = ["Total HC", "CRP HC", "Util %"]
            base = len(flat_rows) + 2
            for fi, lbl in enumerate(footer_labels):
                r_excel = base + fi
                c = ws.cell(r_excel, 1, lbl)
                c.font = f_bold; c.fill = FILL_TOTAL; c.alignment = aln_l
                ws.row_dimensions[r_excel].height = 18

                for ci, col_key in enumerate(col_keys):
                    hc_map    = dist.get(col_key, {})
                    total_hc  = sum(hc_map.values())
                    crp_total = crp_manager.get_total_hc(col_key[0], col_key[1])
                    util_pct  = round(total_hc / crp_total * 100) if crp_total > 0 else 0

                    if fi == 0:
                        val = total_hc
                        fill = FILL_TOTAL
                    elif fi == 1:
                        val = crp_total if crp_total > 0 else ""
                        fill = FILL_TOTAL
                    else:
                        val  = f"{util_pct}%" if crp_total > 0 else ""
                        fill = FILL_TOT_OVER if util_pct >= 100 else FILL_TOTAL

                    c = ws.cell(r_excel, ci + 2, val)
                    c.font = f_bold; c.fill = fill
                    c.alignment = aln_c; c.border = brd

            # ── Column widths ─────────────────────────────────────────────────
            ws.column_dimensions["A"].width = 30
            for ci in range(n_cols):
                col_letter = ws.cell(1, ci + 2).column_letter
                ws.column_dimensions[col_letter].width = 12

            wb.save(path)
            QMessageBox.information(self, "Export Complete",
                                    f"Saved:\n{path}")

        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))
