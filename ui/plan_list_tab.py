"""Plan List Tab — flat date-grouped view of the production plan."""
from __future__ import annotations
from collections import defaultdict
from datetime import date

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QLineEdit, QComboBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QSizePolicy,
    QMessageBox, QInputDialog, QStyledItemDelegate, QStyle,
)
from PyQt6.QtCore import Qt, QDate, QRectF, QRect
from PyQt6.QtGui import QColor, QBrush, QFont, QPainter, QFontMetrics, QPen

from data.repositories import PlanRepo, SORepo, ShiftRepo, RoomRepo

# ── Color palette (mirrors GanttCanvas) ─────────────────────────────────────
_PALETTE = [
    QColor(37,  99, 235),
    QColor( 5, 150, 105),
    QColor(124, 58, 237),
    QColor(217, 119,  6),
    QColor(220,  38,  38),
    QColor(  8, 145, 178),
    QColor(190,  18,  60),
    QColor( 22, 163,  74),
    QColor(234, 179,   8),
    QColor(168,  85, 247),
    QColor( 20, 184, 166),
]
_IO_COLOR = QColor(79, 70, 229)


def _sku_color(entity_code: str, is_io: bool = False) -> QColor:
    return _IO_COLOR if is_io else _PALETTE[hash(entity_code) % len(_PALETTE)]


_PILL_ON  = ("QPushButton{background:#DBEAFE;border:1px solid #93c5fd;"
             "color:#1d4ed8;font-weight:700;border-radius:99px;"
             "padding:2px 10px;font-size:11px;}"
             "QPushButton:hover{background:#bfdbfe;}")
_PILL_OFF = ("QPushButton{background:white;border:1px solid #DDE3ED;"
             "color:#64748B;border-radius:99px;padding:2px 10px;font-size:11px;}"
             "QPushButton:hover{background:#f1f5f9;}")

_COLS       = ["", "Room / Process", "SKU / Entity", "Qty",
               "SO / Customer", "Due", "Status", ""]
_COL_WIDTHS = [6,  0,               0,             70,
               0,                  90,  130,          64]
_COL_STRETCH = {1, 2, 4}   # these columns stretch

_HDR_H = 26
_ROW_H = 44


class _TwoLineDelegate(QStyledItemDelegate):
    """Renders a two-line cell (bold top + small grey bottom) via QPainter."""
    _F_TOP = QFont("Segoe UI", 11, QFont.Weight.Bold)
    _F_BOT = QFont("Segoe UI", 9)

    def paint(self, painter: QPainter, option, index):
        data = index.data(Qt.ItemDataRole.UserRole)
        if not data:
            super().paint(painter, option, index)
            return
        painter.save()
        if option.state & QStyle.StateFlag.State_Selected:
            painter.fillRect(option.rect, QColor("#EEF4FF"))
        r = option.rect.adjusted(8, 4, -4, -4)
        painter.setFont(self._F_TOP)
        painter.setPen(QColor(data.get("top_color", "#1E293B")))
        painter.drawText(r, Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft,
                         data.get("top", ""))
        painter.setFont(self._F_BOT)
        painter.setPen(QColor("#64748B"))
        painter.drawText(r.adjusted(0, 18, 0, 0),
                         Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft,
                         data.get("bottom", ""))
        painter.restore()


class _DueDelegate(QStyledItemDelegate):
    """Renders due-date cell with status color and days-to-due line."""
    _F_DATE = QFont("Segoe UI", 11, QFont.Weight.Bold)
    _F_DAYS = QFont("Segoe UI", 9)
    _C = {"LATE": QColor("#DC2626"), "RISK": QColor("#D97706"), "OK": QColor("#16A34A")}

    def paint(self, painter: QPainter, option, index):
        data = index.data(Qt.ItemDataRole.UserRole)
        if not data:
            super().paint(painter, option, index)
            return
        painter.save()
        if option.state & QStyle.StateFlag.State_Selected:
            painter.fillRect(option.rect, QColor("#EEF4FF"))
        r   = option.rect.adjusted(8, 4, -4, -4)
        due = data.get("due", "")
        dtd = data.get("dtd")
        st  = data.get("status", "")
        if due:
            painter.setFont(self._F_DATE)
            painter.setPen(self._C.get(st, QColor("#64748B")))
            painter.drawText(r, Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft, due)
            if dtd is not None:
                painter.setFont(self._F_DAYS)
                painter.setPen(QColor("#94a3b8"))
                sign = "+" if dtd >= 0 else ""
                painter.drawText(r.adjusted(0, 18, 0, 0),
                                 Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft,
                                 f"{sign}{dtd}d")
        else:
            painter.setFont(self._F_DATE)
            painter.setPen(QColor("#94a3b8"))
            painter.drawText(r, Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft, "—")
        painter.restore()


class _BadgeDelegate(QStyledItemDelegate):
    """Renders coloured pill badges in a cell row — zero widget objects."""
    _F    = QFont("Segoe UI", 8, QFont.Weight.Bold)
    _PH   = 5   # horizontal padding inside pill
    _PV   = 2   # vertical padding
    _GAP  = 4   # gap between pills

    def paint(self, painter: QPainter, option, index):
        badges = index.data(Qt.ItemDataRole.UserRole) or []
        painter.save()
        if option.state & QStyle.StateFlag.State_Selected:
            painter.fillRect(option.rect, QColor("#EEF4FF"))
        painter.setFont(self._F)
        fm  = QFontMetrics(self._F)
        x   = option.rect.x() + 6
        cy  = option.rect.center().y()
        bh  = fm.height() + 2 * self._PV
        for b in badges:
            text = b["text"]
            tw   = fm.horizontalAdvance(text)
            bw   = tw + 2 * self._PH
            rx   = QRectF(x, cy - bh / 2, bw, bh)
            bg   = QColor(b["bg"])
            fg   = QColor(b.get("fg", "white"))
            if b.get("outlined"):
                painter.setPen(QPen(bg, 1))
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawRoundedRect(rx, 3, 3)
                painter.setPen(bg)
            else:
                painter.setPen(Qt.PenStyle.NoPen)
                painter.setBrush(QBrush(bg))
                painter.drawRoundedRect(rx, 3, 3)
                painter.setPen(fg)
            painter.drawText(
                QRect(int(x + self._PH), int(cy - bh / 2), tw, bh),
                Qt.AlignmentFlag.AlignCenter, text)
            x += bw + self._GAP
        painter.restore()


class PlanListTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._filter_key     : str  = "ALL"
        self._collapsed_dates: set  = set()
        self._first_load     : bool = True
        self._build_ui()

    # ── UI construction ──────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._make_toolbar())
        root.addWidget(self._make_pill_bar())
        root.addWidget(self._make_kpi_strip())
        root.addWidget(self._make_scroll(), stretch=1)

    def _make_toolbar(self) -> QFrame:
        tb = QFrame()
        tb.setStyleSheet(
            "QFrame{background:white;border-bottom:1px solid #DDE3ED;}"
            "QLabel{font-size:12px;color:#64748B;}"
            "QDateEdit,QComboBox,QLineEdit{border:1px solid #DDE3ED;"
            "border-radius:5px;padding:4px 8px;font-size:12px;}"
        )
        tb.setFixedHeight(48)
        lay = QHBoxLayout(tb)
        lay.setContentsMargins(14, 0, 14, 0)
        lay.setSpacing(8)

        lay.addWidget(QLabel("From"))
        self._d_from = QDateEdit(QDate.currentDate())
        self._d_from.setCalendarPopup(True)
        self._d_from.setFixedWidth(118)
        self._d_from.dateChanged.connect(self._on_filter)
        lay.addWidget(self._d_from)

        lay.addWidget(QLabel("To"))
        self._d_to = QDateEdit(QDate.currentDate().addDays(13))
        self._d_to.setCalendarPopup(True)
        self._d_to.setFixedWidth(118)
        self._d_to.dateChanged.connect(self._on_filter)
        lay.addWidget(self._d_to)

        self._room_combo = QComboBox()
        self._room_combo.setFixedWidth(130)
        self._room_combo.currentIndexChanged.connect(self._on_filter)
        lay.addWidget(self._room_combo)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search SKU / SO / Customer…")
        self._search.setFixedWidth(200)
        self._search.textChanged.connect(self._on_filter)
        lay.addWidget(self._search)

        lay.addStretch()

        btn_gantt = QPushButton("📅  Open Gantt")
        btn_gantt.setStyleSheet(
            "QPushButton{border:1px solid #DDE3ED;border-radius:5px;"
            "padding:5px 11px;font-size:12px;background:white;color:#1E293B;}"
            "QPushButton:hover{background:#f8fafc;}")
        btn_gantt.clicked.connect(self._open_gantt)
        lay.addWidget(btn_gantt)

        btn_export = QPushButton("📥  Export")
        btn_export.setStyleSheet(
            "QPushButton{background:#2563EB;color:white;font-weight:700;"
            "border:none;border-radius:5px;padding:5px 14px;font-size:12px;}"
            "QPushButton:hover{background:#1d4ed8;}")
        btn_export.clicked.connect(self._export)
        lay.addWidget(btn_export)

        return tb

    def _make_pill_bar(self) -> QFrame:
        bar = QFrame()
        bar.setStyleSheet(
            "QFrame{background:#f8f9fc;border-bottom:1px solid #DDE3ED;}"
            "QLabel{font-size:11px;color:#64748B;font-weight:700;}")
        bar.setFixedHeight(38)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(14, 0, 14, 0)
        lay.setSpacing(6)
        lay.addWidget(QLabel("Show:"))

        self._pills: dict[str, QPushButton] = {}
        for key, label in [("ALL",    "All"),
                            ("SKU",    "SKU"),
                            ("MAT",    "Material"),
                            ("IO",     "Internal Order"),
                            ("LOCKED", "🔒 Locked"),
                            ("LATE",   "⚠️ Late")]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(24)
            btn.setChecked(key == "ALL")
            btn.setStyleSheet(_PILL_ON if key == "ALL" else _PILL_OFF)
            btn.clicked.connect(lambda _, k=key: self._set_pill(k))
            self._pills[key] = btn
            lay.addWidget(btn)

        lay.addSpacing(8)
        self._count_lbl = QLabel()
        lay.addWidget(self._count_lbl)
        lay.addStretch()
        return bar

    def _make_kpi_strip(self) -> QFrame:
        strip = QFrame()
        strip.setFixedHeight(66)
        strip.setStyleSheet("QFrame{background:#f0f2f7;border-bottom:1px solid #DDE3ED;}")
        lay = QHBoxLayout(strip)
        lay.setContentsMargins(14, 8, 14, 8)
        lay.setSpacing(10)
        self._kpi_total = self._kpi_card("Total Plans", "#94a3b8")
        self._kpi_qty   = self._kpi_card("Total Qty",   "#94a3b8")
        self._kpi_risk  = self._kpi_card("AT RISK",     "#D97706")
        self._kpi_late  = self._kpi_card("LATE",        "#DC2626")
        for c in [self._kpi_total, self._kpi_qty, self._kpi_risk, self._kpi_late]:
            lay.addWidget(c)
        return strip

    def _kpi_card(self, label: str, border: str) -> QFrame:
        card = QFrame()
        card.setStyleSheet(
            f"QFrame{{background:white;border:1px solid #DDE3ED;"
            f"border-left:3px solid {border};border-radius:6px;}}"
            "QLabel{border:none;}")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 5, 12, 5)
        lay.setSpacing(1)
        val = QLabel("—")
        val.setStyleSheet("font-size:20px;font-weight:800;color:#1E293B;")
        lbl = QLabel(label)
        lbl.setStyleSheet("font-size:10px;color:#64748B;")
        lay.addWidget(val)
        lay.addWidget(lbl)
        card._val = val
        return card

    def _make_scroll(self) -> QScrollArea:
        sa = QScrollArea()
        sa.setWidgetResizable(True)
        sa.setFrameShape(QFrame.Shape.NoFrame)
        sa.setStyleSheet("QScrollArea{background:#ECEEF3;}")

        self._list_widget = QWidget()
        self._list_widget.setStyleSheet("background:#ECEEF3;")
        self._list_layout = QVBoxLayout(self._list_widget)
        self._list_layout.setContentsMargins(14, 12, 14, 24)
        self._list_layout.setSpacing(12)
        self._list_layout.addStretch()

        sa.setWidget(self._list_widget)
        self._scroll = sa
        return sa

    # ── Refresh ──────────────────────────────────────────────────────────────

    def refresh(self):
        if self._first_load:
            self._first_load = False
            self._reload_rooms()
        self._rebuild()

    def _reload_rooms(self):
        self._room_combo.blockSignals(True)
        self._room_combo.clear()
        self._room_combo.addItem("All Rooms", "")
        for r in RoomRepo.rooms():
            self._room_combo.addItem(r, r)
        self._room_combo.blockSignals(False)

    def _on_filter(self):
        self._rebuild()

    def _set_pill(self, key: str):
        self._filter_key = key
        for k, btn in self._pills.items():
            active = (k == key)
            btn.setChecked(active)
            btn.setStyleSheet(_PILL_ON if active else _PILL_OFF)
        self._rebuild()

    def _open_gantt(self):
        mw = self.window()
        if hasattr(mw, "_on_sidebar_nav"):
            mw._on_sidebar_nav(0)

    # ── Data + rebuild ───────────────────────────────────────────────────────

    def _rebuild(self):
        d_from     = self._d_from.date().toPyDate()
        d_to       = self._d_to.date().toPyDate()
        room_filter = self._room_combo.currentData() or ""
        search     = self._search.text().strip().lower()
        today      = date.today()

        all_plans = PlanRepo.all(date_from=d_from.isoformat(),
                                 date_to=d_to.isoformat())
        sos   = {(r["so_number"], r["sku_code"], r["line_item"]): r
                 for r in SORepo.all()}
        shifts = {s["shift_no"]: s for s in ShiftRepo.all()}

        filtered = []
        for p in all_plans:
            if room_filter and p["room_code"] != room_filter:
                continue

            so  = sos.get((p["so_number"], p["sku_code"], p["line_item"]), {})
            cust = so.get("customer_name") or ""
            due_str = (so.get("committed_due_date") or so.get("due_date") or "")
            is_io   = (p["so_number"] or "").startswith("IO-")
            is_mat  = p["entity_type"] == "MATERIAL"
            is_lk   = bool(p.get("is_locked"))

            due_date = days_to = None
            status   = ""
            if due_str:
                try:
                    due_date = date.fromisoformat(due_str)
                    days_to  = (due_date - today).days
                    status   = "LATE" if days_to < 0 else ("RISK" if days_to <= 3 else "OK")
                except ValueError:
                    pass

            # entity filter
            fk = self._filter_key
            if fk == "SKU"    and (is_mat or is_io): continue
            if fk == "MAT"    and not is_mat:         continue
            if fk == "IO"     and not is_io:          continue
            if fk == "LOCKED" and not is_lk:          continue
            if fk == "LATE"   and status != "LATE":   continue

            # search
            if search:
                hay = " ".join([p.get("entity_code",""), p.get("so_number",""), cust]).lower()
                if search not in hay:
                    continue

            filtered.append(dict(plan=p, so=so, customer=cust,
                                 due_date=due_date, due_str=due_str,
                                 days_to=days_to, status=status,
                                 is_io=is_io, is_mat=is_mat, is_lk=is_lk,
                                 shift=shifts.get(p["shift_no"], {})))

        # KPIs
        total_qty = sum(r["plan"]["qty_planned"] for r in filtered)
        at_risk = sum(1 for r in filtered if r["status"] == "RISK")
        late    = sum(1 for r in filtered if r["status"] == "LATE")
        self._kpi_total._val.setText(str(len(filtered)))
        self._kpi_qty._val.setText(f"{total_qty:,}")
        self._kpi_risk._val.setText(str(at_risk))
        self._kpi_late._val.setText(str(late))

        self._count_lbl.setText(f"{len(filtered)} plans")

        # Group: date_str → shift_no → rows
        by_date: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
        for r in filtered:
            by_date[r["plan"]["plan_date"]][r["plan"]["shift_no"]].append(r)

        # Rebuild list — disable updates to suppress per-insertWidget layout recalc
        self._list_widget.setUpdatesEnabled(False)
        while self._list_layout.count() > 1:
            item = self._list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for ds in sorted(by_date.keys()):
            grp = self._make_date_group(ds, by_date[ds], shifts)
            self._list_layout.insertWidget(self._list_layout.count() - 1, grp)
        self._list_widget.setUpdatesEnabled(True)

    # ── Date group ───────────────────────────────────────────────────────────

    def _make_date_group(self, ds: str, shift_map: dict, shifts: dict) -> QWidget:
        d        = date.fromisoformat(ds)
        day_name = d.strftime("%a, %b %d, %Y")
        n_plans  = sum(len(v) for v in shift_map.values())
        n_shifts = len(shift_map)
        is_wknd  = d.weekday() >= 5
        collapsed = ds in self._collapsed_dates

        outer = QWidget()
        lay   = QVBoxLayout(outer)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Header
        hdr = QFrame()
        hdr.setFixedHeight(38)
        bg  = "#334155" if is_wknd else "#263C80"
        hdr.setStyleSheet(
            f"QFrame{{background:{bg};border-radius:6px 6px 0 0;}}"
            "QLabel{color:white;background:transparent;border:none;}")
        hdr.setCursor(Qt.CursorShape.PointingHandCursor)
        hdr_lay = QHBoxLayout(hdr)
        hdr_lay.setContentsMargins(14, 0, 14, 0)

        lbl_day  = QLabel(day_name)
        lbl_day.setStyleSheet("font-size:13px;font-weight:700;color:white;"
                              "background:transparent;border:none;")
        lbl_meta = QLabel(f"{n_plans} plans · {n_shifts} shifts")
        lbl_meta.setStyleSheet("font-size:11px;color:rgba(255,255,255,0.6);"
                               "background:transparent;border:none;")
        chev = QLabel("▼" if not collapsed else "▶")
        chev.setStyleSheet("font-size:10px;color:rgba(255,255,255,0.5);"
                           "background:transparent;border:none;")
        hdr_lay.addWidget(lbl_day)
        hdr_lay.addStretch()
        hdr_lay.addWidget(lbl_meta)
        hdr_lay.addSpacing(10)
        hdr_lay.addWidget(chev)
        lay.addWidget(hdr)

        # Content
        content = QWidget()
        content.setVisible(not collapsed)
        content.setStyleSheet(
            "QWidget{border:1px solid #DDE3ED;border-top:none;"
            "border-radius:0 0 6px 6px;background:white;}")
        c_lay = QVBoxLayout(content)
        c_lay.setContentsMargins(0, 0, 0, 0)
        c_lay.setSpacing(0)

        for sno in sorted(shift_map.keys()):
            c_lay.addWidget(self._make_shift_section(sno, shifts.get(sno, {}),
                                                     shift_map[sno]))
        lay.addWidget(content)

        def _toggle(_):
            if ds in self._collapsed_dates:
                self._collapsed_dates.discard(ds)
                content.setVisible(True)
                chev.setText("▼")
            else:
                self._collapsed_dates.add(ds)
                content.setVisible(False)
                chev.setText("▶")

        hdr.mousePressEvent = _toggle
        return outer

    # ── Shift section ────────────────────────────────────────────────────────

    def _make_shift_section(self, sno: int, shift_info: dict,
                            rows: list) -> QWidget:
        sec = QWidget()
        lay = QVBoxLayout(sec)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Sub-header
        sh = QFrame()
        sh.setFixedHeight(30)
        sh.setStyleSheet(
            "QFrame{background:#eef2fa;border-top:1px solid #DDE3ED;"
            "border-left:none;border-right:none;border-bottom:none;}"
            "QLabel{background:transparent;border:none;}")
        sh_lay = QHBoxLayout(sh)
        sh_lay.setContentsMargins(14, 0, 14, 0)

        s_name = shift_info.get("shift_name", f"Shift {sno}")
        s_t    = shift_info.get("start_time", "")
        e_t    = shift_info.get("end_time", "")
        time_s = f"  {s_t}–{e_t}" if s_t else ""

        lbl_n = QLabel(f"Shift {sno} · {s_name}")
        lbl_n.setStyleSheet("font-size:11px;font-weight:700;color:#334155;")
        lbl_t = QLabel(time_s)
        lbl_t.setStyleSheet("font-size:10px;color:#94a3b8;")
        cnt   = QLabel(str(len(rows)))
        cnt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cnt.setFixedWidth(28)
        cnt.setStyleSheet(
            "font-size:10px;color:#64748B;background:#DDE3ED;"
            "border-radius:99px;padding:0px 4px;")

        sh_lay.addWidget(lbl_n)
        sh_lay.addWidget(lbl_t)
        sh_lay.addStretch()
        sh_lay.addWidget(cnt)
        lay.addWidget(sh)
        lay.addWidget(self._make_table(rows))
        return sec

    # ── Plan table ───────────────────────────────────────────────────────────

    def _make_table(self, rows: list) -> QTableWidget:
        tbl = QTableWidget(len(rows), len(_COLS))
        tbl.setHorizontalHeaderLabels(_COLS)
        tbl.verticalHeader().setVisible(False)
        tbl.setShowGrid(False)
        tbl.setAlternatingRowColors(False)
        tbl.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        tbl.horizontalHeader().setHighlightSections(False)
        tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        tbl.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        tbl.setSizePolicy(QSizePolicy.Policy.Expanding,
                          QSizePolicy.Policy.Fixed)
        tbl.setFixedHeight(len(rows) * _ROW_H + _HDR_H)
        tbl.setStyleSheet(
            "QTableWidget{border:none;background:white;gridline-color:transparent;outline:0;}"
            "QTableWidget::item{border-bottom:1px solid #f0f2f7;padding:0;}"
            "QTableWidget::item:selected{background:#EEF4FF;color:#1E293B;}"
            "QHeaderView::section{background:#f8f9fc;color:#94a3b8;"
            "font-size:10px;font-weight:700;padding:5px 8px;border:none;"
            "border-bottom:1px solid #DDE3ED;text-transform:uppercase;}")

        hdr = tbl.horizontalHeader()
        for i, w in enumerate(_COL_WIDTHS):
            if i in _COL_STRETCH:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            else:
                hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
                tbl.setColumnWidth(i, w)

        tbl.verticalHeader().setDefaultSectionSize(_ROW_H)

        # Delegate-based rendering: zero widget objects for data columns
        _tl = _TwoLineDelegate(tbl)
        _du = _DueDelegate(tbl)
        _bd = _BadgeDelegate(tbl)
        for col in (1, 2, 4):
            tbl.setItemDelegateForColumn(col, _tl)
        tbl.setItemDelegateForColumn(5, _du)
        tbl.setItemDelegateForColumn(6, _bd)

        tbl.setUpdatesEnabled(False)
        for ri, r in enumerate(rows):
            self._fill_row(tbl, ri, r)
        tbl.setUpdatesEnabled(True)

        return tbl

    def _fill_row(self, tbl: QTableWidget, ri: int, r: dict):
        p      = r["plan"]
        is_io  = r["is_io"]
        is_mat = r["is_mat"]
        is_lk  = r["is_lk"]
        color  = _sku_color(p["entity_code"], is_io=is_io)

        # Col 0: color stripe (plain item with background)
        stripe = QTableWidgetItem()
        stripe.setBackground(QBrush(color))
        stripe.setFlags(Qt.ItemFlag.NoItemFlags)
        tbl.setItem(ri, 0, stripe)

        # Col 1: Room / Process  (rendered by _TwoLineDelegate)
        top_c = "#94a3b8" if is_lk else "#1E293B"
        it1 = QTableWidgetItem()
        it1.setData(Qt.ItemDataRole.UserRole,
                    {"top": p["room_code"], "bottom": p["process_name"],
                     "top_color": top_c})
        tbl.setItem(ri, 1, it1)

        # Col 2: SKU / Entity  (rendered by _TwoLineDelegate)
        type_str = "Material" if is_mat else ("Int. Order" if is_io else "SKU")
        it2 = QTableWidgetItem()
        it2.setData(Qt.ItemDataRole.UserRole,
                    {"top": p["entity_code"], "bottom": type_str,
                     "top_color": top_c})
        tbl.setItem(ri, 2, it2)

        # Col 3: Qty  (plain item — fast path, no widget needed)
        qty_item = QTableWidgetItem(f"{p['qty_planned']:,}")
        qty_item.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        qty_item.setTextAlignment(
            Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        if is_lk:
            qty_item.setForeground(QBrush(QColor("#94a3b8")))
        tbl.setItem(ri, 3, qty_item)

        # Col 4: SO / Customer  (rendered by _TwoLineDelegate)
        so_str = (p["so_number"] or "—")
        if p.get("line_item"):
            so_str += f"  /  {p['line_item']}"
        cust_str = r["customer"] or ("Material demand" if is_mat else "")
        so_top_c = "#94a3b8" if (is_mat and not p["so_number"]) else "#334155"
        it4 = QTableWidgetItem()
        it4.setData(Qt.ItemDataRole.UserRole,
                    {"top": so_str, "bottom": cust_str, "top_color": so_top_c})
        tbl.setItem(ri, 4, it4)

        # Col 5: Due date  (rendered by _DueDelegate)
        it5 = QTableWidgetItem()
        due_display = r["due_date"].strftime("%b %d") if r["due_date"] else ""
        it5.setData(Qt.ItemDataRole.UserRole,
                    {"due": due_display, "dtd": r["days_to"], "status": r["status"]})
        tbl.setItem(ri, 5, it5)

        # Col 6: Status badges  (rendered by _BadgeDelegate)
        badges = []
        if is_lk:                       badges.append({"text": "LOCK",  "bg": "#64748B"})
        if is_io:                       badges.append({"text": "[IO]",  "bg": "#4f46e5"})
        if is_mat:                      badges.append({"text": "MAT",   "bg": "#0891b2"})
        if p.get("is_closing_shift"):   badges.append({"text": "CLOSE", "bg": "#D97706"})
        if p.get("is_consolidated"):    badges.append({"text": "GROUP", "bg": "#D97706",
                                                        "outlined": True})
        it6 = QTableWidgetItem()
        it6.setData(Qt.ItemDataRole.UserRole, badges)
        tbl.setItem(ri, 6, it6)

        # Col 7: Actions — keep as setCellWidget (needs real button click handlers)
        tbl.setCellWidget(ri, 7, self._action_widget(p["plan_id"], is_lk))

    def _action_widget(self, plan_id: int, is_lk: bool) -> QWidget:
        w   = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.setSpacing(2)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        style = ("QPushButton{background:transparent;border:none;font-size:13px;"
                 "border-radius:4px;color:#94a3b8;}"
                 "QPushButton:hover{background:#DBEAFE;color:#2563EB;}")

        btn_m = QPushButton("💬")
        btn_m.setFixedSize(26, 26)
        btn_m.setToolTip("Add Memo")
        btn_m.setStyleSheet(style)
        btn_m.clicked.connect(lambda _, pid=plan_id: self._on_memo(pid))

        btn_l = QPushButton("🔒" if is_lk else "🔓")
        btn_l.setFixedSize(26, 26)
        btn_l.setToolTip("Unlock" if is_lk else "Lock")
        btn_l.setStyleSheet(style)
        btn_l.clicked.connect(lambda _, pid=plan_id, lk=is_lk: self._on_lock(pid, lk))

        lay.addWidget(btn_m)
        lay.addWidget(btn_l)
        return w

    # ── Row actions ──────────────────────────────────────────────────────────

    def _on_memo(self, plan_id: int):
        plan = PlanRepo.get(plan_id)
        if not plan:
            return
        text, ok = QInputDialog.getText(
            self, "Plan Memo", "Memo:", text=plan.get("memo") or "")
        if ok:
            PlanRepo.update(plan_id, {"memo": text})
            self._notify("Memo updated.")
            self._rebuild()

    def _on_lock(self, plan_id: int, currently_locked: bool):
        PlanRepo.lock(plan_id, not currently_locked)
        self._notify("Plan " + ("locked." if not currently_locked else "unlocked."))
        self._rebuild()

    def _notify(self, msg: str):
        mw = self.window()
        if hasattr(mw, "notify"):
            mw.notify(msg)

    # ── Export ───────────────────────────────────────────────────────────────

    def _export(self):
        from PyQt6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Plan List", "PlanList.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            self._export_xlsx(path)
            self._notify(f"Exported: {path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", str(e))

    def _export_xlsx(self, path: str):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        today = date.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PlanList"

        headers = ["Date", "Shift", "Room", "Process", "Entity Type", "SKU/Entity",
                   "Qty", "SO Number", "Line Item", "Customer",
                   "Due Date", "Days to Due", "Status",
                   "Locked", "IO", "MAT", "CLOSE", "GROUP", "Memo"]
        widths   = [12, 14, 14, 18, 12, 20, 8, 16, 10, 22, 12, 12, 10,
                    8, 6, 6, 8, 8, 30]

        hdr_fill = PatternFill("solid", fgColor="263C80")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 20

        d_from     = self._d_from.date().toPyDate()
        d_to       = self._d_to.date().toPyDate()
        room_filter = self._room_combo.currentData() or ""
        search     = self._search.text().strip().lower()
        fk         = self._filter_key

        all_plans = PlanRepo.all(date_from=d_from.isoformat(),
                                 date_to=d_to.isoformat())
        sos   = {(r["so_number"], r["sku_code"], r["line_item"]): r
                 for r in SORepo.all()}
        shifts = {s["shift_no"]: s for s in ShiftRepo.all()}

        status_fills = {
            "OK":   PatternFill("solid", fgColor="C8F7C5"),
            "RISK": PatternFill("solid", fgColor="FFF3CD"),
            "LATE": PatternFill("solid", fgColor="FFD6D6"),
        }
        ri = 2
        for p in all_plans:
            if room_filter and p["room_code"] != room_filter:
                continue
            so   = sos.get((p["so_number"], p["sku_code"], p["line_item"]), {})
            cust = so.get("customer_name") or ""
            due_str = so.get("committed_due_date") or so.get("due_date") or ""
            is_io  = (p["so_number"] or "").startswith("IO-")
            is_mat = p["entity_type"] == "MATERIAL"
            is_lk  = bool(p.get("is_locked"))
            days_to = None
            status  = ""
            if due_str:
                try:
                    dd = date.fromisoformat(due_str)
                    days_to = (dd - today).days
                    status  = "LATE" if days_to < 0 else ("RISK" if days_to <= 3 else "OK")
                except ValueError:
                    pass
            if fk == "SKU"    and (is_mat or is_io): continue
            if fk == "MAT"    and not is_mat:         continue
            if fk == "IO"     and not is_io:          continue
            if fk == "LOCKED" and not is_lk:          continue
            if fk == "LATE"   and status != "LATE":   continue
            if search:
                hay = " ".join([p.get("entity_code",""), p.get("so_number",""), cust]).lower()
                if search not in hay:
                    continue

            sn    = shifts.get(p["shift_no"], {})
            s_lbl = f"Shift {p['shift_no']} {sn.get('shift_name','')}"
            vals  = [
                p["plan_date"], s_lbl, p["room_code"], p["process_name"],
                p["entity_type"], p["entity_code"], p["qty_planned"],
                p["so_number"], p["line_item"], cust, due_str,
                days_to, status,
                "Y" if is_lk else "", "Y" if is_io else "",
                "Y" if is_mat else "",
                "Y" if p.get("is_closing_shift") else "",
                "Y" if p.get("is_consolidated") else "",
                p.get("memo") or "",
            ]
            fill = status_fills.get(status)
            for c, v in enumerate(vals, 1):
                cell = ws.cell(ri, c, v)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="center")
            ri += 1

        ws.freeze_panes = "A2"
        wb.save(path)
