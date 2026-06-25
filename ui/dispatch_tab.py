"""
Dispatch List Tab — production priority queue sorted by Critical Ratio.

Critical Ratio (CR) = Open Shifts until Due ÷ Shifts Needed (1-room, hc_max basis)
  CR < 1.0  → CRITICAL  (behind schedule, must start now)
  CR < 1.5  → AT RISK   (tight, watch closely)
  CR ≥ 1.5  → ON TIME
  rem_qty=0 → COMPLETE
"""
from __future__ import annotations

import math
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Set, Tuple

from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QBrush, QColor, QFont
from PyQt6.QtWidgets import (
    QAbstractItemView, QComboBox, QFileDialog, QFrame, QHBoxLayout,
    QHeaderView, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QVBoxLayout, QWidget,
)

from core.scheduler import calc_uph
from data.repositories import (
    AllocationRepo, CalendarRepo, ProcessRoutingRepo,
    RoomRepo, ShiftRepo, SKURepo, SORepo,
)

# ── Thresholds ────────────────────────────────────────────────────────────────
CR_CRITICAL = 1.0
CR_AT_RISK  = 1.5

STATUS_CRITICAL  = "CRITICAL"
STATUS_AT_RISK   = "AT RISK"
STATUS_ON_TIME   = "ON TIME"
STATUS_COMPLETE  = "COMPLETE"
STATUS_NO_ROUTE  = "NO ROUTING"

_STATUS_COLORS = {
    STATUS_CRITICAL:  ("#FFEBEE", "#B71C1C"),
    STATUS_AT_RISK:   ("#FFF8E1", "#E65100"),
    STATUS_ON_TIME:   ("#E8F5E9", "#1B5E20"),
    STATUS_COMPLETE:  ("#F5F5F5", "#757575"),
    STATUS_NO_ROUTE:  ("#F5F5F5", "#9E9E9E"),
}

_COLS = [
    "Rank", "SO", "Customer", "SKU", "Line",
    "Process (Seq)", "Rem. Qty", "Shifts Needed",
    "Open Shifts", "True Slack", "CR",
    "Req. Due", "Committed", "D-Day", "Status",
]


def _shift_hours(shift: Dict) -> float:
    from datetime import datetime as _dt
    fmt = "%H:%M"
    t0 = _dt.strptime(shift["start_time"], fmt)
    t1 = _dt.strptime(shift["end_time"], fmt)
    h = (t1 - t0).seconds / 3600
    return h if h > 0 else 24 + h


# ── Main Tab ──────────────────────────────────────────────────────────────────

class DispatchListTab(QWidget):
    """Weekly production priority queue — sorted by Critical Ratio."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self._all_rows: List[Dict] = []
        self._build_ui()

        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(30_000)

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(6)

        # ── Toolbar ──
        bar = QHBoxLayout()
        bar.setSpacing(8)

        bar.addWidget(QLabel("Search:"))
        self._search = QLineEdit()
        self._search.setPlaceholderText("SO / SKU / Customer…")
        self._search.setFixedWidth(180)
        self._search.textChanged.connect(self._apply_filter)
        bar.addWidget(self._search)

        bar.addWidget(QLabel("Process:"))
        self._proc_combo = QComboBox()
        self._proc_combo.setFixedWidth(160)
        self._proc_combo.addItem("ALL")
        self._proc_combo.currentTextChanged.connect(self._apply_filter)
        bar.addWidget(self._proc_combo)

        bar.addWidget(QLabel("Status:"))
        self._status_combo = QComboBox()
        self._status_combo.addItems(
            ["ALL", STATUS_CRITICAL, STATUS_AT_RISK, STATUS_ON_TIME,
             STATUS_COMPLETE, STATUS_NO_ROUTE])
        self._status_combo.currentTextChanged.connect(self._apply_filter)
        bar.addWidget(self._status_combo)

        bar.addStretch()

        self._summary_lbl = QLabel()
        self._summary_lbl.setStyleSheet("color:#555; font-size:11px;")
        bar.addWidget(self._summary_lbl)

        btn_refresh = QPushButton("🔄 Refresh")
        btn_refresh.clicked.connect(self.refresh)
        bar.addWidget(btn_refresh)

        btn_export = QPushButton("📥 Export")
        btn_export.clicked.connect(self._export)
        bar.addWidget(btn_export)

        layout.addLayout(bar)

        # ── Legend ──
        legend = QHBoxLayout()
        legend.setSpacing(16)
        for status, (bg, fg) in _STATUS_COLORS.items():
            dot = QLabel(f"● {status}")
            dot.setStyleSheet(
                f"color:{fg}; font-size:10px; font-weight:600;")
            legend.addWidget(dot)
        legend.addStretch()
        cap_note = QLabel("Shifts Needed = 1-room basis (hc_max, best UPH)")
        cap_note.setStyleSheet("color:#888; font-size:10px; font-style:italic;")
        legend.addWidget(cap_note)
        layout.addLayout(legend)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("background:#dde3ed; border:none;")
        sep.setFixedHeight(1)
        layout.addWidget(sep)

        # ── Table ──
        self._table = QTableWidget()
        self._table.setColumnCount(len(_COLS))
        self._table.setHorizontalHeaderLabels(_COLS)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(False)
        self._table.setSortingEnabled(True)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalHeader().setDefaultSectionSize(26)
        self._table.setStyleSheet(
            "QTableWidget { border:none; font-size:11px; }"
            "QHeaderView::section { background:#f0f2f7; font-weight:600;"
            " padding:4px; border:none; border-bottom:1px solid #dde3ed; }")
        layout.addWidget(self._table, stretch=1)

        self.refresh()

    # ── Data ─────────────────────────────────────────────────────────────────

    def refresh(self):
        self._all_rows = self._compute_rows()
        # Rebuild process combo (preserve selection if still valid)
        current_proc = self._proc_combo.currentText()
        procs = sorted({r["process_name"] for r in self._all_rows
                        if r["process_name"] != "(no routing)"})
        self._proc_combo.blockSignals(True)
        self._proc_combo.clear()
        self._proc_combo.addItem("ALL")
        self._proc_combo.addItems(procs)
        if current_proc in procs:
            self._proc_combo.setCurrentText(current_proc)
        self._proc_combo.blockSignals(False)
        self._apply_filter()

    def _compute_rows(self) -> List[Dict]:
        today = date.today()
        today_str = today.strftime("%Y-%m-%d")

        # ── Load masters ─────────────────────────────────────────────────────
        sos = [s for s in SORepo.all() if s["status"] == "OPEN"]
        if not sos:
            return []

        all_rooms = RoomRepo.all()
        shifts = ShiftRepo.all()
        sku_map: Dict[str, Dict] = {s["sku_code"]: s for s in SKURepo.all()}

        # avg shift hours across all defined shifts
        avg_shift_h = (sum(_shift_hours(s) for s in shifts) / len(shifts)
                       if shifts else 12.0)
        n_shifts = len(shifts) if shifts else 2

        # process_name → list of room_proc dicts
        process_rooms: Dict[str, List[Dict]] = {}
        for rp in all_rooms:
            process_rooms.setdefault(rp["process_name"], []).append(rp)

        # room_code → [process_name, ...]
        room_to_procs: Dict[str, List[str]] = {}
        for rp in all_rooms:
            room_to_procs.setdefault(rp["room_code"], []).append(
                rp["process_name"])

        # ── Open calendar slots ───────────────────────────────────────────────
        max_due = max(
            (s.get("committed_due_date") or s["due_date"] for s in sos),
            default=today_str)

        raw_slots = CalendarRepo.get_open_slots(today_str, max_due)
        # {(date_str, room_code, shift_no)} — already is_open=1
        open_set: Set[Tuple] = {
            (s["cal_date"], s["room_code"], s["shift_no"])
            for s in raw_slots if not s.get("is_hold", 0)
        }
        # Which (date, shift) pairs appear at all in calendar (any room)
        calendared: Set[Tuple] = {
            (s["cal_date"], s["shift_no"]) for s in raw_slots}

        # ── Pre-compute cumulative open shifts per process up to max_due ─────
        # proc_cum[proc_name][date_str] = total open shifts from today to date_str
        proc_cum: Dict[str, Dict[str, int]] = {}
        all_procs = list(process_rooms.keys())
        # build: proc → {room_codes}
        proc_room_set: Dict[str, Set[str]] = {
            p: {rp["room_code"] for rp in rps}
            for p, rps in process_rooms.items()
        }

        d = today
        until = datetime.strptime(max_due, "%Y-%m-%d").date()
        running: Dict[str, int] = {p: 0 for p in all_procs}

        while d <= until:
            ds = d.strftime("%Y-%m-%d")
            is_weekday = d.weekday() < 5

            for proc_name in all_procs:
                rooms = proc_room_set[proc_name]
                added = 0
                for sno_idx in range(1, n_shifts + 1):
                    if (ds, sno_idx) in calendared:
                        # calendar data exists → check if any room is open
                        if any((ds, rc, sno_idx) in open_set for rc in rooms):
                            added += 1
                    else:
                        # no calendar entry → weekday fallback
                        if is_weekday:
                            added += 1
                running[proc_name] += added

            for proc_name in all_procs:
                proc_cum.setdefault(proc_name, {})[ds] = running[proc_name]

            d += timedelta(days=1)

        # ── Build rows ────────────────────────────────────────────────────────
        rows: List[Dict] = []

        for so in sos:
            routing = ProcessRoutingRepo.for_entity("SKU", so["sku_code"])
            prod_needed = AllocationRepo.production_needed(
                so["so_number"], so["sku_code"], so["line_item"])

            eff_due_str = so.get("committed_due_date") or so["due_date"]
            req_due_str = so["due_date"]
            committed_str = so.get("committed_due_date") or ""

            eff_due = datetime.strptime(eff_due_str, "%Y-%m-%d").date()
            days_to_due = (eff_due - today).days

            sku = sku_map.get(so["sku_code"], {})
            uom = int(sku.get("uom") or 1)

            if not routing:
                rows.append(_make_row(
                    so, "(no routing)", 0, prod_needed,
                    None, None, None, None,
                    req_due_str, committed_str, days_to_due, STATUS_NO_ROUTE))
                continue

            for step in routing:
                proc_name = step["process_name"]
                proc_seq  = step["process_seq"]

                # Best single-room capacity per shift (inner units, hc_max)
                best_cap = 0.0
                for rp in process_rooms.get(proc_name, []):
                    hc = (int(rp.get("hc_fixed") or 1)
                          if rp.get("process_type") == "AUTO"
                          else int(rp.get("hc_max") or rp.get("hc_min") or 1))
                    uph = calc_uph(rp, hc)
                    cap = uph * avg_shift_h
                    if cap > best_cap:
                        best_cap = cap

                rem_inner = prod_needed * uom

                if rem_inner == 0:
                    shifts_needed = 0
                    open_shifts   = proc_cum.get(proc_name, {}).get(eff_due_str, 0)
                    true_slack    = open_shifts
                    cr            = 999.0
                    status        = STATUS_COMPLETE
                elif best_cap <= 0:
                    shifts_needed = None
                    open_shifts   = None
                    true_slack    = None
                    cr            = None
                    status        = STATUS_NO_ROUTE
                else:
                    shifts_needed = math.ceil(rem_inner / best_cap)
                    open_shifts   = proc_cum.get(proc_name, {}).get(eff_due_str, 0)
                    true_slack    = open_shifts - shifts_needed
                    cr            = open_shifts / shifts_needed
                    if cr < CR_CRITICAL:
                        status = STATUS_CRITICAL
                    elif cr < CR_AT_RISK:
                        status = STATUS_AT_RISK
                    else:
                        status = STATUS_ON_TIME

                rows.append(_make_row(
                    so, proc_name, proc_seq, prod_needed,
                    shifts_needed, open_shifts, true_slack, cr,
                    req_due_str, committed_str, days_to_due, status))

        # ── Sort: CR asc (COMPLETE/NO_ROUTE last) ────────────────────────────
        def _sort_key(r):
            if r["status"] == STATUS_COMPLETE:
                return (2, 999.0, r["days_to_due"])
            if r["status"] == STATUS_NO_ROUTE:
                return (3, 999.0, r["days_to_due"])
            return (0, r["cr"] if r["cr"] is not None else 999.0, r["days_to_due"])

        rows.sort(key=_sort_key)

        # Assign rank (active rows only)
        rank = 1
        for r in rows:
            if r["status"] in (STATUS_CRITICAL, STATUS_AT_RISK, STATUS_ON_TIME):
                r["rank"] = rank
                rank += 1
            else:
                r["rank"] = "—"

        return rows

    # ── Filter & Render ───────────────────────────────────────────────────────

    def _apply_filter(self):
        search = self._search.text().strip().lower()
        proc_f  = self._proc_combo.currentText()
        stat_f  = self._status_combo.currentText()

        filtered = []
        for r in self._all_rows:
            if search and not any(
                search in str(r.get(k, "")).lower()
                for k in ("so_number", "sku_code", "line_item", "customer")):
                continue
            if proc_f != "ALL" and r["process_name"] != proc_f:
                continue
            if stat_f != "ALL" and r["status"] != stat_f:
                continue
            filtered.append(r)

        self._populate_table(filtered)

        # Summary
        n_crit = sum(1 for r in filtered if r["status"] == STATUS_CRITICAL)
        n_risk = sum(1 for r in filtered if r["status"] == STATUS_AT_RISK)
        n_ok   = sum(1 for r in filtered if r["status"] == STATUS_ON_TIME)
        self._summary_lbl.setText(
            f"Showing {len(filtered)} rows  |  "
            f"🔴 Critical: {n_crit}  🟠 At Risk: {n_risk}  🟢 On Time: {n_ok}")

    def _populate_table(self, rows: List[Dict]):
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._table.setSortingEnabled(False)
        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(rows))

        bold = QFont()
        bold.setBold(True)

        for ri, r in enumerate(rows):
            bg, fg = _STATUS_COLORS.get(r["status"], ("#fff", "#000"))
            bg_c = QBrush(QColor(bg))
            fg_c = QBrush(QColor(fg))

            cr_val  = r["cr"]
            cr_str  = f"{cr_val:.2f}" if cr_val is not None and cr_val < 900 else "—"
            tsl_val = r["true_slack"]
            tsl_str = (f"+{tsl_val}" if tsl_val is not None and tsl_val >= 0
                       else str(tsl_val) if tsl_val is not None else "—")
            sn_str  = str(r["shifts_needed"]) if r["shifts_needed"] is not None else "—"
            os_str  = str(r["open_shifts"])   if r["open_shifts"]   is not None else "—"
            d_day   = r["days_to_due"]
            d_str   = (f"+{d_day}d" if d_day >= 0 else f"{d_day}d")

            values = [
                str(r["rank"]),
                r["so_number"],
                r["customer"],
                r["sku_code"],
                r["line_item"],
                f"[{r['process_seq']}] {r['process_name']}",
                str(r["rem_qty"]),
                sn_str,
                os_str,
                tsl_str,
                cr_str,
                r["req_due"],
                r["committed_due"] or "—",
                d_str,
                r["status"],
            ]

            for ci, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setBackground(bg_c)
                item.setForeground(fg_c)
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignCenter
                    if ci not in (1, 2, 3, 5)
                    else Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                # Bold the rank + status columns for CRITICAL rows
                if r["status"] == STATUS_CRITICAL and ci in (0, 14):
                    item.setFont(bold)
                self._table.setItem(ri, ci, item)

        self._table.setUpdatesEnabled(True)
        hdr.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.setSortingEnabled(True)
        self._table.resizeColumnsToContents()

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Export", "openpyxl not installed.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Dispatch List",
            f"DispatchList_{date.today()}.xlsx",
            "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dispatch List"
        ws.append(_COLS)

        # header style
        hdr_fill = PatternFill("solid", fgColor="2F5FD6")
        hdr_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        _hex = {"CRITICAL": "FFCDD2", "AT RISK": "FFF9C4",
                "ON TIME":  "C8E6C9", "COMPLETE": "F5F5F5",
                "NO ROUTING": "EEEEEE"}

        for r in self._all_rows:
            bg = _hex.get(r["status"], "FFFFFF")
            cr_val = r["cr"]
            cr_str = f"{cr_val:.2f}" if cr_val is not None and cr_val < 900 else "—"
            tsl = r["true_slack"]
            tsl_str = (f"+{tsl}" if tsl is not None and tsl >= 0
                       else str(tsl) if tsl is not None else "—")
            d_day = r["days_to_due"]
            d_str = f"+{d_day}d" if d_day >= 0 else f"{d_day}d"

            row_data = [
                str(r["rank"]),
                r["so_number"], r["customer"], r["sku_code"], r["line_item"],
                f"[{r['process_seq']}] {r['process_name']}",
                r["rem_qty"],
                r["shifts_needed"] if r["shifts_needed"] is not None else "",
                r["open_shifts"]   if r["open_shifts"]   is not None else "",
                tsl_str, cr_str,
                r["req_due"], r["committed_due"] or "", d_str, r["status"],
            ]
            ws.append(row_data)
            fill = PatternFill("solid", fgColor=bg)
            for cell in ws[ws.max_row]:
                cell.fill = fill

        wb.save(path)
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Export", f"Saved: {path}")


# ── Helper ────────────────────────────────────────────────────────────────────

def _make_row(so: Dict, proc_name: str, proc_seq: int, rem_qty: int,
              shifts_needed, open_shifts, true_slack, cr,
              req_due: str, committed_due: str, days_to_due: int,
              status: str) -> Dict:
    return {
        "rank":         "—",
        "so_number":    so["so_number"],
        "customer":     so.get("customer_name", ""),
        "sku_code":     so["sku_code"],
        "line_item":    so["line_item"],
        "process_name": proc_name,
        "process_seq":  proc_seq,
        "rem_qty":      rem_qty,
        "shifts_needed": shifts_needed,
        "open_shifts":  open_shifts,
        "true_slack":   true_slack,
        "cr":           cr,
        "req_due":      req_due,
        "committed_due": committed_due,
        "days_to_due":  days_to_due,
        "status":       status,
    }
